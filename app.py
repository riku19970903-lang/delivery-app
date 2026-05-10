import calendar
import csv
from contextlib import contextmanager
from datetime import date, datetime, timedelta, timezone
from email.message import EmailMessage
from hashlib import pbkdf2_hmac, sha256
import hmac
import io
import json
import logging
import secrets
import smtplib
import os
from pathlib import Path
import re
import sqlite3
from threading import Lock, Thread
from time import monotonic
from typing import List, Optional
import unicodedata
from urllib.parse import quote, unquote
from urllib.error import HTTPError, URLError
from urllib.request import Request as UrlRequest, urlopen
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from itsdangerous import BadSignature, URLSafeSerializer


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "delivery_ops.db"
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
USE_POSTGRES = bool(DATABASE_URL)
DB_CONNECT_TIMEOUT = int(os.getenv("DB_CONNECT_TIMEOUT", "3"))
DB_INIT_RETRY_SECONDS = int(os.getenv("DB_INIT_RETRY_SECONDS", "30"))
SCHEMA_VERSION = "2026-05-10-member-results-v1"
DB_INIT_ON_STARTUP = os.getenv("DB_INIT_ON_STARTUP", "0" if USE_POSTGRES else "1").lower() in {"1", "true", "yes", "on"}
APP_NAME = "SPARKLE DRIVE"
SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
SUPABASE_STORAGE_KEY = (
    os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    or os.getenv("SUPABASE_SERVICE_KEY")
    or os.getenv("SUPABASE_STORAGE_KEY")
    or os.getenv("SUPABASE_ANON_KEY")
    or ""
)
SUPABASE_STORAGE_BUCKET = os.getenv("SUPABASE_STORAGE_BUCKET", "inspection-sheets")
SUPABASE_STORAGE_PUBLIC_BASE = os.getenv("SUPABASE_STORAGE_PUBLIC_URL_BASE", "").rstrip("/")
STORAGE_RETENTION_DAYS = int(os.getenv("STORAGE_RETENTION_DAYS", "62"))
APP_BASE_URL = os.getenv("APP_BASE_URL", "").rstrip("/")
SMTP_HOST = os.getenv("SMTP_HOST", "").strip()
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "").strip()
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_FROM = os.getenv("SMTP_FROM", SMTP_USERNAME or "no-reply@sparkle-drive.local").strip()
SMTP_USE_TLS = os.getenv("SMTP_USE_TLS", "1").lower() in {"1", "true", "yes", "on"}
PASSWORD_RESET_EXPIRY_MINUTES = int(os.getenv("PASSWORD_RESET_EXPIRY_MINUTES", "60"))
try:
    APP_TZ = ZoneInfo(os.getenv("APP_TIMEZONE", "Asia/Tokyo"))
except ZoneInfoNotFoundError:
    APP_TZ = timezone(timedelta(hours=9))
UPLOAD_DIR = BASE_DIR / "uploads" / "inspection_sheets"
SLIP_UPLOAD_DIR = BASE_DIR / "uploads" / "inspection_slips"
SUPPORT_UPLOAD_DIR = BASE_DIR / "uploads" / "support_requests"
RECEIPT_UPLOAD_DIR = BASE_DIR / "uploads" / "receipts"
SECRET_KEY = (
    os.getenv("SECRET_KEY")
    or os.getenv("APP_SECRET_KEY")
    or os.getenv("SESSION_SECRET")
    or "local-dev-session-secret-change-me"
)
SESSION_COOKIE_SECURE = os.getenv("SESSION_COOKIE_SECURE", "1" if USE_POSTGRES else "0").lower() in {"1", "true", "yes", "on"}
PASSWORD_ITERATIONS = int(os.getenv("PASSWORD_ITERATIONS", "200000"))
DEFAULT_ADMIN_PASSWORD = os.getenv("DEFAULT_ADMIN_PASSWORD", "admin123")
DEFAULT_MEMBER_PASSWORD = os.getenv("DEFAULT_MEMBER_PASSWORD", "member123")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
SLIP_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
SUPPORT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
RECEIPT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title=APP_NAME)
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
app.mount("/uploads", StaticFiles(directory=BASE_DIR / "uploads"), name="uploads")
templates = Jinja2Templates(directory=BASE_DIR / "templates")
serializer = URLSafeSerializer(SECRET_KEY, salt="delivery-ops-session")
logger = logging.getLogger("delivery_app")
DB_INIT_LOCK = Lock()
DB_INIT_DONE = False
DB_INIT_ERROR = None
DB_INIT_LAST_ATTEMPT = 0.0
FEATURE_CACHE = {}
FEATURE_CACHE_TTL_SECONDS = int(os.getenv("FEATURE_CACHE_TTL_SECONDS", "30"))
SCHEMA_REQUIRED_COLUMNS = {
    "companies": ("id", "name", "status", "next_renewal_date"),
    "users": ("id", "company_id", "username", "password_hash", "role", "notification_email", "member_notes"),
    "deliveries": ("id", "company_id", "user_id", "work_date", "completed", "inspection_sheet_path", "is_deleted", "deleted_at"),
    "work_logs": ("id", "company_id", "user_id", "work_date", "log_type", "start_odometer", "end_odometer", "distance_km", "is_deleted", "deleted_at"),
    "work_log_sessions": ("id", "company_id", "user_id", "work_date", "session_no", "status", "start_odometer", "end_odometer", "inspection_image_path", "is_deleted", "deleted_at"),
    "work_log_corrections": ("id", "company_id", "user_id", "work_date", "before_data", "after_data", "actor_id"),
    "work_day_deletions": ("id", "company_id", "user_id", "work_date", "deleted_at"),
    "vehicles": ("id", "company_id", "name", "active", "last_odometer", "inspection_due", "vehicle_status", "depot_id", "primary_user_id"),
    "vehicle_expenses": ("id", "company_id", "vehicle_id", "expense_date", "category", "amount"),
    "company_expenses": ("id", "company_id", "expense_date", "category", "amount"),
    "depot_rate_settings": ("id", "company_id", "depot_id", "base_unit", "effective_start"),
    "inspection_sheets": ("id", "company_id", "user_id", "file_path", "storage_path", "retention_until"),
    "inspection_slips": ("id", "company_id", "user_id", "file_path", "storage_path", "retention_until"),
    "company_features": ("company_id", "feature_key", "enabled"),
    "audit_logs": ("id", "company_id", "action", "table_name", "created_at"),
}
SHIFT_ALLOWED_STATUSES = {"出勤", "1便", "2便", "3便", "休み"}
SHIFT_IMPORT_HEADERS = ["日付", "名前", "出勤区分", "配達エリア", "便", "備考"]
SHIFT_IMPORT_LIMIT = 500
FULLWIDTH_DIGITS = str.maketrans("０１２３４５６７８９", "0123456789")
PLATFORM_COMPANY_ID = 1
COMPANY_STATUSES = ("利用中", "停止中", "お試し")
SUPPORT_TYPES = ("不具合", "改善要望", "使い方相談")
SUPPORT_URGENCIES = ("低", "中", "高", "緊急")
SUPPORT_STATUSES = ("未対応", "対応中", "完了", "保留")
VEHICLE_STATUSES = ("使用中", "修理中", "予備", "廃車予定")
VEHICLE_EXPENSE_CATEGORIES = ("オイル交換", "タイヤ", "車検", "修理", "保険", "駐車場", "ガソリン", "洗車", "部品", "その他")
COMPANY_EXPENSE_CATEGORIES = ("車両費", "人件費", "駐車場", "保険", "備品", "通信費", "交際費", "福利厚生", "その他")
FEATURE_CATALOG = [
    {"key": "basic", "label": "基本管理", "price": 3000, "note": "課金予定"},
    {"key": "members", "label": "メンバー管理", "price": 0, "note": "テスト中"},
    {"key": "deliveries", "label": "配達個数管理", "price": 2000, "note": "課金予定"},
    {"key": "rewards", "label": "給料・報酬管理", "price": 3000, "note": "課金予定"},
    {"key": "shifts", "label": "シフト管理10人まで", "price": 2000, "note": "課金予定"},
    {"key": "shifts_20", "label": "シフト管理20人まで", "price": 4000, "note": "課金予定"},
    {"key": "inspection_ocr", "label": "点検表OCR", "price": 3000, "note": "課金予定"},
    {"key": "safety", "label": "安全記録", "price": 1000, "note": "テスト中"},
    {"key": "vehicle", "label": "車両管理", "price": 1000, "note": "テスト中"},
    {"key": "issues", "label": "不具合管理", "price": 1000, "note": "テスト中"},
    {"key": "holidays", "label": "休み希望管理", "price": 1000, "note": "テスト中"},
    {"key": "multi_depot", "label": "複数拠点管理", "price": 2000, "note": "テスト中"},
]
FEATURE_KEYS = {item["key"] for item in FEATURE_CATALOG}
FEATURE_ALIASES = {"shifts": ("shifts", "shifts_20"), "inspection": ("inspection_ocr",)}


def app_now():
    return datetime.now(APP_TZ)


def app_today():
    return app_now().date()


def seed_company_features(conn, company_id: int, enabled: int = 1):
    now = datetime.now().isoformat(timespec="seconds")
    for feature in FEATURE_CATALOG:
        conn.execute(
            "INSERT OR IGNORE INTO company_features(company_id, feature_key, enabled, updated_at) VALUES (?,?,?,?)",
            (company_id, feature["key"], enabled, now),
        )


def seed_feature_prices(conn):
    now = datetime.now().isoformat(timespec="seconds")
    for feature in FEATURE_CATALOG:
        conn.execute(
            """INSERT OR IGNORE INTO feature_prices(feature_key, label, monthly_price, note, active, updated_at)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (feature["key"], feature["label"], feature["price"], feature["note"], 1, now),
        )


if USE_POSTGRES:
    import psycopg2
    from psycopg2.extras import RealDictCursor

    DB_INTEGRITY_ERROR = psycopg2.IntegrityError
else:
    psycopg2 = None
    RealDictCursor = None
    DB_INTEGRITY_ERROR = sqlite3.IntegrityError


IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


class DbCursor:
    def __init__(self, cursor, lastrowid=None):
        self.cursor = cursor
        self.lastrowid = lastrowid if lastrowid is not None else getattr(cursor, "lastrowid", None)

    def fetchone(self):
        return self.cursor.fetchone()

    def fetchall(self):
        return self.cursor.fetchall()


class DbConnection:
    def __init__(self, conn, use_postgres: bool):
        self.conn = conn
        self.use_postgres = use_postgres

    def execute(self, sql: str, params=()):
        if not self.use_postgres:
            return self.conn.execute(sql, params or ())

        return_id = _needs_returning_id(sql)
        translated_sql = _postgres_sql(sql, return_id=return_id)
        cur = self.conn.cursor()
        cur.execute(translated_sql, params or ())
        lastrowid = None
        if return_id:
            row = cur.fetchone()
            if row:
                lastrowid = row["id"]
        return DbCursor(cur, lastrowid)

    def executescript(self, sql: str):
        if not self.use_postgres:
            return self.conn.executescript(sql)
        result = None
        for statement in _split_sql_script(sql):
            result = self.execute(statement)
        return result

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        self.conn.close()


def _postgres_sql(sql: str, return_id: bool = False) -> str:
    translated = re.sub(
        r"INTEGER\s+PRIMARY\s+KEY\s+AUTOINCREMENT",
        "SERIAL PRIMARY KEY",
        sql,
        flags=re.IGNORECASE,
    )
    translated = re.sub(r"\bAUTOINCREMENT\b", "", translated, flags=re.IGNORECASE)

    if re.match(r"^\s*INSERT\s+OR\s+IGNORE\s+INTO\b", translated, flags=re.IGNORECASE):
        translated = re.sub(
            r"^\s*INSERT\s+OR\s+IGNORE\s+INTO\b",
            "INSERT INTO",
            translated,
            count=1,
            flags=re.IGNORECASE,
        )
        if "ON CONFLICT" not in translated.upper():
            translated = _append_sql_clause(translated, "ON CONFLICT DO NOTHING")

    translated = translated.replace("?", "%s")
    if return_id:
        translated = _append_sql_clause(translated, "RETURNING id")
    return translated


def _split_sql_script(sql: str):
    return [statement.strip() for statement in sql.split(";") if statement.strip()]


def _append_sql_clause(sql: str, clause: str) -> str:
    stripped = sql.rstrip()
    has_semicolon = stripped.endswith(";")
    if has_semicolon:
        stripped = stripped[:-1].rstrip()
    stripped = f"{stripped} {clause}"
    return f"{stripped};" if has_semicolon else stripped


def _needs_returning_id(sql: str) -> bool:
    return bool(
        re.match(r"^\s*INSERT\s+INTO\s+(users|vehicle_issues|companies|vehicles|vehicle_expenses|company_expenses|depot_rate_settings|work_logs|work_log_sessions|work_log_corrections|work_day_deletions|audit_logs)\b", sql, flags=re.IGNORECASE)
        and "RETURNING" not in sql.upper()
        and "ON CONFLICT" not in sql.upper()
    )


def _quote_identifier(name: str) -> str:
    if not IDENTIFIER_RE.match(name):
        raise ValueError(f"Unsafe SQL identifier: {name}")
    return f'"{name}"' if USE_POSTGRES else name


@contextmanager
def db():
    if USE_POSTGRES:
        connect_kwargs = {
            "cursor_factory": RealDictCursor,
            "connect_timeout": DB_CONNECT_TIMEOUT,
        }
        if "sslmode=" not in DATABASE_URL.lower():
            connect_kwargs["sslmode"] = "require"
        conn = psycopg2.connect(DATABASE_URL, **connect_kwargs)
    else:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row

    wrapped = DbConnection(conn, USE_POSTGRES)
    try:
        yield wrapped
        wrapped.commit()
    except Exception:
        wrapped.rollback()
        raise
    finally:
        wrapped.close()


def ensure_db_initialized():
    if DB_INIT_DONE:
        return
    if USE_POSTGRES and DB_INIT_ERROR and monotonic() - DB_INIT_LAST_ATTEMPT < DB_INIT_RETRY_SECONDS:
        raise HTTPException(status_code=503, detail="DB初期化に失敗しています。少し時間をおいて再読み込みしてください。")
    init_db()


def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    digest = pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        bytes.fromhex(salt),
        PASSWORD_ITERATIONS,
    ).hex()
    return f"pbkdf2_sha256${PASSWORD_ITERATIONS}${salt}${digest}"


def verify_password(password: str, stored_hash: str) -> bool:
    stored_hash = stored_hash or ""
    if stored_hash.startswith("pbkdf2_sha256$"):
        parts = stored_hash.split("$")
        if len(parts) != 4:
            return False
        _, iterations, salt, expected = parts
        try:
            digest = pbkdf2_hmac(
                "sha256",
                password.encode("utf-8"),
                bytes.fromhex(salt),
                int(iterations),
            ).hex()
        except (TypeError, ValueError):
            return False
        return hmac.compare_digest(digest, expected)
    legacy = sha256(password.encode("utf-8")).hexdigest()
    return hmac.compare_digest(legacy, stored_hash)


def needs_password_rehash(stored_hash: str) -> bool:
    return not (stored_hash or "").startswith("pbkdf2_sha256$")


def safe_upload_name(original_name: str):
    suffix = Path(original_name or "").suffix.lower()
    if suffix not in {".jpg", ".jpeg", ".png", ".webp", ".gif", ".heic", ".heif"}:
        suffix = ".jpg"
    return suffix


def schema_table_columns(conn, table: str):
    if USE_POSTGRES:
        rows = conn.execute(
            """SELECT column_name FROM information_schema.columns
               WHERE table_schema='public' AND table_name=?""",
            (table,),
        ).fetchall()
        return {row["column_name"] for row in rows}
    rows = conn.execute(f"PRAGMA table_info({_quote_identifier(table)})").fetchall()
    return {row["name"] for row in rows}


def current_schema_shape_ok():
    try:
        with db() as conn:
            for table, columns in SCHEMA_REQUIRED_COLUMNS.items():
                existing = schema_table_columns(conn, table)
                if not existing or not set(columns).issubset(existing):
                    return False
        return True
    except Exception:
        return False


def mark_schema_initialized():
    now = datetime.now().isoformat(timespec="seconds")
    with db() as conn:
        conn.execute(
            """CREATE TABLE IF NOT EXISTS app_metadata (
                meta_key TEXT PRIMARY KEY,
                meta_value TEXT DEFAULT '',
                updated_at TEXT DEFAULT ''
            )"""
        )
        conn.execute(
            """INSERT INTO app_metadata(meta_key, meta_value, updated_at)
               VALUES (?, ?, ?)
               ON CONFLICT(meta_key) DO UPDATE SET meta_value=excluded.meta_value, updated_at=excluded.updated_at""",
            ("schema_version", SCHEMA_VERSION, now),
        )
        conn.commit()


def supabase_storage_enabled():
    return bool(SUPABASE_URL and SUPABASE_STORAGE_KEY and SUPABASE_STORAGE_BUCKET)


def _storage_path(folder: str, filename: str):
    clean_folder = folder.strip("/").replace("\\", "/")
    clean_name = Path(filename).name
    return f"{clean_folder}/{clean_name}"


def _encoded_storage_path(storage_path: str):
    return "/".join(quote(part, safe="") for part in storage_path.split("/"))


def upload_to_supabase_storage(folder: str, filename: str, data: bytes, content_type: str):
    if not supabase_storage_enabled():
        return ""
    storage_path = _storage_path(folder, filename)
    encoded_path = _encoded_storage_path(storage_path)
    url = f"{SUPABASE_URL}/storage/v1/object/{quote(SUPABASE_STORAGE_BUCKET, safe='')}/{encoded_path}"
    request = UrlRequest(
        url,
        data=data,
        method="POST",
        headers={
            "apikey": SUPABASE_STORAGE_KEY,
            "Authorization": f"Bearer {SUPABASE_STORAGE_KEY}",
            "Content-Type": content_type or "application/octet-stream",
            "Cache-Control": "3600",
            "x-upsert": "true",
        },
    )
    try:
        with urlopen(request, timeout=DB_CONNECT_TIMEOUT) as response:
            if response.status >= 400:
                logger.warning("Supabase Storage upload failed: %s", response.status)
                return ""
    except (HTTPError, URLError, TimeoutError, OSError):
        logger.exception("Supabase Storage upload failed; falling back to local uploads.")
        return ""
    if SUPABASE_STORAGE_PUBLIC_BASE:
        return f"{SUPABASE_STORAGE_PUBLIC_BASE}/{encoded_path}"
    return f"{SUPABASE_URL}/storage/v1/object/public/{quote(SUPABASE_STORAGE_BUCKET, safe='')}/{encoded_path}"


def retention_until(uploaded_at: Optional[datetime] = None):
    uploaded = uploaded_at or app_now()
    return (uploaded + timedelta(days=STORAGE_RETENTION_DAYS)).date().isoformat()


def storage_path_from_url(path: str):
    if not path or not is_remote_image_path(path):
        return ""
    marker = f"/storage/v1/object/public/{quote(SUPABASE_STORAGE_BUCKET, safe='')}/"
    if marker not in path:
        return ""
    encoded_path = path.split(marker, 1)[1].split("?", 1)[0]
    return "/".join(unquote(part) for part in encoded_path.split("/"))


def inspection_storage_meta(folder: str, filename: str, public_path: str, uploaded_at: Optional[datetime] = None):
    uploaded = uploaded_at or app_now()
    uploaded_s = uploaded.isoformat(timespec="seconds")
    remote = is_remote_image_path(public_path)
    return {
        "storage_path": _storage_path(folder, filename) if remote else storage_path_from_url(public_path),
        "public_url": public_path if remote else "",
        "signed_url": "",
        "uploaded_at": uploaded_s,
        "retention_until": retention_until(uploaded),
    }


def save_inspection_image(data: bytes, filename: str, content_type: str, folder: str, local_dir: Path):
    remote_url = upload_to_supabase_storage(folder, filename, data, content_type)
    if remote_url:
        return remote_url, None
    local_dir.mkdir(parents=True, exist_ok=True)
    disk_path = local_dir / filename
    disk_path.write_bytes(data)
    return f"/uploads/{folder}/{filename}", disk_path


def is_remote_image_path(path: str):
    return bool(path and re.match(r"^https?://", path, flags=re.IGNORECASE))


def allowed_local_image_path(path: str):
    return bool(
        path
        and (
            path.startswith("/uploads/inspection_sheets/")
            or path.startswith("/uploads/inspection_slips/")
            or path.startswith("/uploads/support_requests/")
            or path.startswith("/uploads/receipts/")
        )
    )


def local_image_exists(path: str):
    return allowed_local_image_path(path) and (BASE_DIR / path.lstrip("/")).exists()


def inspection_image_info(path: str):
    path = path or ""
    if not path:
        return {"image_path": "", "image_available": False, "image_missing": False, "image_view_url": "", "image_label": "画像なし"}
    view_url = f"/inspection-file?path={quote(path, safe='')}"
    if is_remote_image_path(path):
        return {"image_path": path, "image_available": True, "image_missing": False, "image_view_url": view_url, "image_label": "画像あり"}
    if local_image_exists(path):
        return {"image_path": path, "image_available": True, "image_missing": False, "image_view_url": view_url, "image_label": "画像あり"}
    return {
        "image_path": path,
        "image_available": False,
        "image_missing": True,
        "image_view_url": view_url,
        "image_label": "画像ファイルが見つかりません。再アップロードしてください",
    }


def inspection_image_info_lazy(path: str):
    path = path or ""
    if not path:
        return {"image_path": "", "image_available": False, "image_missing": False, "image_view_url": "", "image_label": "画像なし"}
    view_url = f"/inspection-file?path={quote(path, safe='')}"
    return {"image_path": path, "image_available": True, "image_missing": False, "image_view_url": view_url, "image_label": "画像あり"}


def with_image_info(row, path_field="file_path"):
    item = dict(row)
    item.update(inspection_image_info(item.get(path_field, "")))
    return item


def attach_image_info(rows, path_field="file_path"):
    return [with_image_info(row, path_field) for row in rows]


def with_image_info_lazy(row, path_field="file_path"):
    item = dict(row)
    item.update(inspection_image_info_lazy(item.get(path_field, "")))
    return item


def attach_image_info_lazy(rows, path_field="file_path"):
    return [with_image_info_lazy(row, path_field) for row in rows]


def save_image_data_for_ocr(data: bytes, filename: str, content_type: str, folder: str, local_dir: Path):
    local_dir.mkdir(parents=True, exist_ok=True)
    disk_path = local_dir / filename
    disk_path.write_bytes(data)
    remote_url = upload_to_supabase_storage(folder, filename, data, content_type)
    public_path = remote_url or f"/uploads/{folder}/{filename}"
    return public_path, disk_path


def row_value(row, key, default=None):
    if not row:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    try:
        return row[key] if key in row.keys() else default
    except (AttributeError, KeyError, TypeError):
        return default


def company_id_for(user):
    return int(row_value(user, "company_id", 1) or 1)


def company_name_for(user):
    return row_value(user, "company_name", "") or APP_NAME


def is_platform_admin(user):
    return row_value(user, "role") == "admin" and company_id_for(user) == PLATFORM_COMPANY_ID


def can_manage_companies(user):
    return is_platform_admin(user)


def deleted_username_for(user_id):
    return f"deleted_member_{user_id}_{app_now().strftime('%Y%m%d%H%M%S%f')}"


def release_purged_username(username: str):
    username = (username or "").strip()
    if not username:
        return
    rows = query_all("SELECT id FROM users WHERE username=? AND COALESCE(purged,0)=1", (username,))
    for row in rows:
        released_username = deleted_username_for(row["id"])
        now = app_now().isoformat(timespec="seconds")
        execute(
            """UPDATE users
               SET username=?, password_hash=?, deleted_at=COALESCE(NULLIF(deleted_at,''), ?)
               WHERE id=? AND username=? AND COALESCE(purged,0)=1""",
            (released_username, hash_password(released_username), now, row["id"], username),
        )


def execute(sql: str, params=()):
    ensure_db_initialized()
    with db() as conn:
        cur = conn.execute(sql, params)
        conn.commit()
        return cur


def query_one(sql: str, params=()):
    ensure_db_initialized()
    with db() as conn:
        return conn.execute(sql, params).fetchone()


def query_all(sql: str, params=()):
    ensure_db_initialized()
    with db() as conn:
        return conn.execute(sql, params).fetchall()


def row_to_dict(row):
    if not row:
        return {}
    if isinstance(row, dict):
        return dict(row)
    try:
        return {key: row[key] for key in row.keys()}
    except (AttributeError, KeyError, TypeError):
        return {}


def audit_log(company_id, actor, action, table_name, record_id="", summary="", before=None, after=None, conn=None):
    payload = (
        int(company_id or PLATFORM_COMPANY_ID),
        row_value(actor, "id"),
        row_value(actor, "role", ""),
        action,
        table_name,
        str(record_id or ""),
        summary or "",
        json.dumps(before or {}, ensure_ascii=False, default=str),
        json.dumps(after or {}, ensure_ascii=False, default=str),
        app_now().isoformat(timespec="seconds"),
    )
    sql = """INSERT INTO audit_logs(company_id, actor_id, actor_role, action, table_name, record_id, summary, before_data, after_data, created_at)
             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"""
    if conn:
        conn.execute(sql, payload)
    else:
        execute(sql, payload)


def vehicle_rates_for_company(company_id: int):
    row = query_one("SELECT * FROM company_vehicle_rates WHERE company_id=?", (company_id,))
    if row:
        return row
    legacy = query_one("SELECT * FROM vehicle_rates WHERE id=1")
    if legacy:
        now = app_now().isoformat(timespec="seconds")
        execute(
            """INSERT INTO company_vehicle_rates(company_id, daily_fee, monthly_fee, updated_at)
               VALUES (?, ?, ?, ?)
               ON CONFLICT(company_id) DO UPDATE SET daily_fee=excluded.daily_fee,
               monthly_fee=excluded.monthly_fee, updated_at=excluded.updated_at""",
            (company_id, legacy["daily_fee"], legacy["monthly_fee"], now),
        )
        return query_one("SELECT * FROM company_vehicle_rates WHERE company_id=?", (company_id,))
    return {"company_id": company_id, "daily_fee": 1500, "monthly_fee": 30000}


def feature_price_label(feature):
    price = int(row_value(feature, "monthly_price", row_value(feature, "price", 0)) or 0)
    return "基本プランに含む" if price <= 0 else f"月額{price:,}円"


def is_feature_enabled(features, key: str) -> bool:
    keys = FEATURE_ALIASES.get(key, (key,))
    return any(bool(features.get(item, True)) for item in keys)


def load_company_features(company_id: int):
    cache_key = int(company_id or PLATFORM_COMPANY_ID)
    cached = FEATURE_CACHE.get(cache_key)
    now_m = monotonic()
    if cached and now_m - cached["at"] < FEATURE_CACHE_TTL_SECONDS:
        return dict(cached["features"])
    enabled = {feature["key"]: True for feature in FEATURE_CATALOG}
    rows = query_all("SELECT feature_key, enabled FROM company_features WHERE company_id=?", (company_id,))
    for row in rows:
        key = row["feature_key"]
        if key in FEATURE_KEYS:
            enabled[key] = bool(row["enabled"])
    FEATURE_CACHE[cache_key] = {"at": now_m, "features": dict(enabled)}
    return enabled


def clear_feature_cache(company_id: Optional[int] = None):
    if company_id is None:
        FEATURE_CACHE.clear()
    else:
        FEATURE_CACHE.pop(int(company_id or PLATFORM_COMPANY_ID), None)


def feature_price_rows():
    rows = query_all("SELECT * FROM feature_prices WHERE active=1 ORDER BY feature_key")
    if not rows:
        return [dict(feature, feature_key=feature["key"], monthly_price=feature["price"]) for feature in FEATURE_CATALOG]
    return rows


def company_feature_rows(company_id: int):
    enabled = load_company_features(company_id)
    rows = []
    for feature in feature_price_rows():
        item = dict(feature)
        item["enabled"] = bool(enabled.get(item["feature_key"], True))
        item["key"] = item["feature_key"]
        item["price"] = item["monthly_price"]
        item["price_label"] = feature_price_label(feature)
        rows.append(item)
    return rows


def feature_enabled_for_company(company_id: int, key: str) -> bool:
    return is_feature_enabled(load_company_features(company_id), key)


def require_feature(user, key: str):
    if not feature_enabled_for_company(company_id_for(user), key):
        raise HTTPException(status_code=403, detail="この機能は未契約です")


def monthly_amount_for_company(company_id: int):
    row = query_one(
        """SELECT COALESCE(SUM(fp.monthly_price), 0) AS total
           FROM company_features cf
           JOIN feature_prices fp ON fp.feature_key=cf.feature_key
           WHERE cf.company_id=? AND cf.enabled=1 AND fp.active=1""",
        (company_id,),
    )
    return int(row_value(row, "total", 0) or 0)


def pending_feature_changes(company_id: Optional[int] = None, limit: int = 100):
    params = []
    where = "WHERE ch.status='pending'"
    if company_id is not None:
        where += " AND ch.company_id=?"
        params.append(company_id)
    params.append(limit)
    return query_all(
        f"""SELECT ch.*, c.name AS company_name, fp.label AS feature_label
            FROM company_feature_changes ch
            JOIN companies c ON c.id=ch.company_id
            LEFT JOIN feature_prices fp ON fp.feature_key=ch.feature_key
            {where}
            ORDER BY ch.effective_date, ch.created_at DESC
            LIMIT ?""",
        params,
    )


def apply_pending_feature_changes(company_id: int):
    now = app_now().isoformat(timespec="seconds")
    rows = query_all(
        """SELECT * FROM company_feature_changes
           WHERE company_id=? AND status='pending'
           ORDER BY feature_key, created_at""",
        (company_id,),
    )
    with db() as conn:
        for row in rows:
            conn.execute(
                """INSERT INTO company_features(company_id, feature_key, enabled, updated_at)
                   VALUES (?, ?, ?, ?)
                   ON CONFLICT(company_id, feature_key) DO UPDATE SET
                   enabled=excluded.enabled, updated_at=excluded.updated_at""",
                (company_id, row["feature_key"], row["enabled"], now),
            )
            conn.execute(
                "UPDATE company_feature_changes SET status='applied', applied_at=? WHERE id=?",
                (now, row["id"]),
            )
        conn.commit()
    clear_feature_cache(company_id)
    return len(rows)


def plan_label_for_company(status: str, enabled_count: int, monthly_total: int):
    if status == "お試し":
        return "お試し"
    if enabled_count >= len(FEATURE_CATALOG) - 1:
        return "フル機能"
    if monthly_total <= 3000:
        return "基本プラン"
    return "標準プラン"


def company_feature_totals():
    rows = query_all(
        """SELECT c.id AS company_id,
                  COALESCE(SUM(CASE WHEN cf.enabled=1 THEN fp.monthly_price ELSE 0 END), 0) AS monthly_total,
                  COALESCE(SUM(CASE WHEN cf.enabled=1 THEN 1 ELSE 0 END), 0) AS enabled_count
           FROM companies c
           LEFT JOIN company_features cf ON cf.company_id=c.id
           LEFT JOIN feature_prices fp ON fp.feature_key=cf.feature_key AND fp.active=1
           GROUP BY c.id"""
    )
    return {row["company_id"]: {"monthly_total": int(row["monthly_total"] or 0), "enabled_count": int(row["enabled_count"] or 0)} for row in rows}


def ensure_column(conn, table: str, column: str, definition: str):
    if USE_POSTGRES:
        exists = conn.execute(
            """SELECT 1 FROM information_schema.columns
               WHERE table_schema = 'public' AND table_name = ? AND column_name = ?""",
            (table, column),
        ).fetchone()
        if not exists:
            conn.execute(
                f"ALTER TABLE {_quote_identifier(table)} ADD COLUMN {_quote_identifier(column)} {definition}"
            )
        return

    columns = [row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()]
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


ISSUE_STATUSES = ["未対応", "確認中", "修理依頼中", "対応済", "修理済"]
ISSUE_STATUS_CLASSES = {
    "未対応": "todo",
    "確認中": "checking",
    "修理依頼中": "repair-requested",
    "対応済": "handled",
    "修理済": "repaired",
}


def ensure_vehicle_issue_schema():
    with db() as conn:
        conn.execute(
            """CREATE TABLE IF NOT EXISTS vehicle_issues (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                issue_date TEXT NOT NULL,
                vehicle_name TEXT NOT NULL,
                severity TEXT NOT NULL,
                detail TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT '未対応',
                created_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            )"""
        )
        ensure_column(conn, "vehicle_issues", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "vehicle_issues", "status", "TEXT NOT NULL DEFAULT '未対応'")
        ensure_column(conn, "vehicle_issues", "created_at", "TEXT DEFAULT ''")
        conn.execute(
            """CREATE TABLE IF NOT EXISTS vehicle_issue_status_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                issue_id INTEGER NOT NULL,
                status TEXT NOT NULL,
                changed_by INTEGER NOT NULL,
                changed_at TEXT NOT NULL,
                note TEXT DEFAULT '',
                FOREIGN KEY(issue_id) REFERENCES vehicle_issues(id),
                FOREIGN KEY(changed_by) REFERENCES users(id)
            )"""
        )
        ensure_column(conn, "vehicle_issue_status_logs", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "vehicle_issue_status_logs", "note", "TEXT DEFAULT ''")
        conn.commit()


def normalize_issue_status(status: str) -> str:
    return status if status in ISSUE_STATUSES else "未対応"


def date_label(day: date):
    weekdays = "月火水木金土日"
    return f"{day.isoformat()}（{weekdays[day.weekday()]}）"


def simple_japanese_holidays(year: int):
    days = {
        date(year, 1, 1),
        date(year, 2, 11),
        date(year, 2, 23),
        date(year, 4, 29),
        date(year, 5, 3),
        date(year, 5, 4),
        date(year, 5, 5),
        date(year, 8, 11),
        date(year, 11, 3),
        date(year, 11, 23),
    }

    def nth_monday(month, nth):
        first = date(year, month, 1)
        offset = (7 - first.weekday()) % 7
        return first + timedelta(days=offset + 7 * (nth - 1))

    days.update(
        {
            nth_monday(1, 2),
            nth_monday(7, 3),
            nth_monday(9, 3),
            nth_monday(10, 2),
            date(year, 3, 20),
            date(year, 9, 23),
        }
    )
    return days


def _init_db_schema():
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS app_metadata (
                meta_key TEXT PRIMARY KEY,
                meta_value TEXT DEFAULT '',
                updated_at TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                code TEXT DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                name TEXT NOT NULL,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin','member')),
                email TEXT DEFAULT '',
                phone TEXT DEFAULT '',
                vehicle TEXT DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                purged INTEGER NOT NULL DEFAULT 0,
                deleted_at TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS rates (
                user_id INTEGER PRIMARY KEY,
                company_id INTEGER NOT NULL DEFAULT 1,
                delivery_unit INTEGER NOT NULL DEFAULT 180,
                transfer_unit INTEGER NOT NULL DEFAULT 80,
                night_unit INTEGER NOT NULL DEFAULT 120,
                pickup_unit INTEGER NOT NULL DEFAULT 100,
                large_unit INTEGER NOT NULL DEFAULT 150,
                vehicle_rental_type TEXT NOT NULL DEFAULT 'none',
                vehicle_daily_fee INTEGER NOT NULL DEFAULT 0,
                vehicle_monthly_fee INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS vehicle_rates (
                id INTEGER PRIMARY KEY CHECK(id = 1),
                daily_fee INTEGER NOT NULL DEFAULT 1500,
                monthly_fee INTEGER NOT NULL DEFAULT 30000
            );
            CREATE TABLE IF NOT EXISTS company_vehicle_rates (
                company_id INTEGER PRIMARY KEY,
                daily_fee INTEGER NOT NULL DEFAULT 1500,
                monthly_fee INTEGER NOT NULL DEFAULT 30000,
                updated_at TEXT DEFAULT '',
                FOREIGN KEY(company_id) REFERENCES companies(id)
            );
            CREATE TABLE IF NOT EXISTS vehicles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                name TEXT NOT NULL,
                plate_number TEXT DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                last_odometer INTEGER NOT NULL DEFAULT 0,
                last_used_at TEXT DEFAULT '',
                oil_change_odometer INTEGER NOT NULL DEFAULT 0,
                oil_change_date TEXT DEFAULT '',
                inspection_due TEXT DEFAULT '',
                tire_note TEXT DEFAULT '',
                model TEXT DEFAULT '',
                color TEXT DEFAULT '',
                depot_id INTEGER,
                primary_user_id INTEGER,
                vehicle_status TEXT DEFAULT '使用中',
                oil_change_interval INTEGER NOT NULL DEFAULT 5000,
                oil_note TEXT DEFAULT '',
                inspection_cost INTEGER NOT NULL DEFAULT 0,
                inspection_note TEXT DEFAULT '',
                tire_change_date TEXT DEFAULT '',
                tire_change_odometer INTEGER NOT NULL DEFAULT 0,
                tire_cost INTEGER NOT NULL DEFAULT 0,
                tire_status_note TEXT DEFAULT '',
                repair_cost INTEGER NOT NULL DEFAULT 0,
                parts_cost INTEGER NOT NULL DEFAULT 0,
                car_wash_cost INTEGER NOT NULL DEFAULT 0,
                insurance_fee INTEGER NOT NULL DEFAULT 0,
                parking_fee INTEGER NOT NULL DEFAULT 0,
                other_vehicle_cost INTEGER NOT NULL DEFAULT 0,
                admin_memo TEXT DEFAULT '',
                issue_memo TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT DEFAULT '',
                UNIQUE(company_id, name),
                FOREIGN KEY(company_id) REFERENCES companies(id)
            );
            CREATE TABLE IF NOT EXISTS vehicle_expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                vehicle_id INTEGER,
                depot_id INTEGER,
                expense_date TEXT NOT NULL,
                category TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                payee TEXT DEFAULT '',
                memo TEXT DEFAULT '',
                receipt_path TEXT DEFAULT '',
                created_by INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY(company_id) REFERENCES companies(id),
                FOREIGN KEY(vehicle_id) REFERENCES vehicles(id)
            );
            CREATE TABLE IF NOT EXISTS company_expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                depot_id INTEGER,
                expense_date TEXT NOT NULL,
                category TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                payee TEXT DEFAULT '',
                memo TEXT DEFAULT '',
                receipt_path TEXT DEFAULT '',
                created_by INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY(company_id) REFERENCES companies(id)
            );
            CREATE TABLE IF NOT EXISTS depot_rate_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                depot_id INTEGER NOT NULL,
                base_unit REAL NOT NULL DEFAULT 0,
                transfer_unit REAL NOT NULL DEFAULT 0,
                night_unit REAL NOT NULL DEFAULT 0,
                pickup_unit REAL NOT NULL DEFAULT 0,
                large_unit REAL NOT NULL DEFAULT 0,
                effective_start TEXT NOT NULL,
                effective_end TEXT DEFAULT '',
                memo TEXT DEFAULT '',
                created_by INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY(company_id) REFERENCES companies(id)
            );
            CREATE TABLE IF NOT EXISTS work_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                work_date TEXT NOT NULL,
                log_type TEXT NOT NULL CHECK(log_type IN ('start','end')),
                logged_at TEXT NOT NULL,
                vehicle_id INTEGER,
                vehicle_name TEXT DEFAULT '',
                odometer INTEGER NOT NULL DEFAULT 0,
                start_odometer INTEGER NOT NULL DEFAULT 0,
                end_odometer INTEGER NOT NULL DEFAULT 0,
                distance_km INTEGER NOT NULL DEFAULT 0,
                rest_minutes INTEGER NOT NULL DEFAULT 0,
                waiting_minutes INTEGER NOT NULL DEFAULT 0,
                loading_minutes INTEGER NOT NULL DEFAULT 0,
                alcohol_result TEXT NOT NULL,
                detector_used TEXT NOT NULL,
                intoxicated TEXT NOT NULL,
                health_status TEXT NOT NULL,
                face_check TEXT NOT NULL,
                breath_check TEXT NOT NULL,
                voice_check TEXT NOT NULL,
                admin_confirm TEXT DEFAULT '',
                notes TEXT DEFAULT '',
                is_deleted INTEGER NOT NULL DEFAULT 0,
                deleted_at TEXT DEFAULT '',
                deleted_by INTEGER,
                delete_reason TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS work_log_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                work_date TEXT NOT NULL,
                session_no INTEGER NOT NULL DEFAULT 1,
                status TEXT NOT NULL DEFAULT 'working',
                start_time TEXT DEFAULT '',
                end_time TEXT DEFAULT '',
                start_log_id INTEGER,
                end_log_id INTEGER,
                vehicle_id INTEGER,
                vehicle_name TEXT DEFAULT '',
                start_odometer INTEGER NOT NULL DEFAULT 0,
                end_odometer INTEGER NOT NULL DEFAULT 0,
                distance_km INTEGER NOT NULL DEFAULT 0,
                alcohol_check_start TEXT DEFAULT '',
                alcohol_check_end TEXT DEFAULT '',
                call_check_start TEXT DEFAULT '',
                call_check_end TEXT DEFAULT '',
                delivery_count INTEGER NOT NULL DEFAULT 0,
                transfer_count INTEGER NOT NULL DEFAULT 0,
                night_count INTEGER NOT NULL DEFAULT 0,
                pickup_count INTEGER NOT NULL DEFAULT 0,
                large_count INTEGER NOT NULL DEFAULT 0,
                inspection_sheet_id INTEGER,
                inspection_image_path TEXT DEFAULT '',
                notes TEXT DEFAULT '',
                is_deleted INTEGER NOT NULL DEFAULT 0,
                deleted_at TEXT DEFAULT '',
                deleted_by INTEGER,
                delete_reason TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(start_log_id) REFERENCES work_logs(id),
                FOREIGN KEY(end_log_id) REFERENCES work_logs(id)
            );
            CREATE TABLE IF NOT EXISTS deliveries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                work_date TEXT NOT NULL,
                completed INTEGER NOT NULL DEFAULT 0,
                transfer INTEGER NOT NULL DEFAULT 0,
                night INTEGER NOT NULL DEFAULT 0,
                pickup INTEGER NOT NULL DEFAULT 0,
                large INTEGER NOT NULL DEFAULT 0,
                vehicle_rental TEXT NOT NULL DEFAULT 'none',
                memo TEXT DEFAULT '',
                inspection_sheet_path TEXT DEFAULT '',
                is_deleted INTEGER NOT NULL DEFAULT 0,
                deleted_at TEXT DEFAULT '',
                deleted_by INTEGER,
                delete_reason TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(user_id, work_date),
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS delivery_corrections (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                delivery_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                actor_id INTEGER NOT NULL,
                work_date TEXT NOT NULL,
                before_completed INTEGER NOT NULL DEFAULT 0,
                before_transfer INTEGER NOT NULL DEFAULT 0,
                before_night INTEGER NOT NULL DEFAULT 0,
                before_pickup INTEGER NOT NULL DEFAULT 0,
                before_large INTEGER NOT NULL DEFAULT 0,
                after_completed INTEGER NOT NULL DEFAULT 0,
                after_transfer INTEGER NOT NULL DEFAULT 0,
                after_night INTEGER NOT NULL DEFAULT 0,
                after_pickup INTEGER NOT NULL DEFAULT 0,
                after_large INTEGER NOT NULL DEFAULT 0,
                source TEXT DEFAULT '',
                changed_at TEXT NOT NULL,
                FOREIGN KEY(delivery_id) REFERENCES deliveries(id),
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(actor_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS work_log_corrections (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                work_date TEXT NOT NULL,
                work_log_id INTEGER,
                session_id INTEGER,
                target_table TEXT DEFAULT '',
                before_data TEXT DEFAULT '',
                after_data TEXT DEFAULT '',
                actor_id INTEGER,
                changed_at TEXT NOT NULL,
                note TEXT DEFAULT '',
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(actor_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS work_day_deletions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                work_date TEXT NOT NULL,
                actor_id INTEGER,
                reason TEXT DEFAULT '',
                before_data TEXT DEFAULT '',
                deleted_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(actor_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS vehicle_issues (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                issue_date TEXT NOT NULL,
                vehicle_name TEXT NOT NULL,
                severity TEXT NOT NULL,
                detail TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT '未対応',
                created_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS holiday_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                request_date TEXT NOT NULL,
                reason TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                UNIQUE(user_id, request_date),
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS shifts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                shift_date TEXT NOT NULL,
                status TEXT NOT NULL,
                start_time TEXT DEFAULT '',
                end_time TEXT DEFAULT '',
                note TEXT DEFAULT '',
                district_id INTEGER,
                town_id INTEGER,
                area_label TEXT DEFAULT '',
                decided INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL,
                UNIQUE(user_id, shift_date),
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS depots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                name TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, name)
            );
            CREATE TABLE IF NOT EXISTS districts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                depot_id INTEGER,
                name TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, depot_id, name),
                FOREIGN KEY(depot_id) REFERENCES depots(id)
            );
            CREATE TABLE IF NOT EXISTS towns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                district_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                UNIQUE(company_id, district_id, name),
                FOREIGN KEY(district_id) REFERENCES districts(id)
            );
            CREATE TABLE IF NOT EXISTS shift_towns (
                company_id INTEGER NOT NULL DEFAULT 1,
                shift_id INTEGER NOT NULL,
                town_id INTEGER NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                PRIMARY KEY(shift_id, town_id),
                FOREIGN KEY(shift_id) REFERENCES shifts(id) ON DELETE CASCADE,
                FOREIGN KEY(town_id) REFERENCES towns(id)
            );
            CREATE TABLE IF NOT EXISTS vehicle_issue_status_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                issue_id INTEGER NOT NULL,
                status TEXT NOT NULL,
                changed_by INTEGER NOT NULL,
                changed_at TEXT NOT NULL,
                note TEXT DEFAULT '',
                FOREIGN KEY(issue_id) REFERENCES vehicle_issues(id),
                FOREIGN KEY(changed_by) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS inspection_sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                sheet_date TEXT NOT NULL,
                delivery_date TEXT DEFAULT '',
                file_path TEXT NOT NULL,
                storage_path TEXT DEFAULT '',
                public_url TEXT DEFAULT '',
                signed_url TEXT DEFAULT '',
                original_filename TEXT DEFAULT '',
                content_type TEXT DEFAULT '',
                ocr_status TEXT NOT NULL DEFAULT '未処理',
                ocr_text TEXT DEFAULT '',
                ocr_completed INTEGER NOT NULL DEFAULT 0,
                ocr_pickup INTEGER NOT NULL DEFAULT 0,
                ocr_acceptance INTEGER NOT NULL DEFAULT 0,
                ocr_received INTEGER NOT NULL DEFAULT 0,
                uploaded_at TEXT DEFAULT '',
                retention_until TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id),
                UNIQUE(user_id, sheet_date)
            );
            CREATE TABLE IF NOT EXISTS inspection_slips (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                delivery_id INTEGER NOT NULL,
                slip_date TEXT NOT NULL,
                file_path TEXT NOT NULL,
                storage_path TEXT DEFAULT '',
                public_url TEXT DEFAULT '',
                signed_url TEXT DEFAULT '',
                original_filename TEXT DEFAULT '',
                content_type TEXT DEFAULT '',
                uploaded_at TEXT NOT NULL,
                retention_until TEXT DEFAULT '',
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(delivery_id) REFERENCES deliveries(id),
                UNIQUE(user_id, slip_date)
            );
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                user_id INTEGER NOT NULL,
                email TEXT NOT NULL,
                token_hash TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                used_at TEXT DEFAULT '',
                dev_reset_url TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS company_features (
                company_id INTEGER NOT NULL,
                feature_key TEXT NOT NULL,
                enabled INTEGER NOT NULL DEFAULT 1,
                updated_at TEXT NOT NULL,
                PRIMARY KEY(company_id, feature_key),
                FOREIGN KEY(company_id) REFERENCES companies(id)
            );
            CREATE TABLE IF NOT EXISTS feature_prices (
                feature_key TEXT PRIMARY KEY,
                label TEXT NOT NULL,
                monthly_price INTEGER NOT NULL DEFAULT 0,
                note TEXT DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS company_feature_changes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                feature_key TEXT NOT NULL,
                enabled INTEGER NOT NULL DEFAULT 1,
                effective_date TEXT NOT NULL,
                requested_by INTEGER,
                status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT NOT NULL,
                applied_at TEXT DEFAULT '',
                FOREIGN KEY(company_id) REFERENCES companies(id),
                FOREIGN KEY(requested_by) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS support_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                sender_id INTEGER NOT NULL,
                request_type TEXT NOT NULL,
                subject TEXT NOT NULL,
                body TEXT NOT NULL,
                urgency TEXT NOT NULL DEFAULT '中',
                image_path TEXT DEFAULT '',
                status TEXT NOT NULL DEFAULT '未対応',
                reply_note TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(company_id) REFERENCES companies(id),
                FOREIGN KEY(sender_id) REFERENCES users(id)
            );
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL DEFAULT 1,
                actor_id INTEGER,
                actor_role TEXT DEFAULT '',
                action TEXT NOT NULL,
                table_name TEXT NOT NULL,
                record_id TEXT DEFAULT '',
                summary TEXT DEFAULT '',
                before_data TEXT DEFAULT '',
                after_data TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY(actor_id) REFERENCES users(id)
            );
            """
        )
        ensure_column(conn, "companies", "code", "TEXT DEFAULT ''")
        ensure_column(conn, "companies", "active", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "companies", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "companies", "status", "TEXT DEFAULT '利用中'")
        ensure_column(conn, "companies", "contract_start_date", "TEXT DEFAULT ''")
        ensure_column(conn, "companies", "next_renewal_date", "TEXT DEFAULT ''")
        ensure_column(conn, "companies", "memo", "TEXT DEFAULT ''")
        ensure_column(conn, "users", "phone", "TEXT DEFAULT ''")
        ensure_column(conn, "users", "email", "TEXT DEFAULT ''")
        ensure_column(conn, "users", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "users", "vehicle", "TEXT DEFAULT ''")
        ensure_column(conn, "users", "last_vehicle_id", "INTEGER")
        ensure_column(conn, "users", "oil_change_date", "TEXT DEFAULT ''")
        ensure_column(conn, "users", "vehicle_inspection_due", "TEXT DEFAULT ''")
        ensure_column(conn, "users", "notification_email", "TEXT DEFAULT ''")
        ensure_column(conn, "users", "member_notes", "TEXT DEFAULT ''")
        ensure_column(conn, "users", "active", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "users", "purged", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "users", "deleted_at", "TEXT DEFAULT ''")
        ensure_column(conn, "users", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "rates", "delivery_unit", "INTEGER NOT NULL DEFAULT 180")
        ensure_column(conn, "rates", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "rates", "transfer_unit", "INTEGER NOT NULL DEFAULT 80")
        ensure_column(conn, "rates", "night_unit", "INTEGER NOT NULL DEFAULT 120")
        ensure_column(conn, "rates", "pickup_unit", "INTEGER NOT NULL DEFAULT 100")
        ensure_column(conn, "rates", "large_unit", "INTEGER NOT NULL DEFAULT 150")
        ensure_column(conn, "rates", "vehicle_rental_type", "TEXT NOT NULL DEFAULT 'none'")
        ensure_column(conn, "rates", "vehicle_daily_fee", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "rates", "vehicle_monthly_fee", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicle_rates", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "company_vehicle_rates", "daily_fee", "INTEGER NOT NULL DEFAULT 1500")
        ensure_column(conn, "company_vehicle_rates", "monthly_fee", "INTEGER NOT NULL DEFAULT 30000")
        ensure_column(conn, "company_vehicle_rates", "updated_at", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "alcohol_result", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "work_logs", "vehicle_id", "INTEGER")
        ensure_column(conn, "work_logs", "vehicle_name", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "odometer", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_logs", "start_odometer", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_logs", "end_odometer", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_logs", "distance_km", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_logs", "rest_minutes", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_logs", "waiting_minutes", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_logs", "loading_minutes", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_logs", "detector_used", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "intoxicated", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "health_status", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "face_check", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "breath_check", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "voice_check", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "admin_confirm", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "notes", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "is_deleted", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_logs", "deleted_at", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "deleted_by", "INTEGER")
        ensure_column(conn, "work_logs", "delete_reason", "TEXT DEFAULT ''")
        ensure_column(conn, "work_logs", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "work_log_sessions", "user_id", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_sessions", "work_date", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "session_no", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "work_log_sessions", "status", "TEXT DEFAULT 'working'")
        ensure_column(conn, "work_log_sessions", "start_time", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "end_time", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "start_log_id", "INTEGER")
        ensure_column(conn, "work_log_sessions", "end_log_id", "INTEGER")
        ensure_column(conn, "work_log_sessions", "vehicle_id", "INTEGER")
        ensure_column(conn, "work_log_sessions", "vehicle_name", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "start_odometer", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_sessions", "end_odometer", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_sessions", "distance_km", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_sessions", "alcohol_check_start", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "alcohol_check_end", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "call_check_start", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "call_check_end", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "delivery_count", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_sessions", "transfer_count", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_sessions", "night_count", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_sessions", "pickup_count", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_sessions", "large_count", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_sessions", "inspection_sheet_id", "INTEGER")
        ensure_column(conn, "work_log_sessions", "inspection_image_path", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "notes", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "is_deleted", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_sessions", "deleted_at", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "deleted_by", "INTEGER")
        ensure_column(conn, "work_log_sessions", "delete_reason", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_sessions", "updated_at", "TEXT DEFAULT ''")
        ensure_column(conn, "deliveries", "completed", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "deliveries", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "deliveries", "vehicle_id", "INTEGER")
        ensure_column(conn, "deliveries", "vehicle_name", "TEXT DEFAULT ''")
        ensure_column(conn, "deliveries", "transfer", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "deliveries", "night", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "deliveries", "pickup", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "deliveries", "large", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "deliveries", "vehicle_rental", "TEXT NOT NULL DEFAULT 'none'")
        ensure_column(conn, "deliveries", "memo", "TEXT DEFAULT ''")
        ensure_column(conn, "deliveries", "inspection_sheet_path", "TEXT DEFAULT ''")
        ensure_column(conn, "deliveries", "is_deleted", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "deliveries", "deleted_at", "TEXT DEFAULT ''")
        ensure_column(conn, "deliveries", "deleted_by", "INTEGER")
        ensure_column(conn, "deliveries", "delete_reason", "TEXT DEFAULT ''")
        ensure_column(conn, "deliveries", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "deliveries", "updated_at", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "vehicles", "name", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "plate_number", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "active", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "vehicles", "last_odometer", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "last_used_at", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "oil_change_odometer", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "oil_change_date", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "inspection_due", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "tire_note", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "model", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "color", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "depot_id", "INTEGER")
        ensure_column(conn, "vehicles", "primary_user_id", "INTEGER")
        ensure_column(conn, "vehicles", "vehicle_status", "TEXT DEFAULT '使用中'")
        ensure_column(conn, "vehicles", "oil_change_interval", "INTEGER NOT NULL DEFAULT 5000")
        ensure_column(conn, "vehicles", "oil_note", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "inspection_cost", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "inspection_note", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "tire_change_date", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "tire_change_odometer", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "tire_cost", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "tire_status_note", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "repair_cost", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "parts_cost", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "car_wash_cost", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "insurance_fee", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "parking_fee", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "other_vehicle_cost", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicles", "admin_memo", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "issue_memo", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicles", "updated_at", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicle_expenses", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "vehicle_expenses", "vehicle_id", "INTEGER")
        ensure_column(conn, "vehicle_expenses", "depot_id", "INTEGER")
        ensure_column(conn, "vehicle_expenses", "expense_date", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicle_expenses", "category", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicle_expenses", "amount", "REAL NOT NULL DEFAULT 0")
        ensure_column(conn, "vehicle_expenses", "payee", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicle_expenses", "memo", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicle_expenses", "receipt_path", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicle_expenses", "created_by", "INTEGER")
        ensure_column(conn, "vehicle_expenses", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "company_expenses", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "company_expenses", "depot_id", "INTEGER")
        ensure_column(conn, "company_expenses", "expense_date", "TEXT DEFAULT ''")
        ensure_column(conn, "company_expenses", "category", "TEXT DEFAULT ''")
        ensure_column(conn, "company_expenses", "amount", "REAL NOT NULL DEFAULT 0")
        ensure_column(conn, "company_expenses", "payee", "TEXT DEFAULT ''")
        ensure_column(conn, "company_expenses", "memo", "TEXT DEFAULT ''")
        ensure_column(conn, "company_expenses", "receipt_path", "TEXT DEFAULT ''")
        ensure_column(conn, "company_expenses", "created_by", "INTEGER")
        ensure_column(conn, "company_expenses", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "depot_rate_settings", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "depot_rate_settings", "depot_id", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "depot_rate_settings", "base_unit", "REAL NOT NULL DEFAULT 0")
        ensure_column(conn, "depot_rate_settings", "transfer_unit", "REAL NOT NULL DEFAULT 0")
        ensure_column(conn, "depot_rate_settings", "night_unit", "REAL NOT NULL DEFAULT 0")
        ensure_column(conn, "depot_rate_settings", "pickup_unit", "REAL NOT NULL DEFAULT 0")
        ensure_column(conn, "depot_rate_settings", "large_unit", "REAL NOT NULL DEFAULT 0")
        ensure_column(conn, "depot_rate_settings", "effective_start", "TEXT DEFAULT ''")
        ensure_column(conn, "depot_rate_settings", "effective_end", "TEXT DEFAULT ''")
        ensure_column(conn, "depot_rate_settings", "memo", "TEXT DEFAULT ''")
        ensure_column(conn, "depot_rate_settings", "created_by", "INTEGER")
        ensure_column(conn, "depot_rate_settings", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "delivery_corrections", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "delivery_corrections", "source", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_corrections", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "work_log_corrections", "user_id", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_log_corrections", "work_date", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_corrections", "work_log_id", "INTEGER")
        ensure_column(conn, "work_log_corrections", "session_id", "INTEGER")
        ensure_column(conn, "work_log_corrections", "target_table", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_corrections", "before_data", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_corrections", "after_data", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_corrections", "actor_id", "INTEGER")
        ensure_column(conn, "work_log_corrections", "changed_at", "TEXT DEFAULT ''")
        ensure_column(conn, "work_log_corrections", "note", "TEXT DEFAULT ''")
        ensure_column(conn, "work_day_deletions", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "work_day_deletions", "user_id", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "work_day_deletions", "work_date", "TEXT DEFAULT ''")
        ensure_column(conn, "work_day_deletions", "actor_id", "INTEGER")
        ensure_column(conn, "work_day_deletions", "reason", "TEXT DEFAULT ''")
        ensure_column(conn, "work_day_deletions", "before_data", "TEXT DEFAULT ''")
        ensure_column(conn, "work_day_deletions", "deleted_at", "TEXT DEFAULT ''")
        ensure_column(conn, "shifts", "start_time", "TEXT DEFAULT ''")
        ensure_column(conn, "shifts", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "shifts", "end_time", "TEXT DEFAULT ''")
        ensure_column(conn, "shifts", "note", "TEXT DEFAULT ''")
        ensure_column(conn, "shifts", "district_id", "INTEGER")
        ensure_column(conn, "shifts", "town_id", "INTEGER")
        ensure_column(conn, "shifts", "area_label", "TEXT DEFAULT ''")
        ensure_column(conn, "shifts", "decided", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "shifts", "updated_at", "TEXT DEFAULT ''")
        ensure_column(conn, "depots", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "districts", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "towns", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "shift_towns", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "shift_towns", "active", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "inspection_sheets", "file_path", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_sheets", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "inspection_sheets", "delivery_date", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_sheets", "storage_path", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_sheets", "public_url", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_sheets", "signed_url", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_sheets", "original_filename", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_sheets", "content_type", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_sheets", "ocr_status", "TEXT DEFAULT '未処理'")
        ensure_column(conn, "inspection_sheets", "ocr_text", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_sheets", "ocr_completed", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "inspection_sheets", "ocr_pickup", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "inspection_sheets", "ocr_acceptance", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "inspection_sheets", "ocr_received", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "inspection_sheets", "uploaded_at", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_sheets", "retention_until", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_sheets", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_slips", "file_path", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_slips", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "inspection_slips", "storage_path", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_slips", "public_url", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_slips", "signed_url", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_slips", "original_filename", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_slips", "content_type", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_slips", "uploaded_at", "TEXT DEFAULT ''")
        ensure_column(conn, "inspection_slips", "retention_until", "TEXT DEFAULT ''")
        ensure_column(conn, "password_reset_tokens", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "password_reset_tokens", "email", "TEXT DEFAULT ''")
        ensure_column(conn, "password_reset_tokens", "token_hash", "TEXT DEFAULT ''")
        ensure_column(conn, "password_reset_tokens", "expires_at", "TEXT DEFAULT ''")
        ensure_column(conn, "password_reset_tokens", "used_at", "TEXT DEFAULT ''")
        ensure_column(conn, "password_reset_tokens", "dev_reset_url", "TEXT DEFAULT ''")
        ensure_column(conn, "password_reset_tokens", "created_at", "TEXT DEFAULT ''")
        ensure_column(conn, "vehicle_issues", "status", "TEXT NOT NULL DEFAULT '未対応'")
        ensure_column(conn, "vehicle_issues", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "holiday_requests", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "vehicle_issue_status_logs", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "audit_logs", "company_id", "INTEGER NOT NULL DEFAULT 1")
        ensure_column(conn, "audit_logs", "actor_id", "INTEGER")
        ensure_column(conn, "audit_logs", "actor_role", "TEXT DEFAULT ''")
        ensure_column(conn, "audit_logs", "action", "TEXT DEFAULT ''")
        ensure_column(conn, "audit_logs", "table_name", "TEXT DEFAULT ''")
        ensure_column(conn, "audit_logs", "record_id", "TEXT DEFAULT ''")
        ensure_column(conn, "audit_logs", "summary", "TEXT DEFAULT ''")
        ensure_column(conn, "audit_logs", "before_data", "TEXT DEFAULT ''")
        ensure_column(conn, "audit_logs", "after_data", "TEXT DEFAULT ''")
        ensure_column(conn, "audit_logs", "created_at", "TEXT DEFAULT ''")
        for index_sql in (
            "CREATE INDEX IF NOT EXISTS idx_work_log_sessions_company_date_user ON work_log_sessions(company_id, work_date, user_id)",
            "CREATE INDEX IF NOT EXISTS idx_work_log_sessions_open ON work_log_sessions(company_id, work_date, status)",
            "CREATE INDEX IF NOT EXISTS idx_work_log_corrections_company_changed ON work_log_corrections(company_id, changed_at)",
            "CREATE INDEX IF NOT EXISTS idx_work_day_deletions_company_date ON work_day_deletions(company_id, work_date)",
            "CREATE INDEX IF NOT EXISTS idx_vehicle_expenses_company_month ON vehicle_expenses(company_id, expense_date)",
            "CREATE INDEX IF NOT EXISTS idx_company_expenses_company_month ON company_expenses(company_id, expense_date)",
            "CREATE INDEX IF NOT EXISTS idx_depot_rate_settings_lookup ON depot_rate_settings(company_id, depot_id, effective_start)",
        ):
            conn.execute(index_sql)
        schema_now = datetime.now().isoformat(timespec="seconds")
        conn.execute(
            """INSERT INTO app_metadata(meta_key, meta_value, updated_at)
               VALUES (?, ?, ?)
               ON CONFLICT(meta_key) DO UPDATE SET meta_value=excluded.meta_value, updated_at=excluded.updated_at""",
            ("schema_version", SCHEMA_VERSION, schema_now),
        )
        conn.execute("INSERT OR IGNORE INTO companies(id, name, created_at) VALUES (1, ?, ?)", ("SPARKLE DRIVE", datetime.now().isoformat(timespec="seconds")))
        conn.execute("UPDATE companies SET status='利用中' WHERE COALESCE(status, '')=''")
        seed_feature_prices(conn)
        for company in conn.execute("SELECT id FROM companies").fetchall():
            seed_company_features(conn, company["id"], 1)
            conn.execute(
                "INSERT OR IGNORE INTO company_vehicle_rates(company_id, daily_fee, monthly_fee, updated_at) VALUES (?, 1500, 30000, ?)",
                (company["id"], datetime.now().isoformat(timespec="seconds")),
            )
        admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
        if not admin:
            now = datetime.now().isoformat(timespec="seconds")
            cur = conn.execute(
                "INSERT INTO users(name, username, password_hash, role, phone, vehicle, created_at) VALUES (?,?,?,?,?,?,?)",
                ("管理者", "admin", hash_password(DEFAULT_ADMIN_PASSWORD), "admin", "", "", now),
            )
            conn.execute("INSERT OR IGNORE INTO rates(user_id) VALUES (?)", (cur.lastrowid,))
            samples = [
                ("佐藤 花子", "sato", DEFAULT_MEMBER_PASSWORD, "070-1111-2222", "軽バン 12-34"),
                ("鈴木 一郎", "suzuki", DEFAULT_MEMBER_PASSWORD, "070-3333-4444", "軽バン 56-78"),
            ]
            for name, username, password, phone, vehicle in samples:
                cur = conn.execute(
                    "INSERT INTO users(name, username, password_hash, role, phone, vehicle, created_at) VALUES (?,?,?,?,?,?,?)",
                    (name, username, hash_password(password), "member", phone, vehicle, now),
                )
                conn.execute("INSERT OR IGNORE INTO rates(user_id) VALUES (?)", (cur.lastrowid,))
        vehicle_seed_rows = conn.execute(
            "SELECT id, company_id, vehicle, last_vehicle_id FROM users WHERE COALESCE(vehicle, '')<>''"
        ).fetchall()
        now = datetime.now().isoformat(timespec="seconds")
        for member in vehicle_seed_rows:
            vehicle_name = (member["vehicle"] or "").strip()
            if not vehicle_name:
                continue
            conn.execute(
                "INSERT OR IGNORE INTO vehicles(company_id, name, created_at, updated_at) VALUES (?, ?, ?, ?)",
                (member["company_id"], vehicle_name, now, now),
            )
            vehicle = conn.execute(
                "SELECT id FROM vehicles WHERE company_id=? AND name=?",
                (member["company_id"], vehicle_name),
            ).fetchone()
            if vehicle and not member["last_vehicle_id"]:
                conn.execute("UPDATE users SET last_vehicle_id=? WHERE id=? AND company_id=?", (vehicle["id"], member["id"], member["company_id"]))
        conn.execute("INSERT OR IGNORE INTO vehicle_rates(id, daily_fee, monthly_fee, company_id) VALUES (1, 1500, 30000, 1)")
        conn.execute("INSERT OR IGNORE INTO company_vehicle_rates(company_id, daily_fee, monthly_fee, updated_at) VALUES (1, 1500, 30000, ?)", (now,))
        now = datetime.now().isoformat(timespec="seconds")
        cur = conn.execute("INSERT OR IGNORE INTO depots(company_id, name, created_at) VALUES (?, ?, ?)", (1, "未設定拠点", now))
        depot = conn.execute("SELECT id FROM depots WHERE company_id=? AND name = ?", (1, "未設定拠点")).fetchone()
        seed_areas = {
            "高松": ["高松1丁目", "高松2丁目", "高松3丁目", "高松4丁目", "高松5丁目", "高松6丁目"],
            "土支田": ["土支田2丁目", "土支田3丁目"],
        }
        for district_name, town_names in seed_areas.items():
            conn.execute("INSERT OR IGNORE INTO districts(company_id, depot_id, name, created_at) VALUES (?, ?, ?, ?)", (1, depot["id"], district_name, now))
            district = conn.execute("SELECT id FROM districts WHERE company_id=? AND depot_id = ? AND name = ?", (1, depot["id"], district_name)).fetchone()
            for town_name in town_names:
                conn.execute("INSERT OR IGNORE INTO towns(company_id, district_id, name, created_at) VALUES (?, ?, ?, ?)", (1, district["id"], town_name, now))
        conn.commit()
    ensure_vehicle_issue_schema()


def schema_already_initialized():
    if not USE_POSTGRES and not DB_PATH.exists():
        return False
    try:
        with db() as conn:
            row = conn.execute(
                "SELECT meta_value FROM app_metadata WHERE meta_key=?",
                ("schema_version",),
            ).fetchone()
        return row_value(row, "meta_value", "") == SCHEMA_VERSION
    except Exception:
        if current_schema_shape_ok():
            mark_schema_initialized()
            return True
        return False


def init_db():
    global DB_INIT_DONE, DB_INIT_ERROR, DB_INIT_LAST_ATTEMPT
    if DB_INIT_DONE:
        return
    with DB_INIT_LOCK:
        if DB_INIT_DONE:
            return
        DB_INIT_LAST_ATTEMPT = monotonic()
        try:
            if schema_already_initialized():
                DB_INIT_DONE = True
                DB_INIT_ERROR = None
                return
            _init_db_schema()
        except Exception as exc:
            DB_INIT_ERROR = exc
            raise
        else:
            DB_INIT_DONE = True
            DB_INIT_ERROR = None


def safe_startup_init_db():
    try:
        init_db()
    except Exception:
        logger.exception("Database initialization failed; continuing startup so Render can bind the port.")


@app.on_event("startup")
def startup():
    if USE_POSTGRES and DB_INIT_ON_STARTUP:
        Thread(target=safe_startup_init_db, daemon=True).start()
    elif not USE_POSTGRES:
        safe_startup_init_db()


@app.get("/healthz")
def healthz():
    return {
        "ok": True,
        "database": "postgresql" if USE_POSTGRES else "sqlite",
        "db_initialized": DB_INIT_DONE,
        "db_init_error": DB_INIT_ERROR.__class__.__name__ if DB_INIT_ERROR else "",
    }


@app.get("/health")
def health():
    return healthz()


@app.get("/readyz")
def readyz():
    return healthz()


def current_user(request: Request):
    cached = getattr(request.state, "user", None)
    if cached is not None:
        return cached
    raw = request.cookies.get("session")
    if not raw:
        return None
    try:
        user_id = serializer.loads(raw)
    except BadSignature:
        return None
    user = query_one(
        """SELECT u.*, c.name AS company_name, c.code AS company_code, c.active AS company_active
           FROM users u
           LEFT JOIN companies c ON c.id=u.company_id
           WHERE u.id = ? AND u.active = 1""",
        (user_id,),
    )
    if user and row_value(user, "company_active", 1) == 0:
        user = None
    request.state.user = user
    return user


def require_user(request: Request):
    user = current_user(request)
    if not user:
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    return user


def require_admin(user=Depends(require_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="管理者だけが利用できます")
    return user


def require_platform_admin(user=Depends(require_admin)):
    if not is_platform_admin(user):
        raise HTTPException(status_code=403, detail="全体管理者だけが利用できます")
    return user


def image_path_authorized(user, path: str):
    path = (path or "").strip()
    if not path or not (is_remote_image_path(path) or allowed_local_image_path(path)):
        return False
    company_id = company_id_for(user)
    storage_path = storage_path_from_url(path)
    user_file_marker = f"user{user['id']}_"
    if user["role"] == "member" and user_file_marker in path and (allowed_local_image_path(path) or storage_path):
        return True
    matches = [path]
    if storage_path:
        matches.append(storage_path)
    placeholders = ",".join("?" for _ in matches)
    role_clause = "" if user["role"] == "admin" else " AND user_id=?"
    params = [company_id, *matches]
    if user["role"] != "admin":
        params.append(user["id"])
    for table in ("inspection_sheets", "inspection_slips"):
        row = query_one(
            f"""SELECT id FROM {table}
                WHERE company_id=? AND (file_path IN ({placeholders}) OR public_url IN ({placeholders}) OR storage_path IN ({placeholders}))
                {role_clause}
                LIMIT 1""",
            [company_id, *matches, *matches, *matches] + ([user["id"]] if user["role"] != "admin" else []),
        )
        if row:
            return True
    delivery_clause = "" if user["role"] == "admin" else " AND user_id=?"
    delivery_row = query_one(
        f"""SELECT id FROM deliveries
            WHERE company_id=? AND inspection_sheet_path IN ({placeholders})
            {delivery_clause}
            LIMIT 1""",
        params if user["role"] != "admin" else [company_id, *matches],
    )
    if delivery_row:
        return True
    if user["role"] == "admin":
        if is_platform_admin(user):
            support_row = query_one(
                f"""SELECT id FROM support_requests
                    WHERE image_path IN ({placeholders})
                    LIMIT 1""",
                matches,
            )
            if support_row:
                return True
        support_row = query_one(
            f"""SELECT id FROM support_requests
                WHERE company_id=? AND image_path IN ({placeholders})
                LIMIT 1""",
            [company_id, *matches],
        )
        if support_row:
            return True
    return False


@app.get("/inspection-file", response_class=HTMLResponse)
def inspection_file(request: Request, path: str = "", user=Depends(require_user)):
    if not image_path_authorized(user, path):
        return render(
            request,
            "image_missing.html",
            {
                "path": path,
                "message": "画像が見つかりません。再アップロードしてください。",
            },
        )
    if is_remote_image_path(path):
        try:
            head = UrlRequest(path, method="HEAD")
            with urlopen(head, timeout=DB_CONNECT_TIMEOUT) as response:
                if response.status >= 400:
                    raise HTTPError(path, response.status, "remote image missing", response.headers, None)
        except HTTPError as exc:
            if exc.code in {403, 404, 410}:
                return render(
                    request,
                    "image_missing.html",
                    {
                        "path": path,
                        "message": "画像が見つかりません。再アップロードしてください。",
                    },
                )
        except (URLError, TimeoutError, OSError):
            return render(
                request,
                "image_missing.html",
                {
                    "path": path,
                    "message": "画像の確認に失敗しました。時間を置いて再度確認するか、再アップロードしてください。",
                },
            )
        return RedirectResponse(path, status_code=303)
    if local_image_exists(path):
        return RedirectResponse(path, status_code=303)
    return render(
        request,
        "image_missing.html",
        {
            "path": path,
            "message": "画像ファイルが見つかりません。再アップロードしてください。",
        },
    )


def render(request: Request, name: str, context: dict):
    user = current_user(request)
    features = load_company_features(company_id_for(user)) if user else {feature["key"]: True for feature in FEATURE_CATALOG}
    context.update(
        {
            "request": request,
            "user": user,
            "today": app_today().isoformat(),
            "app_name": APP_NAME,
            "company_name": company_name_for(user) if user else APP_NAME,
            "features": features,
            "feature_enabled": lambda key: is_feature_enabled(features, key),
            "can_manage_companies": can_manage_companies(user) if user else False,
            "platform_admin": is_platform_admin(user) if user else False,
        }
    )
    return templates.TemplateResponse(name, context)


def month_bounds(ym: Optional[str]):
    if not ym:
        base = app_today().replace(day=1)
    else:
        try:
            base = datetime.strptime(ym, "%Y-%m").date().replace(day=1)
        except ValueError:
            base = app_today().replace(day=1)
    if base.month == 12:
        next_month = base.replace(year=base.year + 1, month=1)
    else:
        next_month = base.replace(month=base.month + 1)
    return base, next_month - timedelta(days=1)


def calendar_days_for_month(start: date):
    _, last_day = calendar.monthrange(start.year, start.month)
    return [start.replace(day=day) for day in range(1, last_day + 1)]


def active_towns(company_id: Optional[int] = None):
    where = "t.active=1 AND d.active=1"
    params = []
    if company_id is not None:
        where += " AND t.company_id=? AND d.company_id=? AND COALESCE(p.company_id, t.company_id)=?"
        params.extend([company_id, company_id, company_id])
    return query_all(
        f"""SELECT t.id, t.name, d.id AS district_id, d.name AS district_name, p.name AS depot_name
           FROM towns t
           JOIN districts d ON d.id = t.district_id AND d.company_id = t.company_id
           LEFT JOIN depots p ON p.id = d.depot_id AND p.company_id = d.company_id
           WHERE {where}
           ORDER BY p.name, d.name, t.name""",
        params,
    )


def attach_shift_areas(shifts):
    items = [dict(row) for row in shifts]
    if not items:
        return []
    shift_ids = [item["id"] for item in items]
    placeholders = ",".join("?" for _ in shift_ids)
    area_rows = query_all(
        f"""SELECT st.shift_id, d.name AS district_name, t.name AS town_name
            FROM shift_towns st
            JOIN shifts s ON s.id = st.shift_id AND s.company_id = st.company_id
            JOIN towns t ON t.id = st.town_id AND t.company_id = st.company_id
            JOIN districts d ON d.id = t.district_id AND d.company_id = st.company_id
            WHERE st.shift_id IN ({placeholders})
              AND COALESCE(st.active,1)=1
            ORDER BY d.name, t.name""",
        shift_ids,
    )
    area_map = {}
    short_map = {}
    for area in area_rows:
        area_map.setdefault(area["shift_id"], []).append(area["town_name"])
        short_map.setdefault(area["shift_id"], []).append((area["district_name"], area["town_name"]))
    for item in items:
        label_pairs = parse_area_label(item.get("area_label", ""))
        if label_pairs:
            areas = [town_name for _, town_name in label_pairs]
        else:
            areas = area_map.get(item["id"], [])
        if item["status"] == "休み":
            areas = []
        item["areas"] = areas
        item["area_text"] = "、".join(areas)
        short_pairs = label_pairs or short_map.get(item["id"], [])
        item["area_short_text"] = shorten_areas(short_pairs) if item["status"] != "休み" else ""
    return items


def group_rows_by_depot(rows, name_key: str = "depot_name", default_name: str = "未設定拠点"):
    grouped = []
    index = {}
    for row in rows:
        item = dict(row)
        depot_name = row_value(item, name_key, "") or default_name
        if depot_name not in index:
            index[depot_name] = {"name": depot_name, "items": []}
            grouped.append(index[depot_name])
        index[depot_name]["items"].append(item)
    return grouped


def parse_area_label(area_label):
    pairs = []
    for part in (area_label or "").split("・"):
        if "/" not in part:
            continue
        district_name, town_name = [value.strip() for value in part.split("/", 1)]
        if district_name and town_name:
            pairs.append((district_name, town_name))
    return pairs


def shorten_areas(area_pairs):
    grouped = {}
    for district_name, town_name in area_pairs:
        prefix = district_name[:1]
        match = re.search(r"(\d+|[０-９]+)\s*丁目", town_name)
        number = match.group(1).translate(str.maketrans("０１２３４５６７８９", "0123456789")) if match else town_name
        grouped.setdefault(prefix, []).append(number)
    parts = []
    for prefix, numbers in grouped.items():
        unique_numbers = []
        for number in numbers:
            if number not in unique_numbers:
                unique_numbers.append(number)
        parts.append(prefix + ".".join(unique_numbers))
    return " / ".join(parts)


def normalize_text(value):
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def compact_area_key(value):
    text = normalize_text(value).translate(FULLWIDTH_DIGITS)
    text = re.sub(r"\s+", "", text)
    return text.replace("丁目", "")


def normalize_import_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value).translate(FULLWIDTH_DIGITS)
    if not text:
        return ""
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    if re.fullmatch(r"\d{8}", text):
        try:
            return datetime.strptime(text, "%Y%m%d").date().isoformat()
        except ValueError:
            pass
    return text


def normalize_shift_status(status_value, bin_value=""):
    status = normalize_text(status_value)
    bin_label = normalize_text(bin_value)
    aliases = {"休": "休み", "休日": "休み"}
    status = aliases.get(status, status)
    if status in {"1", "2", "3"}:
        status = f"{status}便"
    if bin_label in {"1", "2", "3"}:
        bin_label = f"{bin_label}便"
    if status in {"", "出勤"} and bin_label in {"1便", "2便", "3便"}:
        return bin_label
    return status


def resolve_import_member(member_name, members):
    name = normalize_text(member_name)
    if not name:
        return None, "名前が空です"
    exact = [member for member in members if normalize_text(member["name"]) == name]
    if len(exact) == 1:
        return exact[0], ""
    partial = [member for member in members if name in normalize_text(member["name"])]
    if len(partial) == 1:
        return partial[0], ""
    if len(partial) > 1:
        return None, f"名前「{name}」に一致するメンバーが複数います"
    return None, f"メンバー「{name}」が見つかりません"


def town_number(town_name):
    match = re.search(r"(\d+)\s*丁目?", normalize_text(town_name).translate(FULLWIDTH_DIGITS))
    return match.group(1) if match else ""


def match_area_town_ids(area_text, towns):
    text = normalize_text(area_text)
    if not text:
        return []
    compact_text = compact_area_key(text)
    matched_ids = []

    def add_town(town_id):
        if town_id not in matched_ids:
            matched_ids.append(town_id)

    for town in towns:
        name_key = compact_area_key(town["name"])
        district_key = compact_area_key(town["district_name"])
        full_key = compact_area_key(f"{town['district_name']}{town['name']}")
        if name_key and (name_key in compact_text or full_key in compact_text):
            add_town(town["id"])
        elif district_key and compact_text == district_key:
            add_town(town["id"])

    normalized = normalize_text(text).translate(FULLWIDTH_DIGITS)
    for match in re.finditer(r"([^\d,、/／\s・･]+)(\d+(?:[・･,、/／\s]+\d+)*)", normalized):
        prefix = compact_area_key(match.group(1))
        numbers = re.findall(r"\d+", match.group(2))
        for number in numbers:
            for town in towns:
                if town_number(town["name"]) != number:
                    continue
                district_key = compact_area_key(town["district_name"])
                name_key = compact_area_key(town["name"])
                if prefix and (
                    district_key.startswith(prefix)
                    or name_key.startswith(prefix)
                    or prefix in district_key
                    or prefix in name_key
                ):
                    add_town(town["id"])

    tokens = [compact_area_key(token) for token in re.split(r"[、,/／\s]+", text) if compact_area_key(token)]
    for token in tokens:
        for town in towns:
            if token in {compact_area_key(town["name"]), compact_area_key(f"{town['district_name']}{town['name']}")}:
                add_town(town["id"])
    return matched_ids


def area_preview(town_ids, towns_by_id):
    labels = []
    for town_id in town_ids:
        town = towns_by_id.get(int(town_id))
        if town:
            labels.append(f"{town['district_name']} / {town['name']}")
    return "、".join(labels)


def load_csv_shift_rows(data):
    text = ""
    for encoding in ("utf-8-sig", "cp932"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        raise HTTPException(status_code=400, detail="CSVの文字コードを読み取れませんでした")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="ヘッダー行が見つかりません")
    reader.fieldnames = [normalize_text(header) for header in reader.fieldnames]
    rows = []
    for row in reader:
        normalized = {normalize_text(key): value for key, value in row.items() if key}
        if any(normalize_text(value) for value in normalized.values()):
            rows.append(normalized)
    return reader.fieldnames, rows


def load_excel_shift_rows(data):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Excel取り込みにはopenpyxlが必要です") from exc
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [normalize_text(value) for value in next(rows_iter)]
    except StopIteration as exc:
        raise HTTPException(status_code=400, detail="Excelにデータがありません") from exc
    rows = []
    for values in rows_iter:
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers)) if headers[index]}
        if any(normalize_text(value) for value in row.values()):
            rows.append(row)
    return headers, rows


async def load_shift_import_rows(file: UploadFile):
    filename = file.filename or ""
    suffix = Path(filename).suffix.lower()
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="ファイルが空です")
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="ファイルサイズは5MB以内にしてください")
    if suffix == ".csv":
        headers, rows = load_csv_shift_rows(data)
    elif suffix in {".xlsx", ".xlsm"}:
        headers, rows = load_excel_shift_rows(data)
    elif suffix == ".xls":
        raise HTTPException(status_code=400, detail="Excelは.xlsx形式で保存してアップロードしてください")
    else:
        raise HTTPException(status_code=400, detail="CSVまたはExcel(.xlsx)をアップロードしてください")

    missing = [header for header in SHIFT_IMPORT_HEADERS if header not in headers]
    if missing:
        raise HTTPException(status_code=400, detail="不足している列: " + "、".join(missing))
    if not rows:
        raise HTTPException(status_code=400, detail="登録できる行がありません")
    if len(rows) > SHIFT_IMPORT_LIMIT:
        raise HTTPException(status_code=400, detail=f"一度に取り込めるシフトは{SHIFT_IMPORT_LIMIT}件までです")
    return rows


def current_shift_conflicts(rows, company_id=1):
    member_ids = sorted({int(row["member_id"]) for row in rows if row.get("member_id")})
    dates = sorted({row["shift_date"] for row in rows if row.get("shift_date")})
    if not member_ids or not dates:
        return set()
    placeholders = ",".join("?" for _ in member_ids)
    existing = query_all(
        f"""SELECT user_id, shift_date FROM shifts
            WHERE company_id=? AND shift_date BETWEEN ? AND ? AND user_id IN ({placeholders})""",
        [company_id, dates[0], dates[-1]] + member_ids,
    )
    return {(int(row["user_id"]), row["shift_date"]) for row in existing}


def build_shift_import_preview(raw_rows, company_id=1):
    members = query_all("SELECT id, name FROM users WHERE role='member' AND active=1 AND COALESCE(purged,0)=0 AND company_id=? ORDER BY name LIMIT 200", (company_id,))
    towns = active_towns(company_id)
    towns_by_id = {int(town["id"]): town for town in towns}
    preview_rows = []
    for index, raw in enumerate(raw_rows, start=2):
        errors = []
        shift_date = normalize_import_date(raw.get("日付"))
        if not shift_date:
            errors.append("日付が空です")
        else:
            try:
                datetime.strptime(shift_date, "%Y-%m-%d")
            except ValueError:
                errors.append("日付はYYYY-MM-DD形式で入力してください")

        member, member_error = resolve_import_member(raw.get("名前"), members)
        if member_error:
            errors.append(member_error)

        status = normalize_shift_status(raw.get("出勤区分"), raw.get("便"))
        if status not in SHIFT_ALLOWED_STATUSES:
            errors.append("出勤区分は「出勤、1便、2便、3便、休み」のいずれかにしてください")

        area_text = normalize_text(raw.get("配達エリア"))
        town_ids = []
        if status != "休み":
            town_ids = match_area_town_ids(area_text, towns)
            if not town_ids:
                errors.append("配達エリアが見つかりません")

        preview_rows.append(
            {
                "row_number": index,
                "shift_date": shift_date,
                "member_id": member["id"] if member else None,
                "member_name": normalize_text(raw.get("名前")),
                "resolved_member_name": member["name"] if member else "",
                "status": status,
                "bin_label": normalize_text(raw.get("便")),
                "area_text": area_text,
                "town_ids": town_ids,
                "area_preview": area_preview(town_ids, towns_by_id),
                "note": normalize_text(raw.get("備考")),
                "errors": errors,
                "conflict": False,
                "company_id": company_id,
            }
        )

    valid_rows = [row for row in preview_rows if not row["errors"]]
    conflicts = current_shift_conflicts(valid_rows, company_id)
    for row in preview_rows:
        if row["member_id"] and row["shift_date"] and (int(row["member_id"]), row["shift_date"]) in conflicts:
            row["conflict"] = True
    payload_rows = [{key: value for key, value in row.items() if key != "errors"} for row in valid_rows]
    return {
        "rows": preview_rows,
        "payload": json.dumps(payload_rows, ensure_ascii=False),
        "has_errors": any(row["errors"] for row in preview_rows),
        "has_conflicts": any(row["conflict"] for row in preview_rows),
    }


def get_town_rows_by_ids(town_ids, company_id: Optional[int] = None):
    town_ids = [int(town_id) for town_id in town_ids]
    if not town_ids:
        return []
    placeholders = ",".join("?" for _ in town_ids)
    company_filter = ""
    params = list(town_ids)
    if company_id is not None:
        company_filter = " AND t.company_id=? AND d.company_id=?"
        params.extend([company_id, company_id])
    return query_all(
        f"""SELECT t.id, t.name, d.id AS district_id, d.name AS district_name
            FROM towns t JOIN districts d ON d.id=t.district_id
            WHERE t.id IN ({placeholders}) AND t.active=1 AND d.active=1 {company_filter}
            ORDER BY d.name, t.name""",
        params,
    )


def upsert_shift_with_towns(conn, member_id, shift_date, status, town_rows, note="", company_id=1, actor=None):
    first_town = town_rows[0] if town_rows else None
    area_label = "・".join(f"{town['district_name']} / {town['name']}" for town in town_rows)
    existing = conn.execute("SELECT * FROM shifts WHERE user_id=? AND shift_date=? AND company_id=?", (member_id, shift_date, company_id)).fetchone()
    conn.execute(
        """INSERT INTO shifts(company_id, user_id, shift_date, status, start_time, end_time, note, district_id, town_id, area_label, decided, updated_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
           ON CONFLICT(user_id, shift_date) DO UPDATE SET status=excluded.status, start_time=excluded.start_time,
           end_time=excluded.end_time, note=excluded.note, district_id=excluded.district_id, town_id=excluded.town_id,
           area_label=excluded.area_label, decided=excluded.decided, updated_at=excluded.updated_at, company_id=excluded.company_id""",
        (
            company_id,
            member_id,
            shift_date,
            status,
            "",
            "",
            note or "",
            first_town["district_id"] if first_town else None,
            first_town["id"] if first_town else None,
            area_label,
            1,
            app_now().isoformat(timespec="seconds"),
        ),
    )
    shift = conn.execute("SELECT * FROM shifts WHERE user_id=? AND shift_date=? AND company_id=?", (member_id, shift_date, company_id)).fetchone()
    conn.execute("UPDATE shift_towns SET active=0 WHERE shift_id=? AND company_id=?", (shift["id"], company_id))
    if status != "休み":
        for town in town_rows:
            conn.execute(
                """INSERT INTO shift_towns(company_id, shift_id, town_id, active) VALUES (?, ?, ?, 1)
                   ON CONFLICT(shift_id, town_id) DO UPDATE SET company_id=excluded.company_id, active=1""",
                (company_id, shift["id"], town["id"]),
            )
    audit_log(
        company_id,
        actor or {"id": member_id, "role": ""},
        "shift.update" if existing else "shift.create",
        "shifts",
        shift["id"],
        "Shift saved",
        before=row_to_dict(existing),
        after=row_to_dict(shift),
        conn=conn,
    )


def upsert_shift_record(member_id, shift_date, status, town_ids, note="", company_id=1, actor=None):
    if status not in SHIFT_ALLOWED_STATUSES:
        raise HTTPException(status_code=400, detail="出勤区分が正しくありません")
    if status != "休み" and not town_ids:
        raise HTTPException(status_code=400, detail="休み以外は町会・配達エリアを1つ以上選択してください")
    member = query_one("SELECT id FROM users WHERE id=? AND role='member' AND company_id=? AND active=1 AND COALESCE(purged,0)=0", (member_id, company_id))
    if not member:
        raise HTTPException(status_code=404, detail="メンバーが見つかりません")
    town_rows = get_town_rows_by_ids(town_ids, company_id)
    if status != "休み" and len(town_rows) != len(set(int(town_id) for town_id in town_ids)):
        raise HTTPException(status_code=400, detail="選択された町会・配達エリアが正しくありません")
    with db() as conn:
        upsert_shift_with_towns(conn, member_id, shift_date, status, town_rows, note, company_id, actor)
        conn.commit()


def shifts_by_date(shifts):
    grouped = {}
    for shift in shifts:
        grouped.setdefault(shift["shift_date"], []).append(shift)
    return grouped


def calc_reward(row, rates, vehicle_rates=None, monthly_mode=False):
    if not row or not rates:
        return 0
    reward = (
        row["completed"] * rates["delivery_unit"]
        + row["transfer"] * rates["transfer_unit"]
        + row["night"] * rates["night_unit"]
        + row["pickup"] * rates["pickup_unit"]
        + row["large"] * rates["large_unit"]
    )
    rental_type = rates["vehicle_rental_type"] if "vehicle_rental_type" in rates.keys() else "none"
    if rental_type == "daily":
        reward -= rates["vehicle_daily_fee"]
    if monthly_mode and rental_type == "monthly":
        reward -= rates["vehicle_monthly_fee"]
    return reward


def calc_reward_gross(row, rates):
    if not row or not rates:
        return 0
    return (
        row["completed"] * rates["delivery_unit"]
        + row["transfer"] * rates["transfer_unit"]
        + row["night"] * rates["night_unit"]
        + row["pickup"] * rates["pickup_unit"]
        + row["large"] * rates["large_unit"]
    )


def monthly_reward_summary(user_id: int, target_month: date, company_id: Optional[int] = None):
    start = target_month.replace(day=1)
    if start.month == 12:
        next_month = start.replace(year=start.year + 1, month=1)
    else:
        next_month = start.replace(month=start.month + 1)
    end = next_month - timedelta(days=1)
    company_filter = ""
    params = [user_id, start.isoformat(), end.isoformat()]
    if company_id is not None:
        company_filter = " AND company_id=?"
        params.append(company_id)
    summary = query_one(
        f"""SELECT
              COALESCE(SUM(completed), 0) AS completed_total,
              COALESCE(SUM(transfer), 0) AS transfer_total,
              COALESCE(SUM(night), 0) AS night_total,
              COALESCE(SUM(pickup), 0) AS pickup_total,
              COALESCE(SUM(large), 0) AS large_total,
              COUNT(*) AS delivery_days
           FROM deliveries
           WHERE user_id=? AND work_date BETWEEN ? AND ? AND COALESCE(is_deleted,0)=0{company_filter}""",
        params,
    )
    if company_id is not None:
        rates = query_one("SELECT * FROM rates WHERE user_id=? AND company_id=?", (user_id, company_id))
    else:
        rates = query_one("SELECT * FROM rates WHERE user_id=?", (user_id,))
    if not rates:
        return {"gross_total": 0, "net_total": 0, "vehicle_deduction": 0, "completed_total": 0, "rows": []}
    gross_total = (
        summary["completed_total"] * rates["delivery_unit"]
        + summary["transfer_total"] * rates["transfer_unit"]
        + summary["night_total"] * rates["night_unit"]
        + summary["pickup_total"] * rates["pickup_unit"]
        + summary["large_total"] * rates["large_unit"]
    )
    rental_type = rates["vehicle_rental_type"] if rates and "vehicle_rental_type" in rates.keys() else "none"
    vehicle_deduction = 0
    if rental_type == "daily":
        vehicle_deduction = summary["delivery_days"] * rates["vehicle_daily_fee"]
    elif rental_type == "monthly" and summary["delivery_days"]:
        vehicle_deduction = rates["vehicle_monthly_fee"]
    net_total = gross_total - vehicle_deduction
    return {
        "gross_total": gross_total,
        "net_total": net_total,
        "vehicle_deduction": vehicle_deduction,
        "completed_total": summary["completed_total"],
        "rows": [],
    }


def upcoming_member_shifts(user_id: int, today_s: str, company_id: Optional[int] = None):
    company_filter = ""
    params = [user_id, today_s]
    if company_id is not None:
        company_filter = " AND s.company_id=?"
        params.append(company_id)
    rows = query_all(
        f"""SELECT s.*, d.name AS district_name, t.name AS town_name
           FROM shifts s
           LEFT JOIN districts d ON d.id=s.district_id AND d.company_id=s.company_id
           LEFT JOIN towns t ON t.id=s.town_id AND t.company_id=s.company_id
           WHERE s.user_id=? AND s.shift_date>=? AND s.decided=1
           {company_filter}
           ORDER BY s.shift_date LIMIT 3""",
        params,
    )
    return attach_shift_areas(rows)


def today_work_state(user_id: int, company_id: int, today_s: str):
    sessions = query_all(
        """SELECT * FROM work_log_sessions
           WHERE user_id=? AND company_id=? AND work_date=? AND COALESCE(is_deleted,0)=0
           ORDER BY session_no, id""",
        (user_id, company_id, today_s),
    )
    if sessions:
        open_session = None
        for session in sessions:
            if row_value(session, "status") == "working":
                open_session = session
        if open_session:
            return {"status": "working", "label": f"出勤中（{open_session['session_no']}回目）", "start_log": None, "end_log": None, "sessions": sessions, "open_session": open_session}
        return {"status": "finished", "label": f"退勤済（{len(sessions)}回）", "start_log": None, "end_log": None, "sessions": sessions, "open_session": None}
    start_log = query_one(
        """SELECT * FROM work_logs
           WHERE user_id=? AND company_id=? AND work_date=? AND log_type='start' AND COALESCE(is_deleted,0)=0
           ORDER BY logged_at DESC LIMIT 1""",
        (user_id, company_id, today_s),
    )
    end_log = query_one(
        """SELECT * FROM work_logs
           WHERE user_id=? AND company_id=? AND work_date=? AND log_type='end' AND COALESCE(is_deleted,0)=0
           ORDER BY logged_at DESC LIMIT 1""",
        (user_id, company_id, today_s),
    )
    if end_log and (not start_log or str(end_log["logged_at"]) >= str(start_log["logged_at"])):
        return {"status": "finished", "label": "退勤済", "start_log": start_log, "end_log": end_log}
    if start_log:
        return {"status": "working", "label": "出勤中", "start_log": start_log, "end_log": end_log}
    return {"status": "not_started", "label": "未出勤", "start_log": start_log, "end_log": end_log}


def latest_odometer_for_user(user_id: int, company_id: int):
    row = query_one(
        """SELECT odometer FROM work_logs
           WHERE user_id=? AND company_id=? AND COALESCE(odometer, 0)>0 AND COALESCE(is_deleted,0)=0
           ORDER BY work_date DESC, logged_at DESC LIMIT 1""",
        (user_id, company_id),
    )
    return row["odometer"] if row else 0


def int_value(value, default=0):
    try:
        return int(value or default)
    except (TypeError, ValueError):
        return default


def money_value(value, default=0.0):
    try:
        return float(value or default)
    except (TypeError, ValueError):
        return default


def vehicle_rows_for_company(company_id: int, include_inactive: bool = False):
    condition = "" if include_inactive else " AND active=1"
    return query_all(
        f"SELECT * FROM vehicles WHERE company_id=?{condition} ORDER BY active DESC, name LIMIT 200",
        (company_id,),
    )


async def save_receipt_upload(upload: Optional[UploadFile], company_id: int, prefix: str):
    if not upload or not upload.filename:
        return ""
    content_type = upload.content_type or ""
    if content_type and not content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="領収書画像は画像ファイルを選択してください")
    data = await upload.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="領収書画像は10MB以内にしてください")
    suffix = safe_upload_name(upload.filename)
    filename = f"company{company_id}_{prefix}_{app_now().strftime('%Y%m%d%H%M%S%f')}{suffix}"
    public_path, _ = save_inspection_image(data, filename, content_type or "image/jpeg", "receipts", RECEIPT_UPLOAD_DIR)
    return public_path


def depot_rows_for_company(company_id: int):
    return query_all("SELECT * FROM depots WHERE company_id=? AND active=1 ORDER BY name LIMIT 200", (company_id,))


def member_rows_for_company(company_id: int):
    return query_all(
        """SELECT id, name, username, vehicle, last_vehicle_id
           FROM users
           WHERE company_id=? AND role='member' AND active=1 AND COALESCE(purged,0)=0
           ORDER BY id LIMIT 300""",
        (company_id,),
    )


def vehicle_expense_month_bounds(target_month: Optional[str] = None):
    start, end = month_bounds(target_month)
    return start.isoformat(), end.isoformat()


def vehicle_dashboard_rows(company_id: int, target_month: Optional[str] = None):
    month_start, month_end = vehicle_expense_month_bounds(target_month)
    vehicles = []
    issue_statuses = {"対応済", "修理済", "handled", "repaired"}
    for vehicle in vehicle_rows_for_company(company_id, include_inactive=True):
        item = dict(vehicle)
        item["alert"] = vehicle_alert(vehicle)
        item["assigned_members"] = query_all(
            """SELECT id, name FROM users
               WHERE company_id=? AND role='member' AND active=1 AND COALESCE(purged,0)=0
                 AND (last_vehicle_id=? OR vehicle=?)
               ORDER BY name LIMIT 8""",
            (company_id, vehicle["id"], row_value(vehicle, "name", "")),
        )
        expense_row = query_one(
            """SELECT COALESCE(SUM(amount), 0) AS total
               FROM vehicle_expenses
               WHERE company_id=? AND vehicle_id=? AND expense_date BETWEEN ? AND ?""",
            (company_id, vehicle["id"], month_start, month_end),
        )
        item["month_vehicle_cost"] = money_value(row_value(expense_row, "total", 0))
        issue_rows = query_all(
            """SELECT id FROM vehicle_issues
               WHERE company_id=? AND (vehicle_name=? OR vehicle_name=?)
               ORDER BY created_at DESC LIMIT 20""",
            (company_id, row_value(vehicle, "name", ""), row_value(vehicle, "plate_number", "")),
        )
        open_issue_count = 0
        if issue_rows:
            status_rows = query_all(
                """SELECT status FROM vehicle_issues
                   WHERE company_id=? AND (vehicle_name=? OR vehicle_name=?)
                   ORDER BY created_at DESC LIMIT 20""",
                (company_id, row_value(vehicle, "name", ""), row_value(vehicle, "plate_number", "")),
            )
            open_issue_count = sum(1 for row in status_rows if row_value(row, "status", "") not in issue_statuses)
        item["open_issue_count"] = open_issue_count
        if open_issue_count:
            item["alert"]["level"] = "red"
        item["oil_remaining"] = max(int_value(row_value(vehicle, "oil_change_interval", 5000), 5000) - int_value(row_value(item["alert"], "oil_km", 0)), 0)
        item["fixed_cost_total"] = sum(
            int_value(row_value(vehicle, key, 0))
            for key in ("inspection_cost", "tire_cost", "repair_cost", "parts_cost", "car_wash_cost", "insurance_fee", "parking_fee", "other_vehicle_cost")
        )
        vehicles.append(item)
    return vehicles


def depot_rate_for_delivery(rate_rows, depot_id: int, work_date: str):
    selected = None
    for rate in rate_rows:
        if int_value(row_value(rate, "depot_id")) != int_value(depot_id):
            continue
        start = row_value(rate, "effective_start", "") or ""
        end = row_value(rate, "effective_end", "") or ""
        if start and start > work_date:
            continue
        if end and end < work_date:
            continue
        if not selected or (row_value(rate, "effective_start", "") or "") >= (row_value(selected, "effective_start", "") or ""):
            selected = rate
    return selected


def delivery_revenue_from_depot_rate(delivery, rate):
    if not rate:
        return 0.0
    return (
        int_value(row_value(delivery, "completed", 0)) * money_value(row_value(rate, "base_unit", 0))
        + int_value(row_value(delivery, "transfer", 0)) * money_value(row_value(rate, "transfer_unit", 0))
        + int_value(row_value(delivery, "night", 0)) * money_value(row_value(rate, "night_unit", 0))
        + int_value(row_value(delivery, "pickup", 0)) * money_value(row_value(rate, "pickup_unit", 0))
        + int_value(row_value(delivery, "large", 0)) * money_value(row_value(rate, "large_unit", 0))
    )


def finance_delivery_rows(company_id: int, start_s: str, end_s: str, limit: int = 2000):
    return query_all(
        """SELECT d.*, u.name AS member_name,
                  COALESCE(dep.id, 0) AS depot_id,
                  COALESCE(dep.name, '未設定拠点') AS depot_name,
                  r.delivery_unit AS rate_delivery_unit,
                  r.transfer_unit AS rate_transfer_unit,
                  r.night_unit AS rate_night_unit,
                  r.pickup_unit AS rate_pickup_unit,
                  r.large_unit AS rate_large_unit,
                  r.vehicle_rental_type AS rate_vehicle_rental_type,
                  r.vehicle_daily_fee AS rate_vehicle_daily_fee,
                  r.vehicle_monthly_fee AS rate_vehicle_monthly_fee
           FROM deliveries d
           JOIN users u ON u.id=d.user_id AND u.company_id=d.company_id
           LEFT JOIN rates r ON r.user_id=d.user_id AND r.company_id=d.company_id
           LEFT JOIN shifts s ON s.user_id=d.user_id AND s.company_id=d.company_id AND s.shift_date=d.work_date
           LEFT JOIN districts dist ON dist.id=s.district_id AND dist.company_id=s.company_id
           LEFT JOIN depots dep ON dep.id=dist.depot_id AND dep.company_id=d.company_id
           WHERE d.company_id=? AND d.work_date BETWEEN ? AND ? AND COALESCE(d.is_deleted,0)=0
           ORDER BY d.work_date, depot_name, u.name
           LIMIT ?""",
        (company_id, start_s, end_s, limit),
    )


def finance_rate_rows(company_id: int, start_s: str, end_s: str):
    return query_all(
        """SELECT rs.*, dep.name AS depot_name
           FROM depot_rate_settings rs
           JOIN depots dep ON dep.id=rs.depot_id AND dep.company_id=rs.company_id
           WHERE rs.company_id=? AND rs.effective_start<=? AND (COALESCE(rs.effective_end,'')='' OR rs.effective_end>=?)
           ORDER BY rs.depot_id, rs.effective_start DESC, rs.id DESC
           LIMIT 500""",
        (company_id, end_s, start_s),
    )


def reward_rate_dict(row):
    return {
        "delivery_unit": money_value(row_value(row, "rate_delivery_unit", 0)),
        "transfer_unit": money_value(row_value(row, "rate_transfer_unit", 0)),
        "night_unit": money_value(row_value(row, "rate_night_unit", 0)),
        "pickup_unit": money_value(row_value(row, "rate_pickup_unit", 0)),
        "large_unit": money_value(row_value(row, "rate_large_unit", 0)),
        "vehicle_rental_type": row_value(row, "rate_vehicle_rental_type", "none") or "none",
        "vehicle_daily_fee": money_value(row_value(row, "rate_vehicle_daily_fee", 0)),
        "vehicle_monthly_fee": money_value(row_value(row, "rate_vehicle_monthly_fee", 0)),
    }


def finance_totals_for_period(company_id: int, start_s: str, end_s: str):
    deliveries = finance_delivery_rows(company_id, start_s, end_s)
    rate_rows = finance_rate_rows(company_id, start_s, end_s)
    revenue_total = 0.0
    driver_reward_total = 0.0
    delivery_count_total = 0
    monthly_vehicle_fees = {}
    depot_totals = {}
    for row in deliveries:
        depot_id = int_value(row_value(row, "depot_id", 0))
        depot_name = row_value(row, "depot_name", "") or "未設定拠点"
        rate = depot_rate_for_delivery(rate_rows, depot_id, row_value(row, "work_date", ""))
        revenue = delivery_revenue_from_depot_rate(row, rate)
        reward_rates = reward_rate_dict(row)
        gross_reward = calc_reward_gross(row, reward_rates)
        if reward_rates["vehicle_rental_type"] == "daily":
            gross_reward -= reward_rates["vehicle_daily_fee"]
        elif reward_rates["vehicle_rental_type"] == "monthly":
            monthly_vehicle_fees[int_value(row_value(row, "user_id"))] = reward_rates["vehicle_monthly_fee"]
        reward = gross_reward
        counts = {
            "completed": int_value(row_value(row, "completed", 0)),
            "transfer": int_value(row_value(row, "transfer", 0)),
            "night": int_value(row_value(row, "night", 0)),
            "pickup": int_value(row_value(row, "pickup", 0)),
            "large": int_value(row_value(row, "large", 0)),
        }
        delivery_count_total += counts["completed"]
        revenue_total += revenue
        driver_reward_total += reward
        depot = depot_totals.setdefault(
            depot_id,
            {"depot_id": depot_id, "depot_name": depot_name, "revenue": 0.0, "driver_reward": 0.0, "completed": 0, "transfer": 0, "night": 0, "pickup": 0, "large": 0},
        )
        depot["revenue"] += revenue
        depot["driver_reward"] += reward
        for key, value in counts.items():
            depot[key] += value
    driver_reward_total -= sum(monthly_vehicle_fees.values())
    for depot in depot_totals.values():
        depot["gross_profit_before_costs"] = depot["revenue"] - depot["driver_reward"]
    vehicle_expense = money_value(row_value(query_one("SELECT COALESCE(SUM(amount), 0) AS total FROM vehicle_expenses WHERE company_id=? AND expense_date BETWEEN ? AND ?", (company_id, start_s, end_s)), "total", 0))
    company_vehicle_expense = money_value(row_value(query_one("SELECT COALESCE(SUM(amount), 0) AS total FROM company_expenses WHERE company_id=? AND category=? AND expense_date BETWEEN ? AND ?", (company_id, "車両費", start_s, end_s)), "total", 0))
    other_expense = money_value(row_value(query_one("SELECT COALESCE(SUM(amount), 0) AS total FROM company_expenses WHERE company_id=? AND category<>? AND expense_date BETWEEN ? AND ?", (company_id, "車両費", start_s, end_s)), "total", 0))
    gross_profit = revenue_total - driver_reward_total - vehicle_expense - company_vehicle_expense - other_expense
    margin = (gross_profit / revenue_total * 100) if revenue_total else 0.0
    return {
        "revenue_total": revenue_total,
        "driver_reward_total": driver_reward_total,
        "delivery_count_total": delivery_count_total,
        "vehicle_expense_total": vehicle_expense + company_vehicle_expense,
        "other_expense_total": other_expense,
        "gross_profit": gross_profit,
        "gross_margin": margin,
        "depot_rows": sorted(depot_totals.values(), key=lambda item: item["depot_name"]),
    }


def build_finance_dashboard(company_id: int, target_month: Optional[str]):
    start, end = month_bounds(target_month)
    start_s, end_s = start.isoformat(), end.isoformat()
    totals = finance_totals_for_period(company_id, start_s, end_s)
    prev_month = (start - timedelta(days=1)).replace(day=1)
    prev_start, prev_end = month_bounds(prev_month.strftime("%Y-%m"))
    prev_totals = finance_totals_for_period(company_id, prev_start.isoformat(), prev_end.isoformat())
    vehicle_category_rows = query_all(
        """SELECT category, COALESCE(SUM(amount),0) AS total
           FROM vehicle_expenses
           WHERE company_id=? AND expense_date BETWEEN ? AND ?
           GROUP BY category ORDER BY total DESC LIMIT 20""",
        (company_id, start_s, end_s),
    )
    vehicle_rank_rows = query_all(
        """SELECT COALESCE(v.name, '未設定車両') AS vehicle_name, COALESCE(SUM(e.amount),0) AS total
           FROM vehicle_expenses e
           LEFT JOIN vehicles v ON v.id=e.vehicle_id AND v.company_id=e.company_id
           WHERE e.company_id=? AND e.expense_date BETWEEN ? AND ?
           GROUP BY e.vehicle_id, v.name ORDER BY total DESC LIMIT 20""",
        (company_id, start_s, end_s),
    )
    company_expense_rows = query_all(
        """SELECT e.*, COALESCE(d.name, '未設定拠点') AS depot_name
           FROM company_expenses e
           LEFT JOIN depots d ON d.id=e.depot_id AND d.company_id=e.company_id
           WHERE e.company_id=? AND e.expense_date BETWEEN ? AND ?
           ORDER BY e.expense_date DESC, e.id DESC LIMIT 50""",
        (company_id, start_s, end_s),
    )
    trend_rows = []
    trend_base = start
    for offset in range(5, -1, -1):
        month = trend_base.month - offset
        year = trend_base.year
        while month <= 0:
            month += 12
            year -= 1
        trend_start, trend_end = month_bounds(f"{year:04d}-{month:02d}")
        trend = finance_totals_for_period(company_id, trend_start.isoformat(), trend_end.isoformat())
        trend_rows.append({"month": trend_start.strftime("%Y-%m"), **trend})
    return {
        **totals,
        "start": start,
        "end": end,
        "target_month": start.strftime("%Y-%m"),
        "prev_revenue_total": prev_totals["revenue_total"],
        "prev_gross_profit": prev_totals["gross_profit"],
        "revenue_mom": totals["revenue_total"] - prev_totals["revenue_total"],
        "profit_mom": totals["gross_profit"] - prev_totals["gross_profit"],
        "vehicle_category_rows": vehicle_category_rows,
        "vehicle_rank_rows": vehicle_rank_rows,
        "company_expense_rows": company_expense_rows,
        "trend_rows": trend_rows,
    }


def vehicle_for_user(user, company_id: Optional[int] = None):
    company_id = company_id or company_id_for(user)
    vehicle_id = row_value(user, "last_vehicle_id")
    if vehicle_id:
        vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=?", (vehicle_id, company_id))
        if vehicle:
            return vehicle
    latest = query_one(
        """SELECT vehicle_id FROM work_logs
           WHERE user_id=? AND company_id=? AND vehicle_id IS NOT NULL AND COALESCE(is_deleted,0)=0
           ORDER BY work_date DESC, logged_at DESC LIMIT 1""",
        (user["id"], company_id),
    )
    if latest and latest["vehicle_id"]:
        vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=?", (latest["vehicle_id"], company_id))
        if vehicle:
            return vehicle
    vehicle_name = (row_value(user, "vehicle", "") or "").strip()
    if vehicle_name:
        vehicle = query_one("SELECT * FROM vehicles WHERE company_id=? AND name=?", (company_id, vehicle_name))
        if vehicle:
            return vehicle
    return None


def selected_vehicle_context(user):
    company_id = company_id_for(user)
    vehicles = vehicle_rows_for_company(company_id)
    selected = vehicle_for_user(user, company_id)
    if not selected and vehicles:
        selected = vehicles[0]
    return vehicles, selected


def latest_vehicle_odometer(company_id: int, vehicle_id: Optional[int]):
    if not vehicle_id:
        return 0
    vehicle = query_one("SELECT last_odometer FROM vehicles WHERE id=? AND company_id=?", (vehicle_id, company_id))
    if vehicle and vehicle["last_odometer"]:
        return vehicle["last_odometer"]
    row = query_one(
        """SELECT odometer FROM work_logs
           WHERE company_id=? AND vehicle_id=? AND COALESCE(odometer, 0)>0 AND COALESCE(is_deleted,0)=0
           ORDER BY work_date DESC, logged_at DESC LIMIT 1""",
        (company_id, vehicle_id),
    )
    return row["odometer"] if row else 0


def vehicle_label(vehicle):
    if not vehicle:
        return "車両未登録"
    plate = row_value(vehicle, "plate_number", "")
    return f"{vehicle['name']}（{plate}）" if plate else vehicle["name"]


def ensure_vehicle_for_user(company_id: int, user_id: int, vehicle_name: str):
    vehicle_name = (vehicle_name or "").strip()
    if not vehicle_name:
        return
    now = app_now().isoformat(timespec="seconds")
    with db() as conn:
        conn.execute(
            "INSERT OR IGNORE INTO vehicles(company_id, name, created_at, updated_at) VALUES (?, ?, ?, ?)",
            (company_id, vehicle_name, now, now),
        )
        vehicle = conn.execute("SELECT id FROM vehicles WHERE company_id=? AND name=?", (company_id, vehicle_name)).fetchone()
        if vehicle:
            conn.execute("UPDATE users SET last_vehicle_id=? WHERE id=? AND company_id=?", (vehicle["id"], user_id, company_id))
        conn.commit()


def update_vehicle_odometer(conn, company_id: int, vehicle_id: Optional[int], vehicle_name: str, odometer: int, used_at: str):
    if not vehicle_id and not vehicle_name:
        return
    if not vehicle_id and vehicle_name:
        conn.execute(
            "INSERT OR IGNORE INTO vehicles(company_id, name, created_at, updated_at) VALUES (?, ?, ?, ?)",
            (company_id, vehicle_name, used_at, used_at),
        )
        vehicle = conn.execute("SELECT id FROM vehicles WHERE company_id=? AND name=?", (company_id, vehicle_name)).fetchone()
        vehicle_id = vehicle["id"] if vehicle else None
    if vehicle_id:
        conn.execute(
            """UPDATE vehicles
               SET last_odometer=CASE WHEN COALESCE(last_odometer,0)<? THEN ? ELSE last_odometer END,
                   last_used_at=?, updated_at=?
               WHERE id=? AND company_id=?""",
            (max(odometer, 0), max(odometer, 0), used_at, used_at, vehicle_id, company_id),
        )


def odometer_log_values(conn, company_id: int, user_id: int, work_date: str, log_type: str, odometer: int):
    current = max(odometer, 0)
    if log_type == "start":
        return current, 0, 0
    start_log = conn.execute(
        """SELECT start_odometer, odometer FROM work_logs
           WHERE company_id=? AND user_id=? AND work_date=? AND log_type='start' AND COALESCE(is_deleted,0)=0
           ORDER BY logged_at DESC LIMIT 1""",
        (company_id, user_id, work_date),
    ).fetchone()
    start_odo = int_value(row_value(start_log, "start_odometer") or row_value(start_log, "odometer"))
    distance = max(current - start_odo, 0) if current and start_odo else 0
    return start_odo, current, distance


def open_work_session(company_id: int, user_id: int, work_date: str):
    return query_one(
        """SELECT * FROM work_log_sessions
           WHERE company_id=? AND user_id=? AND work_date=? AND status='working' AND COALESCE(is_deleted,0)=0
           ORDER BY session_no DESC, id DESC LIMIT 1""",
        (company_id, user_id, work_date),
    )


def next_work_session_no(conn, company_id: int, user_id: int, work_date: str):
    row = conn.execute(
        "SELECT COALESCE(MAX(session_no), 0) AS max_no FROM work_log_sessions WHERE company_id=? AND user_id=? AND work_date=?",
        (company_id, user_id, work_date),
    ).fetchone()
    return int_value(row["max_no"]) + 1


def session_delivery_totals(company_id: int, user_id: int, work_date: str):
    return query_one(
        """SELECT
             COALESCE(SUM(delivery_count), 0) AS completed,
             COALESCE(SUM(transfer_count), 0) AS transfer,
             COALESCE(SUM(night_count), 0) AS night,
             COALESCE(SUM(pickup_count), 0) AS pickup,
             COALESCE(SUM(large_count), 0) AS large,
             COUNT(*) AS session_count
           FROM work_log_sessions
           WHERE company_id=? AND user_id=? AND work_date=? AND COALESCE(is_deleted,0)=0""",
        (company_id, user_id, work_date),
    )


def sync_delivery_from_sessions(company_id: int, user_id: int, work_date: str, actor_id: int, source: str = "勤務セッション集計", vehicle_id: Optional[int] = None, vehicle_name: str = "", image_path: str = ""):
    totals = session_delivery_totals(company_id, user_id, work_date)
    if not totals or int_value(row_value(totals, "session_count")) == 0:
        return None
    return upsert_delivery_counts(
        company_id,
        user_id,
        work_date,
        int_value(totals["completed"]),
        int_value(totals["transfer"]),
        int_value(totals["night"]),
        int_value(totals["pickup"]),
        int_value(totals["large"]),
        "none",
        source,
        image_path,
        actor_id,
        source,
        vehicle_id,
        vehicle_name,
    )


def insert_work_log_correction(conn, company_id: int, user_id: int, work_date: str, work_log_id, session_id, target_table: str, before, after, actor_id: int, note: str = ""):
    before_dict = row_to_dict(before)
    after_dict = row_to_dict(after)
    if before_dict == after_dict:
        return
    changed_at = app_now().isoformat(timespec="seconds")
    cur = conn.execute(
        """INSERT INTO work_log_corrections(company_id, user_id, work_date, work_log_id, session_id, target_table,
           before_data, after_data, actor_id, changed_at, note)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            company_id,
            user_id,
            work_date,
            work_log_id,
            session_id,
            target_table,
            json.dumps(before_dict, ensure_ascii=False, default=str),
            json.dumps(after_dict, ensure_ascii=False, default=str),
            actor_id,
            changed_at,
            note,
        ),
    )
    audit_log(
        company_id,
        {"id": actor_id, "role": ""},
        f"{target_table}.correction",
        target_table,
        session_id or work_log_id or cur.lastrowid,
        note or "Work record corrected",
        before=before_dict,
        after=after_dict,
        conn=conn,
    )


def update_work_session_from_form(
    session_id: int,
    company_id: int,
    actor,
    user_id_guard: Optional[int] = None,
    start_time: str = "",
    end_time: str = "",
    vehicle_id: int = 0,
    start_odometer: int = 0,
    end_odometer: int = 0,
    alcohol_check_start: str = "",
    alcohol_check_end: str = "",
    call_check_start: str = "",
    call_check_end: str = "",
    delivery_count: int = 0,
    transfer_count: int = 0,
    night_count: int = 0,
    pickup_count: int = 0,
    large_count: int = 0,
    inspection_image_path: str = "",
    notes: str = "",
):
    guard_sql = " AND user_id=?" if user_id_guard else ""
    guard_params = [user_id_guard] if user_id_guard else []
    before = query_one(
        f"SELECT * FROM work_log_sessions WHERE id=? AND company_id=? AND COALESCE(is_deleted,0)=0{guard_sql}",
        [session_id, company_id] + guard_params,
    )
    if not before:
        raise HTTPException(status_code=404, detail="勤務セッションが見つかりません")
    vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=?", (vehicle_id, company_id)) if vehicle_id else None
    vehicle_name = vehicle_label(vehicle) if vehicle else row_value(before, "vehicle_name", "")
    start_time = start_time or row_value(before, "start_time", "")
    end_time = end_time or row_value(before, "end_time", "")
    start_odo = max(int_value(start_odometer), 0)
    end_odo = max(int_value(end_odometer), 0)
    distance_km = max(end_odo - start_odo, 0) if start_odo and end_odo else int_value(row_value(before, "distance_km"))
    status = "finished" if end_time else "working"
    inspection_image_path = inspection_image_path.strip() or row_value(before, "inspection_image_path", "")
    now = app_now().isoformat(timespec="seconds")
    with db() as conn:
        conn.execute(
            """UPDATE work_log_sessions
               SET start_time=?, end_time=?, vehicle_id=?, vehicle_name=?, start_odometer=?, end_odometer=?,
                   distance_km=?, status=?, alcohol_check_start=?, alcohol_check_end=?, call_check_start=?,
                   call_check_end=?, delivery_count=?, transfer_count=?, night_count=?, pickup_count=?,
                   large_count=?, inspection_image_path=?, notes=?, updated_at=?
               WHERE id=? AND company_id=? AND COALESCE(is_deleted,0)=0""",
            (
                start_time,
                end_time,
                vehicle["id"] if vehicle else row_value(before, "vehicle_id"),
                vehicle_name,
                start_odo,
                end_odo,
                distance_km,
                status,
                alcohol_check_start,
                alcohol_check_end,
                call_check_start,
                call_check_end,
                max(delivery_count, 0),
                max(transfer_count, 0),
                max(night_count, 0),
                max(pickup_count, 0),
                max(large_count, 0),
                inspection_image_path,
                notes,
                now,
                session_id,
                company_id,
            ),
        )
        if row_value(before, "start_log_id"):
            conn.execute(
                """UPDATE work_logs
                   SET logged_at=?, vehicle_id=?, vehicle_name=?, odometer=?, start_odometer=?,
                       alcohol_result=?, health_status=?, notes=?
                   WHERE id=? AND company_id=? AND COALESCE(is_deleted,0)=0""",
                (start_time, vehicle["id"] if vehicle else row_value(before, "vehicle_id"), vehicle_name, start_odo, start_odo, alcohol_check_start, call_check_start, notes, before["start_log_id"], company_id),
            )
        if row_value(before, "end_log_id"):
            conn.execute(
                """UPDATE work_logs
                   SET logged_at=?, vehicle_id=?, vehicle_name=?, odometer=?, start_odometer=?, end_odometer=?,
                       distance_km=?, alcohol_result=?, detector_used=?, notes=?
                   WHERE id=? AND company_id=? AND COALESCE(is_deleted,0)=0""",
                (end_time, vehicle["id"] if vehicle else row_value(before, "vehicle_id"), vehicle_name, end_odo, start_odo, end_odo, distance_km, alcohol_check_end, call_check_end, notes, before["end_log_id"], company_id),
            )
        if vehicle:
            conn.execute("UPDATE users SET last_vehicle_id=?, vehicle=? WHERE id=? AND company_id=?", (vehicle["id"], vehicle["name"], before["user_id"], company_id))
        update_vehicle_odometer(conn, company_id, vehicle["id"] if vehicle else row_value(before, "vehicle_id"), vehicle_name, end_odo or start_odo, now)
        after = conn.execute("SELECT * FROM work_log_sessions WHERE id=? AND company_id=? AND COALESCE(is_deleted,0)=0", (session_id, company_id)).fetchone()
        insert_work_log_correction(
            conn,
            company_id,
            before["user_id"],
            before["work_date"],
            row_value(before, "end_log_id") or row_value(before, "start_log_id"),
            session_id,
            "work_log_sessions",
            before,
            after,
            row_value(actor, "id"),
            "勤務セッション修正",
        )
        conn.commit()
    sync_delivery_from_sessions(company_id, before["user_id"], before["work_date"], row_value(actor, "id"), "勤務セッション修正", vehicle["id"] if vehicle else row_value(before, "vehicle_id"), vehicle_name, inspection_image_path)
    return after


def parse_iso_date(value: str):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def vehicle_alert(vehicle, today: Optional[date] = None):
    today = today or app_today()
    if not vehicle:
        return {"level": "", "oil_message": "車両未登録", "inspection_message": "車検未設定", "oil_km": 0, "inspection_days": None}
    last_odometer = int_value(row_value(vehicle, "last_odometer", 0))
    oil_base = int_value(row_value(vehicle, "oil_change_odometer", 0))
    oil_interval = int_value(row_value(vehicle, "oil_change_interval", 5000), 5000) or 5000
    oil_km = max(last_odometer - oil_base, 0) if oil_base else 0
    oil_level = ""
    if oil_base:
        remaining_for_alert = oil_interval - oil_km
        if remaining_for_alert <= 0:
            oil_level = "red"
        elif remaining_for_alert <= 250:
            oil_level = "orange"
        elif remaining_for_alert <= 500:
            oil_level = "yellow"
    if oil_base:
        remaining = max(oil_interval - oil_km, 0)
        oil_message = "オイル交換時期です" if oil_km >= oil_interval else f"オイル交換まであと{remaining}kmです"
    else:
        oil_message = "オイル交換距離未設定"

    due_date = parse_iso_date(row_value(vehicle, "inspection_due", ""))
    inspection_days = None
    inspection_level = ""
    inspection_message = "車検期限未設定"
    if due_date:
        inspection_days = (due_date - today).days
        if inspection_days < 0:
            inspection_level = "expired"
            inspection_message = "車検期限切れ"
        elif inspection_days <= 7:
            inspection_level = "red"
            inspection_message = f"車検期限まであと{inspection_days}日です"
        elif inspection_days <= 14:
            inspection_level = "orange"
            inspection_message = f"車検期限まであと{inspection_days}日です"
        elif inspection_days <= 30:
            inspection_level = "yellow"
            inspection_message = f"車検期限まであと{inspection_days}日です"
        else:
            inspection_message = f"車検期限: {due_date.isoformat()}"

    level_order = {"": 0, "yellow": 1, "orange": 2, "red": 3, "expired": 4}
    level = inspection_level if level_order.get(inspection_level, 0) >= level_order.get(oil_level, 0) else oil_level
    return {
        "level": level,
        "oil_level": oil_level,
        "inspection_level": inspection_level,
        "oil_message": oil_message,
        "inspection_message": inspection_message,
        "oil_km": oil_km,
        "inspection_days": inspection_days,
        "tire_note": row_value(vehicle, "tire_note", ""),
    }


def holiday_deadline_alert(today: Optional[date] = None):
    today = today or app_today()
    if today.day > 25:
        return {"level": "ended", "message": "今月分の休み希望締切は終了しました"}
    if today.day >= 20:
        return {"level": "near", "message": "休み希望の締切が近いです"}
    return {"level": "", "message": ""}


def holiday_submission_status(company_id: int, target_month: Optional[date] = None):
    base = target_month or (app_today().replace(day=1) + timedelta(days=32)).replace(day=1)
    start = base.replace(day=1)
    end = (start.replace(year=start.year + 1, month=1) if start.month == 12 else start.replace(month=start.month + 1)) - timedelta(days=1)
    member_count = query_one(
        "SELECT COUNT(*) AS count FROM users WHERE role='member' AND active=1 AND COALESCE(purged,0)=0 AND company_id=?",
        (company_id,),
    )["count"]
    submitted = query_one(
        """SELECT COUNT(DISTINCT user_id) AS count FROM holiday_requests
           WHERE company_id=? AND request_date BETWEEN ? AND ?""",
        (company_id, start.isoformat(), end.isoformat()),
    )["count"]
    return {
        "ym": start.strftime("%Y-%m"),
        "member_count": member_count,
        "submitted_count": submitted,
        "missing_count": max(member_count - submitted, 0),
        "deadline_alert": holiday_deadline_alert(),
    }


def home_flow_message(vehicle_alert_info, holiday_alert_info):
    if vehicle_alert_info.get("inspection_level") == "expired":
        return "車検期限切れです。管理者に確認してください"
    if vehicle_alert_info.get("inspection_level") == "red":
        return "車検期限が近いです"
    if vehicle_alert_info.get("oil_km", 0) >= 5000:
        return "オイル交換時期です"
    if holiday_alert_info.get("level") == "near":
        return holiday_alert_info["message"]
    if holiday_alert_info.get("level") == "ended":
        return holiday_alert_info["message"]
    if vehicle_alert_info.get("tire_note"):
        return "タイヤの減りを確認してください"
    now = app_now()
    if now.hour >= 17:
        return "夜間配達はライト確認を忘れずに"
    return "今日も無事故で頑張ってください！"


def delivery_values(completed: int, transfer: int, night: int, pickup: int, large: int):
    return {
        "completed": max(completed, 0),
        "transfer": max(transfer, 0),
        "night": max(night, 0),
        "pickup": max(pickup, 0),
        "large": max(large, 0),
    }


def delivery_changed(existing, values):
    if not existing:
        return False
    return any(int_value(existing[key]) != int_value(values[key]) for key in ("completed", "transfer", "night", "pickup", "large"))


def insert_delivery_correction(conn, company_id: int, delivery_id: int, user_id: int, actor_id: int, work_date: str, before, after, source: str):
    if not before or not delivery_changed(before, after):
        return
    conn.execute(
        """INSERT INTO delivery_corrections(company_id, delivery_id, user_id, actor_id, work_date,
           before_completed, before_transfer, before_night, before_pickup, before_large,
           after_completed, after_transfer, after_night, after_pickup, after_large, source, changed_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            company_id,
            delivery_id,
            user_id,
            actor_id,
            work_date,
            int_value(before["completed"]),
            int_value(before["transfer"]),
            int_value(before["night"]),
            int_value(before["pickup"]),
            int_value(before["large"]),
            int_value(after["completed"]),
            int_value(after["transfer"]),
            int_value(after["night"]),
            int_value(after["pickup"]),
            int_value(after["large"]),
            source,
            app_now().isoformat(timespec="seconds"),
        ),
    )


def upsert_delivery_counts(company_id: int, user_id: int, work_date: str, completed: int, transfer: int, night: int, pickup: int, large: int, vehicle_rental: str = "none", memo: str = "", inspection_sheet_path: str = "", actor_id: Optional[int] = None, source: str = "", vehicle_id: Optional[int] = None, vehicle_name: str = ""):
    now = app_now().isoformat(timespec="seconds")
    values = delivery_values(completed, transfer, night, pickup, large)
    with db() as conn:
        existing = conn.execute("SELECT * FROM deliveries WHERE user_id=? AND company_id=? AND work_date=?", (user_id, company_id, work_date)).fetchone()
        conn.execute(
            """INSERT INTO deliveries(company_id, user_id, work_date, completed, transfer, night, pickup, large, vehicle_rental, memo, inspection_sheet_path, created_at, updated_at, vehicle_id, vehicle_name, is_deleted, deleted_at, deleted_by, delete_reason)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
               ON CONFLICT(user_id, work_date) DO UPDATE SET completed=excluded.completed, transfer=excluded.transfer,
               night=excluded.night, pickup=excluded.pickup, large=excluded.large, vehicle_rental=excluded.vehicle_rental,
               memo=excluded.memo, inspection_sheet_path=COALESCE(NULLIF(excluded.inspection_sheet_path, ''), deliveries.inspection_sheet_path),
               vehicle_id=COALESCE(excluded.vehicle_id, deliveries.vehicle_id),
               vehicle_name=COALESCE(NULLIF(excluded.vehicle_name, ''), deliveries.vehicle_name),
               is_deleted=0, deleted_at='', deleted_by=NULL, delete_reason='',
               updated_at=excluded.updated_at, company_id=excluded.company_id""",
            (
                company_id,
                user_id,
                work_date,
                values["completed"],
                values["transfer"],
                values["night"],
                values["pickup"],
                values["large"],
                vehicle_rental,
                memo,
                inspection_sheet_path or "",
                now,
                now,
                vehicle_id,
                vehicle_name or "",
                0,
                "",
                None,
                "",
            ),
        )
        delivery = conn.execute("SELECT * FROM deliveries WHERE user_id=? AND company_id=? AND work_date=?", (user_id, company_id, work_date)).fetchone()
        if existing and delivery:
            insert_delivery_correction(conn, company_id, delivery["id"], user_id, actor_id or user_id, work_date, existing, values, source)
        if delivery:
            audit_log(
                company_id,
                {"id": actor_id or user_id, "role": ""},
                "delivery.update" if existing else "delivery.create",
                "deliveries",
                delivery["id"],
                source or "Delivery counts saved",
                before=row_to_dict(existing),
                after=row_to_dict(delivery),
                conn=conn,
            )
        conn.commit()
        return delivery


def active_delivery_for_day(company_id: int, user_id: int, work_date: str):
    return query_one(
        "SELECT * FROM deliveries WHERE company_id=? AND user_id=? AND work_date=? AND COALESCE(is_deleted,0)=0",
        (company_id, user_id, work_date),
    )


def active_sessions_for_day(company_id: int, user_id: int, work_date: str):
    return query_all(
        """SELECT s.*, v.plate_number
           FROM work_log_sessions s
           LEFT JOIN vehicles v ON v.id=s.vehicle_id AND v.company_id=s.company_id
           WHERE s.company_id=? AND s.user_id=? AND s.work_date=? AND COALESCE(s.is_deleted,0)=0
           ORDER BY s.session_no, s.id""",
        (company_id, user_id, work_date),
    )


def active_work_logs_for_day(company_id: int, user_id: int, work_date: str):
    return query_all(
        """SELECT * FROM work_logs
           WHERE company_id=? AND user_id=? AND work_date=? AND COALESCE(is_deleted,0)=0
           ORDER BY logged_at, id""",
        (company_id, user_id, work_date),
    )


def member_result_day_context(company_id: int, user_id: int, work_date: str):
    delivery = active_delivery_for_day(company_id, user_id, work_date)
    sessions = active_sessions_for_day(company_id, user_id, work_date)
    logs = active_work_logs_for_day(company_id, user_id, work_date)
    rates = query_one("SELECT * FROM rates WHERE user_id=? AND company_id=?", (user_id, company_id))
    vehicle_rates = vehicle_rates_for_company(company_id)
    reward = calc_reward(delivery, rates, vehicle_rates) if delivery else 0
    slip = query_one(
        "SELECT * FROM inspection_slips WHERE user_id=? AND company_id=? AND slip_date=?",
        (user_id, company_id, work_date),
    )
    sheet = query_one(
        "SELECT * FROM inspection_sheets WHERE user_id=? AND company_id=? AND (sheet_date=? OR COALESCE(delivery_date,'')=?)",
        (user_id, company_id, work_date, work_date),
    )
    image_path = row_value(slip, "file_path", "") or row_value(delivery, "inspection_sheet_path", "") or row_value(sheet, "file_path", "")
    corrections = query_all(
        """SELECT c.*, a.name AS actor_name FROM delivery_corrections c
           LEFT JOIN users a ON a.id=c.actor_id
           WHERE c.user_id=? AND c.company_id=? AND c.work_date=?
           ORDER BY c.changed_at DESC LIMIT 30""",
        (user_id, company_id, work_date),
    )
    deletions = query_all(
        """SELECT d.*, a.name AS actor_name FROM work_day_deletions d
           LEFT JOIN users a ON a.id=d.actor_id
           WHERE d.user_id=? AND d.company_id=? AND d.work_date=?
           ORDER BY d.deleted_at DESC LIMIT 10""",
        (user_id, company_id, work_date),
    )
    return {
        "delivery": delivery,
        "sessions": sessions,
        "logs": logs,
        "rates": rates,
        "reward": reward,
        "slip": slip,
        "sheet": sheet,
        "inspection_image_path": image_path,
        "inspection_image": inspection_image_info(image_path),
        "corrections": corrections,
        "deletions": deletions,
    }


def member_month_calendar(company_id: int, user_id: int, month_text: Optional[str]):
    start, end = month_bounds(month_text)
    start_s, end_s = start.isoformat(), end.isoformat()
    delivery_rows = query_all(
        """SELECT * FROM deliveries
           WHERE company_id=? AND user_id=? AND work_date BETWEEN ? AND ? AND COALESCE(is_deleted,0)=0
           ORDER BY work_date""",
        (company_id, user_id, start_s, end_s),
    )
    session_rows = query_all(
        """SELECT work_date,
                  COUNT(*) AS session_count,
                  SUM(CASE WHEN status='working' THEN 1 ELSE 0 END) AS working_count,
                  SUM(CASE WHEN COALESCE(end_time,'')<>'' OR status='finished' THEN 1 ELSE 0 END) AS finished_count
           FROM work_log_sessions
           WHERE company_id=? AND user_id=? AND work_date BETWEEN ? AND ? AND COALESCE(is_deleted,0)=0
           GROUP BY work_date""",
        (company_id, user_id, start_s, end_s),
    )
    log_rows = query_all(
        """SELECT work_date,
                  SUM(CASE WHEN log_type='start' THEN 1 ELSE 0 END) AS start_count,
                  SUM(CASE WHEN log_type='end' THEN 1 ELSE 0 END) AS end_count
           FROM work_logs
           WHERE company_id=? AND user_id=? AND work_date BETWEEN ? AND ? AND COALESCE(is_deleted,0)=0
           GROUP BY work_date""",
        (company_id, user_id, start_s, end_s),
    )
    rates = query_one("SELECT * FROM rates WHERE user_id=? AND company_id=?", (user_id, company_id))
    vehicle_rates = vehicle_rates_for_company(company_id)
    by_day = {}
    for delivery in delivery_rows:
        item = by_day.setdefault(delivery["work_date"], {})
        item["delivery"] = delivery
        item["reward"] = calc_reward(delivery, rates, vehicle_rates) if rates else 0
    for session in session_rows:
        item = by_day.setdefault(session["work_date"], {})
        item["session_count"] = int_value(session["session_count"])
        item["working_count"] = int_value(session["working_count"])
        item["finished_count"] = int_value(session["finished_count"])
    for log in log_rows:
        item = by_day.setdefault(log["work_date"], {})
        item["start_count"] = int_value(log["start_count"])
        item["end_count"] = int_value(log["end_count"])
    weeks = []
    cal = calendar.Calendar(firstweekday=6)
    for week in cal.monthdatescalendar(start.year, start.month):
        cells = []
        for day_obj in week:
            day_s = day_obj.isoformat()
            data = by_day.get(day_s, {})
            delivery = data.get("delivery")
            start_count = int_value(data.get("start_count")) or int_value(data.get("session_count"))
            end_count = int_value(data.get("end_count")) or int_value(data.get("finished_count"))
            working_count = int_value(data.get("working_count"))
            cells.append(
                {
                    "date": day_s,
                    "day": day_obj.day,
                    "current_month": day_obj.month == start.month,
                    "is_today": day_s == app_today().isoformat(),
                    "delivery": delivery,
                    "reward": int_value(data.get("reward")),
                    "has_work": bool(start_count),
                    "finished": bool(end_count) and not working_count,
                    "working": bool(working_count),
                }
            )
        weeks.append(cells)
    return {"start": start, "end": end, "target_month": start.strftime("%Y-%m"), "weeks": weeks}


def logically_delete_work_day(company_id: int, user_id: int, work_date: str, actor, reason: str):
    reason = (reason or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="削除理由を入力してください")
    deleted_at = app_now().isoformat(timespec="seconds")
    with db() as conn:
        sessions = conn.execute("SELECT * FROM work_log_sessions WHERE company_id=? AND user_id=? AND work_date=? AND COALESCE(is_deleted,0)=0", (company_id, user_id, work_date)).fetchall()
        logs = conn.execute("SELECT * FROM work_logs WHERE company_id=? AND user_id=? AND work_date=? AND COALESCE(is_deleted,0)=0", (company_id, user_id, work_date)).fetchall()
        delivery = conn.execute("SELECT * FROM deliveries WHERE company_id=? AND user_id=? AND work_date=? AND COALESCE(is_deleted,0)=0", (company_id, user_id, work_date)).fetchone()
        before = {
            "sessions": [row_to_dict(row) for row in sessions],
            "work_logs": [row_to_dict(row) for row in logs],
            "delivery": row_to_dict(delivery),
        }
        if not sessions and not logs and not delivery:
            raise HTTPException(status_code=404, detail="削除対象の出勤データが見つかりません")
        conn.execute(
            "UPDATE work_log_sessions SET is_deleted=1, deleted_at=?, deleted_by=?, delete_reason=?, updated_at=? WHERE company_id=? AND user_id=? AND work_date=? AND COALESCE(is_deleted,0)=0",
            (deleted_at, row_value(actor, "id"), reason, deleted_at, company_id, user_id, work_date),
        )
        conn.execute(
            "UPDATE work_logs SET is_deleted=1, deleted_at=?, deleted_by=?, delete_reason=? WHERE company_id=? AND user_id=? AND work_date=? AND COALESCE(is_deleted,0)=0",
            (deleted_at, row_value(actor, "id"), reason, company_id, user_id, work_date),
        )
        conn.execute(
            "UPDATE deliveries SET is_deleted=1, deleted_at=?, deleted_by=?, delete_reason=?, updated_at=? WHERE company_id=? AND user_id=? AND work_date=? AND COALESCE(is_deleted,0)=0",
            (deleted_at, row_value(actor, "id"), reason, deleted_at, company_id, user_id, work_date),
        )
        cur = conn.execute(
            """INSERT INTO work_day_deletions(company_id, user_id, work_date, actor_id, reason, before_data, deleted_at)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (company_id, user_id, work_date, row_value(actor, "id"), reason, json.dumps(before, ensure_ascii=False, default=str), deleted_at),
        )
        audit_log(
            company_id,
            actor,
            "work_day.delete",
            "work_day_deletions",
            cur.lastrowid,
            "Work day logically deleted",
            before=before,
            after={"work_date": work_date, "reason": reason, "deleted_at": deleted_at},
            conn=conn,
        )
        conn.commit()


def save_inspection_slip_record(company_id: int, user_id: int, delivery_id: int, work_date: str, image_path: str, original_filename: str = "", content_type: str = ""):
    if not image_path:
        return None
    now_dt = app_now()
    now = now_dt.isoformat(timespec="seconds")
    remote_filename = unquote(image_path.split("?", 1)[0].rstrip("/").split("/")[-1]) if is_remote_image_path(image_path) else ""
    filename = remote_filename or Path(original_filename or image_path).name
    if is_remote_image_path(image_path):
        storage_path = storage_path_from_url(image_path) or _storage_path("inspection_slips", filename)
    elif allowed_local_image_path(image_path):
        storage_path = ""
    else:
        storage_path = ""
    public_url = image_path if is_remote_image_path(image_path) else ""
    retention = retention_until(now_dt)
    with db() as conn:
        conn.execute(
            """INSERT INTO inspection_slips(company_id, user_id, delivery_id, slip_date, file_path, storage_path, public_url, signed_url,
               original_filename, content_type, uploaded_at, retention_until)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(user_id, slip_date) DO UPDATE SET delivery_id=excluded.delivery_id,
               company_id=excluded.company_id, file_path=excluded.file_path, storage_path=excluded.storage_path,
               public_url=excluded.public_url, signed_url=excluded.signed_url, original_filename=excluded.original_filename,
               content_type=excluded.content_type, uploaded_at=excluded.uploaded_at, retention_until=excluded.retention_until""",
            (company_id, user_id, delivery_id, work_date, image_path, storage_path, public_url, "", original_filename, content_type, now, retention),
        )
        conn.execute("UPDATE deliveries SET inspection_sheet_path=?, updated_at=? WHERE id=? AND company_id=?", (image_path, now, delivery_id, company_id))
        slip = conn.execute("SELECT id FROM inspection_slips WHERE user_id=? AND company_id=? AND slip_date=?", (user_id, company_id, work_date)).fetchone()
        audit_log(
            company_id,
            {"id": user_id, "role": ""},
            "inspection_slip.upsert",
            "inspection_slips",
            row_value(slip, "id", delivery_id),
            "Inspection slip image saved",
            after={"delivery_id": delivery_id, "work_date": work_date, "image_path": image_path},
            conn=conn,
        )
        conn.commit()
        return row_value(slip, "id")


def extract_delivery_counts(ocr_text: str = ""):
    fields = extract_inspection_fields(ocr_text)
    counts = {
        "completed": fields["completed"],
        "transfer": _count_after_labels(ocr_text, ["転送"]),
        "night": _count_after_labels(ocr_text, ["夜間"]),
        "pickup": fields["pickup"],
        "large": _count_after_labels(ocr_text, ["大型", "大口"]),
    }
    return counts


def normalize_ocr_text(ocr_text: str = ""):
    text = unicodedata.normalize("NFKC", ocr_text or "")
    text = text.translate(FULLWIDTH_DIGITS)
    replacements = {
        "配達完了": "配達完了",
        "配達完ア": "配達完了",
        "配達完丁": "配達完了",
        "引受け": "引受",
        "引き受け": "引受",
        "引 受": "引受",
        "受 入": "受入",
    }
    for before, after in replacements.items():
        text = text.replace(before, after)
    return text.replace(",", "")


def parse_inspection_date(ocr_text: str = "", fallback_date: str = ""):
    text = normalize_ocr_text(ocr_text)
    patterns = [
        r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日",
        r"(\d{4})[-/\.](\d{1,2})[-/\.](\d{1,2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        year, month, day = [int(value) for value in match.groups()]
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            pass
    return normalize_import_date(fallback_date) if fallback_date else ""


def _numbers_after_label(line: str, label: str):
    match = re.search(rf"{re.escape(label)}[^\d]{{0,18}}(\d+)", line)
    if match:
        return int(match.group(1))
    return None


def _count_after_labels(ocr_text: str = "", labels=None):
    labels = labels or []
    text = normalize_ocr_text(ocr_text)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for label in labels:
        for line in lines:
            value = _numbers_after_label(line, label)
            if value is not None:
                return value
    compact = re.sub(r"\s+", "", text)
    for label in labels:
        value = _numbers_after_label(compact, label)
        if value is not None:
            return value
    return 0


def extract_inspection_fields(ocr_text: str = "", fallback_date: str = ""):
    completed = _count_after_labels(ocr_text, ["配達完了", "配完", "完了"])
    acceptance = _count_after_labels(ocr_text, ["引受", "集荷"])
    received = _count_after_labels(ocr_text, ["受入"])
    return {
        "work_date": parse_inspection_date(ocr_text, fallback_date),
        "completed": completed,
        "pickup": acceptance,
        "acceptance": acceptance,
        "received": received,
    }


def extract_ocr_text(image_path: Path):
    try:
        from PIL import Image
        from PIL import ImageEnhance, ImageFilter, ImageOps
        import pytesseract

        lang = os.getenv("OCR_LANG", "jpn+eng")
        config = os.getenv("OCR_CONFIG", "--psm 6")
        with Image.open(image_path) as image:
            image = ImageOps.exif_transpose(image)
            image.thumbnail((1800, 2600))
            variants = [image]
            gray = ImageOps.grayscale(image)
            gray = ImageOps.autocontrast(gray)
            gray = ImageEnhance.Contrast(gray).enhance(1.8)
            if max(gray.size) < 2200:
                gray = gray.resize((gray.width * 2, gray.height * 2))
            variants.append(gray.filter(ImageFilter.SHARPEN))
            texts = [pytesseract.image_to_string(item, lang=lang, config=config).strip() for item in variants]
            return max(texts, key=len).strip()
    except Exception:
        return ""


def auto_comment(delivery, reward, has_shift=True):
    if not delivery:
        return "今日の入力をすると、ここに応援コメントが表示されます。"
    if delivery["night"] > 0:
        return "夜間までありがとうございました！帰り道も安全第一でお願いします。"
    if delivery["completed"] >= 120:
        return "今日は配達個数が多かったですね。しっかり休んでください！"
    if reward >= 20000:
        return "今月も順調です。この調子でいきましょう！"
    if not has_shift:
        return "休み希望やシフトも忘れず確認しましょう。"
    return "今日もお疲れ様でした！明日も安全運転でお願いします！"


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    user = current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    return RedirectResponse("/admin" if user["role"] == "admin" else "/member", status_code=303)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return render(request, "login.html", {"error": ""})


def smtp_configured():
    return bool(SMTP_HOST and SMTP_FROM)


def external_url(request: Request, path: str):
    if APP_BASE_URL:
        return f"{APP_BASE_URL}{path}"
    return f"{str(request.base_url).rstrip('/')}{path}"


def send_password_reset_email(to_email: str, reset_url: str):
    if not smtp_configured():
        return False
    message = EmailMessage()
    message["Subject"] = "SPARKLE DRIVE パスワード再設定"
    message["From"] = SMTP_FROM
    message["To"] = to_email
    message.set_content(
        "SPARKLE DRIVE のパスワード再設定リンクです。\n"
        f"{PASSWORD_RESET_EXPIRY_MINUTES}分以内に以下のURLから新しいパスワードを登録してください。\n\n"
        f"{reset_url}\n\n"
        "このメールに覚えがない場合は破棄してください。"
    )
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as server:
            if SMTP_USE_TLS:
                server.starttls()
            if SMTP_USERNAME:
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.send_message(message)
        return True
    except Exception:
        logger.exception("Password reset email failed")
        return False


def create_password_reset(conn, request: Request, user_row):
    token = secrets.token_urlsafe(32)
    token_hash = sha256(token.encode("utf-8")).hexdigest()
    now_dt = app_now()
    expires_at = (now_dt + timedelta(minutes=PASSWORD_RESET_EXPIRY_MINUTES)).isoformat(timespec="seconds")
    reset_url = external_url(request, f"/password-reset/{token}")
    dev_reset_url = "" if smtp_configured() else reset_url
    conn.execute(
        """INSERT INTO password_reset_tokens(company_id, user_id, email, token_hash, expires_at, used_at, dev_reset_url, created_at)
           VALUES (?, ?, ?, ?, ?, '', ?, ?)""",
        (
            row_value(user_row, "company_id", 1),
            user_row["id"],
            row_value(user_row, "email", ""),
            token_hash,
            expires_at,
            dev_reset_url,
            now_dt.isoformat(timespec="seconds"),
        ),
    )
    return reset_url


def password_reset_user_for_token(token: str):
    token_hash = sha256((token or "").encode("utf-8")).hexdigest()
    now_s = app_now().isoformat(timespec="seconds")
    return query_one(
        """SELECT t.*, u.username, u.name, u.password_hash, u.active, u.purged
           FROM password_reset_tokens t
           JOIN users u ON u.id=t.user_id AND u.company_id=t.company_id
           WHERE t.token_hash=? AND COALESCE(t.used_at, '')='' AND t.expires_at>=?
             AND u.active=1 AND COALESCE(u.purged,0)=0
           ORDER BY t.id DESC LIMIT 1""",
        (token_hash, now_s),
    )


@app.get("/password-reset/request", response_class=HTMLResponse)
def password_reset_request_page(request: Request, message: str = "", error: str = ""):
    return render(request, "password_reset_request.html", {"message": message, "error": error})


@app.post("/password-reset/request", response_class=HTMLResponse)
def password_reset_request(request: Request, email: str = Form(...)):
    email = email.strip().lower()
    message = "登録メールアドレスが見つかった場合、再設定リンクを発行しました。"
    if not email:
        return render(request, "password_reset_request.html", {"message": "", "error": "メールアドレスを入力してください"})
    user_rows = query_all(
        """SELECT * FROM users
           WHERE LOWER(COALESCE(email, ''))=? AND active=1 AND COALESCE(purged,0)=0
           ORDER BY id LIMIT 10""",
        (email,),
    )
    for user_row in user_rows:
        with db() as conn:
            reset_url = create_password_reset(conn, request, user_row)
            audit_log(
                company_id_for(user_row),
                {"id": user_row["id"], "role": row_value(user_row, "role", "")},
                "password_reset.request",
                "users",
                user_row["id"],
                "Password reset requested",
                conn=conn,
            )
            conn.commit()
        if not send_password_reset_email(email, reset_url):
            logger.warning("Password reset URL for %s: %s", email, reset_url)
    return render(request, "password_reset_request.html", {"message": message, "error": ""})


@app.get("/password-reset/{token}", response_class=HTMLResponse)
def password_reset_form(request: Request, token: str):
    token_row = password_reset_user_for_token(token)
    if not token_row:
        return render(request, "password_reset_form.html", {"token": "", "error": "再設定リンクが無効、または期限切れです。", "message": ""})
    return render(request, "password_reset_form.html", {"token": token, "error": "", "message": ""})


@app.post("/password-reset/{token}", response_class=HTMLResponse)
def password_reset_save(request: Request, token: str, new_password: str = Form(...), confirm_password: str = Form(...)):
    token_row = password_reset_user_for_token(token)
    if not token_row:
        return render(request, "password_reset_form.html", {"token": "", "error": "再設定リンクが無効、または期限切れです。", "message": ""})
    if not new_password:
        return render(request, "password_reset_form.html", {"token": token, "error": "新しいパスワードを入力してください。", "message": ""})
    if new_password != confirm_password:
        return render(request, "password_reset_form.html", {"token": token, "error": "新しいパスワードが一致しません。", "message": ""})
    now_s = app_now().isoformat(timespec="seconds")
    with db() as conn:
        before = conn.execute("SELECT id, username, role, company_id FROM users WHERE id=? AND company_id=?", (token_row["user_id"], token_row["company_id"])).fetchone()
        conn.execute("UPDATE users SET password_hash=? WHERE id=? AND company_id=?", (hash_password(new_password), token_row["user_id"], token_row["company_id"]))
        conn.execute("UPDATE password_reset_tokens SET used_at=? WHERE id=?", (now_s, token_row["id"]))
        audit_log(
            token_row["company_id"],
            {"id": token_row["user_id"], "role": row_value(before, "role", "")},
            "password_reset.complete",
            "users",
            token_row["user_id"],
            "Password reset completed",
            before=row_to_dict(before),
            conn=conn,
        )
        conn.commit()
    return render(request, "password_reset_form.html", {"token": "", "error": "", "message": "パスワードを変更しました。ログインしてください。"})


@app.get("/admin/password-resets", response_class=HTMLResponse)
def admin_password_resets(request: Request, user=Depends(require_admin)):
    company_id = company_id_for(user)
    rows = query_all(
        """SELECT t.*, u.name, u.username
           FROM password_reset_tokens t
           JOIN users u ON u.id=t.user_id AND u.company_id=t.company_id
           WHERE t.company_id=? AND COALESCE(t.used_at, '')='' AND t.expires_at>=?
           ORDER BY t.created_at DESC LIMIT 50""",
        (company_id, app_now().isoformat(timespec="seconds")),
    )
    return render(request, "admin_password_resets.html", {"rows": rows, "smtp_configured": smtp_configured()})


@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...)):
    user = query_one(
        """SELECT u.*, c.name AS company_name, c.code AS company_code, c.active AS company_active
           FROM users u
           LEFT JOIN companies c ON c.id=u.company_id
           WHERE u.username = ? AND u.active = 1""",
        (username,),
    )
    if not user or not verify_password(password, user["password_hash"]):
        return render(request, "login.html", {"error": "IDまたはパスワードが違います"})
    if row_value(user, "company_active", 1) == 0:
        return render(request, "login.html", {"error": "この会社アカウントは停止中です"})
    if needs_password_rehash(user["password_hash"]):
        execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password(password), user["id"]))
    res = RedirectResponse("/", status_code=303)
    res.set_cookie(
        "session",
        serializer.dumps(user["id"]),
        httponly=True,
        samesite="lax",
        secure=SESSION_COOKIE_SECURE,
        max_age=60 * 60 * 24 * 7,
    )
    return res


@app.get("/logout")
def logout():
    res = RedirectResponse("/login", status_code=303)
    res.delete_cookie("session")
    return res


@app.get("/admin", response_class=HTMLResponse)
def admin_dashboard(request: Request, user=Depends(require_admin)):
    if is_platform_admin(user):
        total_companies = query_one("SELECT COUNT(*) AS count FROM companies")["count"]
        active_contracts = query_one(
            "SELECT COUNT(*) AS count FROM companies WHERE active=1 AND COALESCE(status, '利用中')='利用中'"
        )["count"]
        monthly_revenue = int_value(row_value(query_one(
            """SELECT COALESCE(SUM(company_total), 0) AS total
               FROM (
                 SELECT c.id, COALESCE(SUM(CASE WHEN cf.enabled=1 THEN fp.monthly_price ELSE 0 END), 0) AS company_total
                 FROM companies c
                 LEFT JOIN company_features cf ON cf.company_id=c.id
                 LEFT JOIN feature_prices fp ON fp.feature_key=cf.feature_key AND fp.active=1
                 WHERE c.active=1 AND COALESCE(c.status, '利用中')='利用中'
                 GROUP BY c.id
               ) totals"""
        ), "total"))
        recent_companies = query_all(
            """SELECT c.*,
                      (SELECT username FROM users au
                       WHERE au.company_id=c.id AND au.role='admin' AND au.active=1
                       ORDER BY au.id LIMIT 1) AS admin_username,
                      COALESCE(ft.monthly_total, 0) AS monthly_total,
                      COALESCE(ft.enabled_count, 0) AS enabled_count
               FROM companies c
               LEFT JOIN (
                 SELECT cf.company_id,
                        COALESCE(SUM(CASE WHEN cf.enabled=1 THEN fp.monthly_price ELSE 0 END), 0) AS monthly_total,
                        COALESCE(SUM(CASE WHEN cf.enabled=1 THEN 1 ELSE 0 END), 0) AS enabled_count
                 FROM company_features cf
                 LEFT JOIN feature_prices fp ON fp.feature_key=cf.feature_key AND fp.active=1
                 GROUP BY cf.company_id
               ) ft ON ft.company_id=c.id
               ORDER BY c.id DESC
               LIMIT 5"""
        )
        plan_counts = {}
        plan_rows = query_all(
            """SELECT c.status, COALESCE(ft.monthly_total, 0) AS monthly_total, COALESCE(ft.enabled_count, 0) AS enabled_count
               FROM companies c
               LEFT JOIN (
                 SELECT cf.company_id,
                        COALESCE(SUM(CASE WHEN cf.enabled=1 THEN fp.monthly_price ELSE 0 END), 0) AS monthly_total,
                        COALESCE(SUM(CASE WHEN cf.enabled=1 THEN 1 ELSE 0 END), 0) AS enabled_count
                 FROM company_features cf
                 LEFT JOIN feature_prices fp ON fp.feature_key=cf.feature_key AND fp.active=1
                 GROUP BY cf.company_id
               ) ft ON ft.company_id=c.id"""
        )
        for company in plan_rows:
            company_status = row_value(company, "status", "利用中") or "利用中"
            plan_label = plan_label_for_company(company_status, int_value(row_value(company, "enabled_count")), int_value(row_value(company, "monthly_total")))
            plan_counts[plan_label] = plan_counts.get(plan_label, 0) + 1
        recent_requests = attach_image_info_lazy(
            query_all(
                """SELECT r.*, c.name AS company_name, u.name AS sender_name
                   FROM support_requests r
                   JOIN companies c ON c.id=r.company_id
                   JOIN users u ON u.id=r.sender_id
                   ORDER BY r.created_at DESC
                   LIMIT 10""",
            ),
            "image_path",
        )
        unresolved_count = query_one("SELECT COUNT(*) AS count FROM support_requests WHERE status='未対応'")["count"]
        return render(
            request,
            "platform_admin_dashboard.html",
            {
                "total_companies": total_companies,
                "active_contracts": active_contracts,
                "monthly_revenue": monthly_revenue,
                "unresolved_count": unresolved_count,
                "recent_companies": recent_companies,
                "recent_requests": recent_requests,
                "plan_counts": [{"label": key, "count": value} for key, value in sorted(plan_counts.items())],
                "pending_changes": pending_feature_changes(limit=20),
            },
        )

    today_s = app_today().isoformat()
    tomorrow_s = (app_today() + timedelta(days=1)).isoformat()
    day_after_s = (app_today() + timedelta(days=2)).isoformat()
    company_id = company_id_for(user)
    stats = query_one(
        """SELECT
             (SELECT COUNT(*) FROM users WHERE role='member' AND active=1 AND COALESCE(purged,0)=0 AND company_id=?) AS member_count,
             (SELECT COUNT(*) FROM work_logs WHERE work_date=? AND company_id=? AND COALESCE(is_deleted,0)=0) AS log_count,
             (SELECT COUNT(*) FROM deliveries WHERE work_date=? AND company_id=? AND COALESCE(is_deleted,0)=0) AS delivery_count""",
        (company_id, today_s, company_id, today_s, company_id),
    )
    logs = query_all(
        """SELECT w.*, u.name FROM work_logs w JOIN users u ON u.id = w.user_id
           WHERE w.work_date = ? AND w.company_id=? AND COALESCE(w.is_deleted,0)=0 ORDER BY w.logged_at DESC LIMIT 30""",
        (today_s, company_id),
    )
    issues = query_all(
        """SELECT v.*, u.name FROM vehicle_issues v JOIN users u ON u.id = v.user_id
           WHERE v.company_id=? ORDER BY v.created_at DESC LIMIT 5""",
        (company_id,),
    )
    deliveries = query_all(
        """SELECT d.*, u.name, COALESCE(s.file_path, d.inspection_sheet_path) AS slip_path
           FROM deliveries d
           JOIN users u ON u.id = d.user_id
           LEFT JOIN inspection_slips s ON s.delivery_id=d.id AND s.company_id=d.company_id
           WHERE d.work_date = ? AND d.company_id=? AND COALESCE(d.is_deleted,0)=0 ORDER BY u.name LIMIT 100""",
        (today_s, company_id),
    )
    deliveries = attach_image_info_lazy(deliveries, "slip_path")
    today_shift_rows = attach_shift_areas(
        query_all(
            """SELECT s.*, u.name, u.vehicle AS user_vehicle,
                      COALESCE(v.name, u.vehicle, '') AS vehicle_name,
                      v.plate_number,
                      COALESCE(p.name, '未設定拠点') AS depot_name,
                      d.name AS district_name,
                      t.name AS town_name
               FROM shifts s
               JOIN users u ON u.id=s.user_id AND u.company_id=s.company_id
               LEFT JOIN vehicles v ON v.id=u.last_vehicle_id AND v.company_id=u.company_id
               LEFT JOIN districts d ON d.id=s.district_id AND d.company_id=s.company_id
               LEFT JOIN depots p ON p.id=d.depot_id AND p.company_id=s.company_id
               LEFT JOIN towns t ON t.id=s.town_id AND t.company_id=s.company_id
               WHERE s.company_id=? AND s.shift_date=? AND s.decided=1
               ORDER BY depot_name, u.name LIMIT 200""",
            (company_id, today_s),
        )
    )
    today_shift_depots = group_rows_by_depot(today_shift_rows)
    mobile_delivery_status = query_all(
        """SELECT u.id, u.name,
                  MAX(CASE WHEN w.log_type='start' THEN 1 ELSE 0 END) AS has_start,
                  MAX(CASE WHEN w.log_type='end' THEN 1 ELSE 0 END) AS has_end,
                  MAX(d.id) AS delivery_id,
                  COALESCE(MAX(d.completed), 0) AS completed,
                  COALESCE(MAX(d.night), 0) AS night,
                  COALESCE(MAX(d.pickup), 0) AS pickup,
                  COALESCE(MAX(d.large), 0) AS large,
                  MAX(CASE WHEN COALESCE(sl.file_path, '')<>'' OR COALESCE(sh.file_path, '')<>'' OR COALESCE(d.inspection_sheet_path, '')<>'' THEN 1 ELSE 0 END) AS has_inspection
           FROM users u
           LEFT JOIN work_logs w ON w.user_id=u.id AND w.company_id=u.company_id AND w.work_date=? AND COALESCE(w.is_deleted,0)=0
           LEFT JOIN deliveries d ON d.user_id=u.id AND d.company_id=u.company_id AND d.work_date=? AND COALESCE(d.is_deleted,0)=0
           LEFT JOIN inspection_slips sl ON sl.user_id=u.id AND sl.company_id=u.company_id AND sl.slip_date=?
           LEFT JOIN inspection_sheets sh ON sh.user_id=u.id AND sh.company_id=u.company_id AND (sh.sheet_date=? OR COALESCE(sh.delivery_date, '')=?)
           WHERE u.company_id=? AND u.role='member' AND u.active=1 AND COALESCE(u.purged,0)=0
           GROUP BY u.id, u.name
           ORDER BY u.name LIMIT 200""",
        (today_s, today_s, today_s, today_s, today_s, company_id),
    )
    mobile_delivery_status = [dict(row) for row in mobile_delivery_status]
    session_status_rows = query_all(
        """SELECT user_id,
                  COUNT(*) AS session_count,
                  SUM(CASE WHEN status='working' THEN 1 ELSE 0 END) AS open_count,
                  SUM(CASE WHEN status='finished' THEN 1 ELSE 0 END) AS finished_count,
                  SUM(distance_km) AS distance_total,
                  MAX(CASE WHEN COALESCE(inspection_image_path, '')<>'' OR inspection_sheet_id IS NOT NULL THEN 1 ELSE 0 END) AS has_session_inspection
           FROM work_log_sessions
           WHERE company_id=? AND work_date=? AND COALESCE(is_deleted,0)=0
           GROUP BY user_id""",
        (company_id, today_s),
    )
    session_status_map = {row["user_id"]: row for row in session_status_rows}
    for row in mobile_delivery_status:
        session_info = session_status_map.get(row["id"])
        row["session_count"] = int_value(row_value(session_info, "session_count"))
        row["open_session_count"] = int_value(row_value(session_info, "open_count"))
        row["finished_session_count"] = int_value(row_value(session_info, "finished_count"))
        row["distance_total"] = int_value(row_value(session_info, "distance_total"))
        if row["session_count"]:
            row["has_start"] = 1
        if row["finished_session_count"]:
            row["has_end"] = 1
        if int_value(row_value(session_info, "has_session_inspection")):
            row["has_inspection"] = 1
    mobile_missing_status = {
        "not_started": [row for row in mobile_delivery_status if not int_value(row["has_start"])],
        "not_finished": [row for row in mobile_delivery_status if int_value(row.get("open_session_count")) or (int_value(row["has_start"]) and not int_value(row["has_end"]))],
        "delivery_missing": [row for row in mobile_delivery_status if not row_value(row, "delivery_id")],
        "roll_call_missing": [row for row in mobile_delivery_status if not int_value(row["has_start"]) or not int_value(row["has_end"])],
        "inspection_missing": [row for row in mobile_delivery_status if not int_value(row["has_inspection"])],
    }
    mobile_work_counts = {
        "member_total": len(mobile_delivery_status),
        "started": sum(1 for row in mobile_delivery_status if int_value(row["has_start"])),
        "working": sum(1 for row in mobile_delivery_status if int_value(row.get("open_session_count"))),
        "finished": sum(1 for row in mobile_delivery_status if int_value(row["has_end"])),
        "not_started": len(mobile_missing_status["not_started"]),
        "not_finished": len(mobile_missing_status["not_finished"]),
        "delivery_done": sum(1 for row in mobile_delivery_status if row_value(row, "delivery_id")),
    }
    delivery_depot_rows = query_all(
        """SELECT u.id AS user_id, u.name,
                  COALESCE(p.name, '未設定拠点') AS depot_name,
                  COALESCE(d.completed, 0) AS completed,
                  COALESCE(d.transfer, 0) AS transfer,
                  COALESCE(d.night, 0) AS night,
                  COALESCE(d.pickup, 0) AS pickup,
                  COALESCE(d.large, 0) AS large,
                  d.id AS delivery_id,
                  COALESCE(r.delivery_unit, 0) AS delivery_unit,
                  COALESCE(r.transfer_unit, 0) AS transfer_unit,
                  COALESCE(r.night_unit, 0) AS night_unit,
                  COALESCE(r.pickup_unit, 0) AS pickup_unit,
                  COALESCE(r.large_unit, 0) AS large_unit,
                  COALESCE(r.vehicle_rental_type, 'none') AS vehicle_rental_type,
                  COALESCE(r.vehicle_daily_fee, 0) AS vehicle_daily_fee
           FROM users u
           LEFT JOIN shifts s ON s.user_id=u.id AND s.company_id=u.company_id AND s.shift_date=? AND s.decided=1
           LEFT JOIN districts dist ON dist.id=s.district_id AND dist.company_id=s.company_id
           LEFT JOIN depots p ON p.id=dist.depot_id AND p.company_id=s.company_id
           LEFT JOIN deliveries d ON d.user_id=u.id AND d.company_id=u.company_id AND d.work_date=? AND COALESCE(d.is_deleted,0)=0
           LEFT JOIN rates r ON r.user_id=u.id AND r.company_id=u.company_id
           WHERE u.company_id=? AND u.role='member' AND u.active=1 AND COALESCE(u.purged,0)=0
           ORDER BY depot_name, u.name LIMIT 200""",
        (today_s, today_s, company_id),
    )
    delivery_depot_items = []
    for row in delivery_depot_rows:
        item = dict(row)
        gross = (
            int_value(item["completed"]) * int_value(item["delivery_unit"])
            + int_value(item["transfer"]) * int_value(item["transfer_unit"])
            + int_value(item["night"]) * int_value(item["night_unit"])
            + int_value(item["pickup"]) * int_value(item["pickup_unit"])
            + int_value(item["large"]) * int_value(item["large_unit"])
        )
        if item["delivery_id"] and row_value(item, "vehicle_rental_type") == "daily":
            gross -= int_value(item["vehicle_daily_fee"])
        item["reward_estimate"] = gross
        delivery_depot_items.append(item)
    delivery_depots = group_rows_by_depot(delivery_depot_items)
    for depot in delivery_depots:
        depot["totals"] = {
            "completed": sum(int_value(row["completed"]) for row in depot["items"]),
            "transfer": sum(int_value(row["transfer"]) for row in depot["items"]),
            "night": sum(int_value(row["night"]) for row in depot["items"]),
            "pickup": sum(int_value(row["pickup"]) for row in depot["items"]),
            "large": sum(int_value(row["large"]) for row in depot["items"]),
            "reward": sum(int_value(row["reward_estimate"]) for row in depot["items"]),
        }
    today_sessions = query_all(
        """SELECT s.*, u.name, v.plate_number
           FROM work_log_sessions s
           JOIN users u ON u.id=s.user_id AND u.company_id=s.company_id
           LEFT JOIN vehicles v ON v.id=s.vehicle_id AND v.company_id=s.company_id
           WHERE s.company_id=? AND s.work_date=? AND COALESCE(s.is_deleted,0)=0
           ORDER BY s.session_no, s.start_time LIMIT 120""",
        (company_id, today_s),
    )
    open_sessions = [session for session in today_sessions if row_value(session, "status") == "working"]
    member_session_totals = query_all(
        """SELECT u.name, s.user_id, COUNT(*) AS session_count, SUM(s.distance_km) AS distance_km,
                  SUM(s.delivery_count) AS completed, SUM(s.transfer_count) AS transfer,
                  SUM(s.night_count) AS night, SUM(s.pickup_count) AS pickup, SUM(s.large_count) AS large
           FROM work_log_sessions s
           JOIN users u ON u.id=s.user_id AND u.company_id=s.company_id
           WHERE s.company_id=? AND s.work_date=? AND COALESCE(s.is_deleted,0)=0
           GROUP BY s.user_id, u.name
           ORDER BY u.name LIMIT 200""",
        (company_id, today_s),
    )
    vehicle_session_totals = query_all(
        """SELECT COALESCE(NULLIF(s.vehicle_name, ''), '車両未設定') AS vehicle_name,
                  COUNT(*) AS session_count, SUM(s.distance_km) AS distance_km
           FROM work_log_sessions s
           WHERE s.company_id=? AND s.work_date=? AND COALESCE(s.is_deleted,0)=0
           GROUP BY COALESCE(NULLIF(s.vehicle_name, ''), '車両未設定')
           ORDER BY vehicle_name LIMIT 100""",
        (company_id, today_s),
    )
    session_corrections = query_all(
        """SELECT c.*, u.name AS member_name, a.name AS actor_name
           FROM work_log_corrections c
           JOIN users u ON u.id=c.user_id AND u.company_id=c.company_id
           LEFT JOIN users a ON a.id=c.actor_id
           WHERE c.company_id=?
           ORDER BY c.changed_at DESC LIMIT 30""",
        (company_id,),
    )
    frequent_corrections = query_all(
        """SELECT u.name, COUNT(*) AS correction_count
           FROM work_log_corrections c
           JOIN users u ON u.id=c.user_id AND u.company_id=c.company_id
           WHERE c.company_id=? AND c.changed_at>=?
           GROUP BY c.user_id, u.name
           ORDER BY correction_count DESC, u.name LIMIT 10""",
        (company_id, (app_today() - timedelta(days=30)).isoformat()),
    )
    upcoming_shifts = attach_shift_areas(
        query_all(
            """SELECT s.*, u.name, u.vehicle AS user_vehicle,
                      COALESCE(v.name, u.vehicle, '') AS vehicle_name,
                      v.plate_number,
                      COALESCE(p.name, '未設定拠点') AS depot_name,
                      d.name AS district_name,
                      t.name AS town_name
               FROM shifts s
               JOIN users u ON u.id=s.user_id AND u.company_id=s.company_id
               LEFT JOIN vehicles v ON v.id=u.last_vehicle_id AND v.company_id=u.company_id
               LEFT JOIN districts d ON d.id=s.district_id AND d.company_id=s.company_id
               LEFT JOIN depots p ON p.id=d.depot_id AND p.company_id=s.company_id
               LEFT JOIN towns t ON t.id=s.town_id AND t.company_id=s.company_id
               WHERE s.company_id=? AND s.shift_date BETWEEN ? AND ? AND s.decided=1
               ORDER BY s.shift_date, depot_name, u.name LIMIT 120""",
            (company_id, tomorrow_s, day_after_s),
        )
    )
    shifts_by_mobile_day = []
    for day_s in (tomorrow_s, day_after_s):
        day_items = [shift for shift in upcoming_shifts if shift["shift_date"] == day_s]
        shifts_by_mobile_day.append({"date": day_s, "items": day_items, "depots": group_rows_by_depot(day_items)})
    vehicle_alerts = []
    for vehicle in vehicle_rows_for_company(company_id):
        info = vehicle_alert(vehicle)
        if info["level"]:
            item = dict(vehicle)
            item["alert"] = info
            vehicle_alerts.append(item)
    open_vehicle_issues = query_all(
        """SELECT v.*, u.name FROM vehicle_issues v
           JOIN users u ON u.id=v.user_id
           WHERE v.company_id=? AND v.status IN ('未対応','確認中','修理依頼中')
           ORDER BY CASE v.severity WHEN '高' THEN 1 WHEN '中' THEN 2 ELSE 3 END, v.created_at DESC LIMIT 10""",
        (company_id,),
    )
    holiday_status = holiday_submission_status(company_id)
    notifications = []
    deadline_alert = row_value(holiday_status, "deadline_alert", {}) if isinstance(holiday_status, dict) else {}
    if deadline_alert and deadline_alert.get("message"):
        notifications.append({"level": deadline_alert.get("level", "yellow"), "title": "休み希望締切", "body": deadline_alert["message"], "time": today_s})
    for vehicle in vehicle_alerts[:8]:
        alert = vehicle["alert"]
        if alert.get("inspection_level"):
            notifications.append({"level": alert["inspection_level"], "title": "車検期限", "body": f"{vehicle['name']} / {alert['inspection_message']}", "time": today_s})
        if alert.get("oil_level"):
            notifications.append({"level": alert["oil_level"], "title": "オイル交換時期", "body": f"{vehicle['name']} / {alert['oil_message']}", "time": today_s})
    health_logs = query_all(
        """SELECT w.*, u.name FROM work_logs w
           JOIN users u ON u.id=w.user_id
           WHERE w.company_id=? AND w.work_date=? AND w.health_status IN ('注意','不調')
           ORDER BY w.logged_at DESC LIMIT 10""",
        (company_id, today_s),
    )
    for log in health_logs:
        notifications.append({"level": "red" if log["health_status"] == "不調" else "yellow", "title": "体調不良報告", "body": f"{log['name']} / {log['health_status']}", "time": log["logged_at"]})
    for issue in open_vehicle_issues:
        notifications.append({"level": "red" if issue["severity"] == "高" else "yellow", "title": "不具合報告", "body": f"{issue['vehicle_name']} / {issue['detail']}", "time": issue["created_at"]})
    holiday_notice_rows = query_all(
        """SELECT h.*, u.name FROM holiday_requests h
           JOIN users u ON u.id=h.user_id
           WHERE h.company_id=? AND h.request_date>=?
           ORDER BY h.created_at DESC LIMIT 10""",
        (company_id, today_s),
    )
    for holiday in holiday_notice_rows:
        notifications.append({"level": "yellow", "title": "休み希望", "body": f"{holiday['name']} / {holiday['request_date']}", "time": holiday["created_at"]})
    support_rows = query_all(
        """SELECT r.*, u.name AS sender_name FROM support_requests r
           JOIN users u ON u.id=r.sender_id
           WHERE r.company_id=? AND r.status IN ('未対応','対応中')
           ORDER BY CASE r.urgency WHEN '緊急' THEN 1 WHEN '高' THEN 2 WHEN '中' THEN 3 ELSE 4 END, r.created_at DESC LIMIT 10""",
        (company_id,),
    )
    for req in support_rows:
        notifications.append({"level": "red" if req["urgency"] in {"緊急", "高"} else "yellow", "title": req["request_type"], "body": f"{req['subject']} / {req['sender_name']}", "time": req["created_at"]})
    rank = {"red": 0, "orange": 1, "yellow": 2}
    notifications = sorted(notifications, key=lambda item: item.get("time", ""), reverse=True)
    notifications = sorted(notifications, key=lambda item: rank.get(item["level"], 9))[:12]
    return render(
        request,
        "company_admin_dashboard.html",
        {
            "member_count": stats["member_count"],
            "log_count": stats["log_count"],
            "delivery_count": stats["delivery_count"],
            "logs": logs,
            "issues": issues,
            "deliveries": deliveries,
            "vehicle_alerts": vehicle_alerts[:5],
            "holiday_status": holiday_status,
            "today_shift_depots": today_shift_depots,
            "delivery_depots": delivery_depots,
            "mobile_delivery_status": mobile_delivery_status,
            "mobile_missing_status": mobile_missing_status,
            "mobile_work_counts": mobile_work_counts,
            "today_sessions": today_sessions,
            "open_sessions": open_sessions,
            "member_session_totals": member_session_totals,
            "vehicle_session_totals": vehicle_session_totals,
            "session_corrections": session_corrections,
            "frequent_corrections": frequent_corrections,
            "mobile_shift_days": shifts_by_mobile_day,
            "open_vehicle_issues": open_vehicle_issues,
            "priority_notifications": notifications,
        },
    )


@app.get("/admin/members", response_class=HTMLResponse)
def members_page(request: Request, message: str = "", error: str = "", user=Depends(require_admin)):
    require_feature(user, "members")
    company_id = company_id_for(user)
    active_members = query_all("SELECT * FROM users WHERE role = 'member' AND active = 1 AND COALESCE(purged,0)=0 AND company_id=? ORDER BY id LIMIT 200", (company_id,))
    retired_members = query_all("SELECT * FROM users WHERE role = 'member' AND active = 0 AND COALESCE(purged,0)=0 AND company_id=? ORDER BY id LIMIT 200", (company_id,))
    return render(request, "members.html", {"active_members": active_members, "retired_members": retired_members, "message": message, "error": error})


@app.post("/admin/members")
def save_member(
    member_id: str = Form(""),
    name: str = Form(...),
    username: str = Form(...),
    password: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    vehicle: str = Form(""),
    user=Depends(require_admin),
):
    require_feature(user, "members")
    company_id = company_id_for(user)
    now = app_now().isoformat(timespec="seconds")
    username = username.strip()
    email = email.strip()
    if not username:
        return RedirectResponse(f"/admin/members?error={quote('ログインIDを入力してください')}", status_code=303)
    release_purged_username(username)
    if member_id:
        duplicate = query_one("SELECT id FROM users WHERE username=? AND id<>? AND COALESCE(purged,0)=0", (username, member_id))
        if duplicate:
            return RedirectResponse(f"/admin/members?error={quote('同じログインIDが既に使われています')}", status_code=303)
        if password:
            execute(
                "UPDATE users SET name=?, username=?, password_hash=?, email=?, phone=?, vehicle=? WHERE id=? AND role='member' AND company_id=? AND COALESCE(purged,0)=0",
                (name, username, hash_password(password), email, phone, vehicle, member_id, company_id),
            )
        else:
            execute("UPDATE users SET name=?, username=?, email=?, phone=?, vehicle=? WHERE id=? AND role='member' AND company_id=? AND COALESCE(purged,0)=0", (name, username, email, phone, vehicle, member_id, company_id))
        ensure_vehicle_for_user(company_id, int(member_id), vehicle)
        return RedirectResponse("/admin/members?message=" + quote("メンバー情報を更新しました"), status_code=303)
    existing = query_one("SELECT * FROM users WHERE username=? AND COALESCE(purged,0)=0", (username,))
    if existing and row_value(existing, "role") == "member" and existing["active"] == 1:
        return RedirectResponse(f"/admin/members?error={quote('同じログインIDの在籍中メンバーがいます')}", status_code=303)
    if existing and row_value(existing, "role") == "member" and existing["active"] == 0:
        return RedirectResponse(f"/admin/members?error={quote('同じログインIDの退職済みメンバーがいます。完全削除後に再登録できます')}", status_code=303)
    if existing:
        return RedirectResponse(f"/admin/members?error={quote('同じログインIDが既に使われています')}", status_code=303)
    try:
        cur = execute(
            "INSERT INTO users(company_id, name, username, password_hash, role, email, phone, vehicle, created_at) VALUES (?,?,?,?,?,?,?,?,?)",
            (company_id, name, username, hash_password(password or DEFAULT_MEMBER_PASSWORD), "member", email, phone, vehicle, now),
        )
        execute("INSERT OR IGNORE INTO rates(user_id, company_id) VALUES (?, ?)", (cur.lastrowid, company_id))
        ensure_vehicle_for_user(company_id, cur.lastrowid, vehicle)
    except DB_INTEGRITY_ERROR:
        return RedirectResponse(f"/admin/members?error={quote('ログインIDが重複しています。別のIDを入力してください')}", status_code=303)
    return RedirectResponse("/admin/members?message=" + quote("メンバーを追加しました"), status_code=303)


@app.post("/admin/members/{member_id}/delete")
def delete_member(member_id: int, user=Depends(require_admin)):
    require_feature(user, "members")
    execute("UPDATE users SET active = 0 WHERE id=? AND role='member' AND company_id=?", (member_id, company_id_for(user)))
    return RedirectResponse("/admin/members", status_code=303)


@app.post("/admin/members/{member_id}/purge")
def purge_retired_member(member_id: int, user=Depends(require_admin)):
    require_feature(user, "members")
    company_id = company_id_for(user)
    member = query_one("SELECT * FROM users WHERE id=? AND role='member' AND company_id=?", (member_id, company_id))
    if not member:
        return RedirectResponse(f"/admin/members?error={quote('メンバーが見つかりません')}", status_code=303)
    if member["active"] == 1:
        return RedirectResponse(f"/admin/members?error={quote('在籍中メンバーは完全削除できません。先に退職扱いにしてください')}", status_code=303)
    now_dt = app_now()
    now = now_dt.isoformat(timespec="seconds")
    anonymized_username = deleted_username_for(member_id)
    execute(
        """UPDATE users
           SET name=?, username=?, password_hash=?, phone='', vehicle='', active=0, purged=1, deleted_at=?
           WHERE id=? AND role='member' AND active=0 AND company_id=? AND COALESCE(purged,0)=0""",
        (f"削除済みメンバー{member_id}", anonymized_username, hash_password(anonymized_username), now, member_id, company_id),
    )
    return RedirectResponse("/admin/members?message=" + quote("退職済みメンバーを完全削除しました。過去記録は匿名化された名前で保持されます"), status_code=303)


@app.get("/admin/settings", response_class=HTMLResponse)
def admin_settings_page(
    request: Request,
    message: str = "",
    error: str = "",
    feature_company_id: Optional[int] = None,
    user=Depends(require_admin),
):
    company_id = company_id_for(user)
    can_manage = can_manage_companies(user)
    members = query_all(
        """SELECT id, name, username, email, active
           FROM users
           WHERE role='member' AND company_id=? AND COALESCE(purged,0)=0
           ORDER BY active DESC, name
           LIMIT 200""",
        (company_id,),
    )
    companies = []
    if can_manage:
        companies = query_all(
            """SELECT c.*,
                      (SELECT username FROM users au
                       WHERE au.company_id=c.id AND au.role='admin' AND au.active=1
                       ORDER BY au.id LIMIT 1) AS admin_username,
                      (SELECT email FROM users au
                       WHERE au.company_id=c.id AND au.role='admin' AND au.active=1
                       ORDER BY au.id LIMIT 1) AS admin_email
               FROM companies c
               ORDER BY c.id
               LIMIT 200"""
        )
    feature_companies = companies if can_manage else query_all("SELECT id, name FROM companies WHERE id=?", (company_id,))
    feature_target_id = company_id
    if can_manage and feature_company_id:
        selected_company = query_one("SELECT id FROM companies WHERE id=?", (feature_company_id,))
        if selected_company:
            feature_target_id = feature_company_id
    feature_target_company = query_one("SELECT id, name FROM companies WHERE id=?", (feature_target_id,))
    return render(
        request,
        "admin_settings.html",
        {
            "members": members,
            "companies": companies,
            "can_manage_companies": can_manage,
            "feature_companies": feature_companies,
            "feature_target_company": feature_target_company,
            "feature_target_id": feature_target_id,
            "feature_rows": company_feature_rows(feature_target_id),
            "pending_changes": pending_feature_changes(feature_target_id, limit=100),
            "feature_prices": feature_price_rows(),
            "company_monthly_amount": monthly_amount_for_company(feature_target_id),
            "company_statuses": COMPANY_STATUSES,
            "message": message,
            "error": error,
        },
    )


@app.post("/admin/settings/account")
def update_admin_account(
    username: str = Form(...),
    password: str = Form(""),
    email: str = Form(""),
    user=Depends(require_admin),
):
    username = username.strip()
    email = email.strip()
    if not username:
        return RedirectResponse("/admin/settings?error=" + quote("ログインIDを入力してください"), status_code=303)
    release_purged_username(username)
    duplicate = query_one("SELECT id FROM users WHERE username=? AND id<>? AND COALESCE(purged,0)=0", (username, user["id"]))
    if duplicate:
        return RedirectResponse("/admin/settings?error=" + quote("同じログインIDが既に使われています"), status_code=303)
    if password:
        execute("UPDATE users SET username=?, password_hash=?, email=? WHERE id=? AND role='admin'", (username, hash_password(password), email, user["id"]))
    else:
        execute("UPDATE users SET username=?, email=? WHERE id=? AND role='admin'", (username, email, user["id"]))
    return RedirectResponse("/admin/settings?message=" + quote("管理者アカウントを更新しました"), status_code=303)


@app.post("/admin/settings/member")
def update_member_login(
    member_id: int = Form(...),
    username: str = Form(...),
    password: str = Form(""),
    email: str = Form(""),
    user=Depends(require_admin),
):
    require_feature(user, "members")
    company_id = company_id_for(user)
    username = username.strip()
    email = email.strip()
    if not username:
        return RedirectResponse("/admin/settings?error=" + quote("メンバーのログインIDを入力してください"), status_code=303)
    release_purged_username(username)
    member = query_one(
        "SELECT id FROM users WHERE id=? AND role='member' AND company_id=? AND COALESCE(purged,0)=0",
        (member_id, company_id),
    )
    if not member:
        return RedirectResponse("/admin/settings?error=" + quote("メンバーが見つかりません"), status_code=303)
    duplicate = query_one("SELECT id FROM users WHERE username=? AND id<>? AND COALESCE(purged,0)=0", (username, member_id))
    if duplicate:
        return RedirectResponse("/admin/settings?error=" + quote("同じログインIDが既に使われています"), status_code=303)
    if password:
        execute("UPDATE users SET username=?, password_hash=?, email=? WHERE id=? AND role='member' AND company_id=?", (username, hash_password(password), email, member_id, company_id))
    else:
        execute("UPDATE users SET username=?, email=? WHERE id=? AND role='member' AND company_id=?", (username, email, member_id, company_id))
    return RedirectResponse("/admin/settings?message=" + quote("メンバーのログイン情報を更新しました"), status_code=303)


@app.post("/admin/settings/company")
def save_company(
    company_id: str = Form(""),
    name: str = Form(...),
    code: str = Form(...),
    admin_username: str = Form(...),
    admin_email: str = Form(""),
    admin_password: str = Form(""),
    status: str = Form("利用中"),
    contract_start_date: str = Form(""),
    next_renewal_date: str = Form(""),
    memo: str = Form(""),
    user=Depends(require_admin),
):
    if not can_manage_companies(user):
        raise HTTPException(status_code=403, detail="会社管理の権限がありません")

    cid = int(company_id) if company_id.strip().isdigit() else 0
    name = name.strip()
    code = code.strip()
    admin_username = admin_username.strip()
    admin_email = admin_email.strip()
    release_purged_username(admin_username)
    if status not in COMPANY_STATUSES:
        status = "利用中"
    active_int = 0 if status == "停止中" else 1
    contract_start_date = contract_start_date.strip() or app_today().isoformat()
    next_renewal_date = next_renewal_date.strip() or (app_today() + timedelta(days=30)).isoformat()
    memo = memo.strip()
    admin_name = f"{name} 管理者" if name else "管理者"

    if not name or not code or not admin_username:
        return RedirectResponse("/admin/settings?error=" + quote("会社名、会社コード、管理者ログインIDを入力してください"), status_code=303)
    if cid == company_id_for(user) and active_int == 0:
        return RedirectResponse("/admin/settings?error=" + quote("ログイン中の会社は停止できません"), status_code=303)

    admin = None
    if cid:
        existing_company = query_one("SELECT id FROM companies WHERE id=?", (cid,))
        if not existing_company:
            return RedirectResponse("/admin/settings?error=" + quote("会社が見つかりません"), status_code=303)
        admin = query_one(
            "SELECT id FROM users WHERE company_id=? AND role='admin' ORDER BY id LIMIT 1",
            (cid,),
        )

    if not cid and not admin_password:
        return RedirectResponse("/admin/settings?error=" + quote("新しい会社の管理者パスワードを入力してください"), status_code=303)

    duplicate_code = query_one("SELECT id FROM companies WHERE code=? AND id<>?", (code, cid or 0))
    if duplicate_code:
        return RedirectResponse("/admin/settings?error=" + quote("同じ会社コードが既に使われています"), status_code=303)

    ignore_admin_id = admin["id"] if admin else 0
    duplicate_username = query_one("SELECT id FROM users WHERE username=? AND id<>? AND COALESCE(purged,0)=0", (admin_username, ignore_admin_id))
    if duplicate_username:
        return RedirectResponse("/admin/settings?error=" + quote("同じログインIDが既に使われています"), status_code=303)

    now = app_now().isoformat(timespec="seconds")
    if cid:
        execute(
            "UPDATE companies SET name=?, code=?, active=?, status=?, contract_start_date=?, next_renewal_date=?, memo=? WHERE id=?",
            (name, code, active_int, status, contract_start_date, next_renewal_date, memo, cid),
        )
        if admin:
            if admin_password:
                execute(
                    "UPDATE users SET name=?, username=?, password_hash=?, active=1 WHERE id=? AND role='admin' AND company_id=?",
                    (admin_name, admin_username, hash_password(admin_password), admin["id"], cid),
                )
                execute("UPDATE users SET email=? WHERE id=? AND role='admin' AND company_id=?", (admin_email, admin["id"], cid))
            else:
                execute(
                    "UPDATE users SET name=?, username=?, email=?, active=1 WHERE id=? AND role='admin' AND company_id=?",
                    (admin_name, admin_username, admin_email, admin["id"], cid),
                )
        else:
            if not admin_password:
                return RedirectResponse("/admin/settings?error=" + quote("管理者が未登録です。パスワードを入力してください"), status_code=303)
            execute(
                "INSERT INTO users(company_id, name, username, password_hash, role, email, phone, vehicle, active, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (cid, admin_name, admin_username, hash_password(admin_password), "admin", admin_email, "", "", 1, now),
            )
        return RedirectResponse("/admin/settings?message=" + quote("会社情報を更新しました"), status_code=303)

    cur = execute(
        "INSERT INTO companies(name, code, active, status, contract_start_date, next_renewal_date, memo, created_at) VALUES (?,?,?,?,?,?,?,?)",
        (name, code, active_int, status, contract_start_date, next_renewal_date, memo, now),
    )
    new_company_id = cur.lastrowid
    execute(
        "INSERT INTO users(company_id, name, username, password_hash, role, email, phone, vehicle, active, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (new_company_id, admin_name, admin_username, hash_password(admin_password), "admin", admin_email, "", "", 1, now),
    )
    with db() as conn:
        seed_company_features(conn, new_company_id, 1)
        conn.execute(
            "INSERT OR IGNORE INTO company_vehicle_rates(company_id, daily_fee, monthly_fee, updated_at) VALUES (?, 1500, 30000, ?)",
            (new_company_id, now),
        )
        audit_log(
            company_id_for(user),
            user,
            "company.create",
            "companies",
            new_company_id,
            "Company created",
            after={"name": name, "code": code, "status": status},
            conn=conn,
        )
        conn.commit()
    return RedirectResponse("/admin/settings?message=" + quote("会社と管理者を追加しました"), status_code=303)


@app.post("/admin/settings/features")
def update_company_features(
    feature_company_id: int = Form(...),
    features: Optional[List[str]] = Form(None),
    immediate: str = Form(""),
    user=Depends(require_admin),
):
    own_company_id = company_id_for(user)
    platform = is_platform_admin(user)
    target_company_id = feature_company_id if platform else own_company_id
    if not platform and feature_company_id != own_company_id:
        raise HTTPException(status_code=403, detail="他社の機能設定は変更できません")
    company = query_one("SELECT * FROM companies WHERE id=?", (target_company_id,))
    if not company:
        return RedirectResponse("/admin/settings?error=" + quote("会社が見つかりません"), status_code=303)
    selected = set(features or [])
    unknown = selected - FEATURE_KEYS
    if unknown:
        return RedirectResponse("/admin/settings?error=" + quote("不明な機能が含まれています"), status_code=303)
    now = app_now().isoformat(timespec="seconds")
    if platform and immediate != "schedule":
        with db() as conn:
            for feature in FEATURE_CATALOG:
                enabled = 1 if feature["key"] in selected else 0
                conn.execute(
                    """INSERT INTO company_features(company_id, feature_key, enabled, updated_at)
                       VALUES (?, ?, ?, ?)
                       ON CONFLICT(company_id, feature_key) DO UPDATE SET
                       enabled=excluded.enabled, updated_at=excluded.updated_at""",
                    (target_company_id, feature["key"], enabled, now),
                )
            conn.commit()
        clear_feature_cache(target_company_id)
        message = "利用機能・プラン設定を即時更新しました"
    else:
        effective_date = row_value(company, "next_renewal_date", "") or (app_today() + timedelta(days=30)).isoformat()
        with db() as conn:
            for feature in FEATURE_CATALOG:
                enabled = 1 if feature["key"] in selected else 0
                existing = conn.execute(
                    """SELECT id FROM company_feature_changes
                       WHERE company_id=? AND feature_key=? AND status='pending'
                       ORDER BY created_at DESC LIMIT 1""",
                    (target_company_id, feature["key"]),
                ).fetchone()
                if existing:
                    conn.execute(
                        """UPDATE company_feature_changes
                           SET enabled=?, effective_date=?, requested_by=?, created_at=?, applied_at=''
                           WHERE id=?""",
                        (enabled, effective_date, user["id"], now, existing["id"]),
                    )
                else:
                    conn.execute(
                        """INSERT INTO company_feature_changes(company_id, feature_key, enabled, effective_date, requested_by, status, created_at)
                           VALUES (?, ?, ?, ?, ?, 'pending', ?)""",
                        (target_company_id, feature["key"], enabled, effective_date, user["id"], now),
                    )
            conn.commit()
        message = "次回更新で反映予定の機能変更を保存しました"
    return RedirectResponse(
        f"/admin/settings?feature_company_id={target_company_id}&message=" + quote(message),
        status_code=303,
    )


@app.post("/admin/settings/features/apply-pending")
def apply_company_feature_changes(feature_company_id: int = Form(...), user=Depends(require_platform_admin)):
    applied = apply_pending_feature_changes(feature_company_id)
    return RedirectResponse(
        f"/admin/settings?feature_company_id={feature_company_id}&message=" + quote(f"{applied}件の予定変更を即時反映しました"),
        status_code=303,
    )


@app.post("/admin/feature-prices")
def update_feature_prices(
    feature_key: List[str] = Form(...),
    monthly_price: List[int] = Form(...),
    note: Optional[List[str]] = Form(None),
    user=Depends(require_platform_admin),
):
    now = app_now().isoformat(timespec="seconds")
    notes = list(note or [])
    with db() as conn:
        for index, key in enumerate(feature_key):
            if key not in FEATURE_KEYS:
                continue
            price = max(int(monthly_price[index]), 0) if index < len(monthly_price) else 0
            item_note = notes[index] if index < len(notes) else ""
            conn.execute(
                "UPDATE feature_prices SET monthly_price=?, note=?, updated_at=? WHERE feature_key=?",
                (price, item_note, now, key),
            )
        conn.commit()
    clear_feature_cache()
    return RedirectResponse("/admin/settings?message=" + quote("機能料金を更新しました"), status_code=303)


@app.get("/support-requests", response_class=HTMLResponse)
def support_request_page(request: Request, message: str = "", error: str = "", user=Depends(require_admin)):
    if is_platform_admin(user):
        return RedirectResponse("/admin/support-requests", status_code=303)
    rows = attach_image_info(
        query_all(
            """SELECT r.*, c.name AS company_name, u.name AS sender_name
               FROM support_requests r
               JOIN companies c ON c.id=r.company_id
               JOIN users u ON u.id=r.sender_id
               WHERE r.company_id=?
               ORDER BY r.created_at DESC
               LIMIT 100""",
            (company_id_for(user),),
        ),
        "image_path",
    )
    return render(
        request,
        "support_requests.html",
        {
            "requests": rows,
            "support_types": SUPPORT_TYPES,
            "support_urgencies": SUPPORT_URGENCIES,
            "support_statuses": SUPPORT_STATUSES,
            "platform_view": False,
            "message": message,
            "error": error,
        },
    )


@app.post("/support-requests")
async def create_support_request(
    request_type: str = Form(...),
    subject: str = Form(...),
    body: str = Form(...),
    urgency: str = Form("中"),
    image: Optional[UploadFile] = File(None),
    user=Depends(require_admin),
):
    if is_platform_admin(user):
        raise HTTPException(status_code=403, detail="会社管理者として送信してください")
    if request_type not in SUPPORT_TYPES:
        request_type = "不具合"
    if urgency not in SUPPORT_URGENCIES:
        urgency = "中"
    subject = subject.strip()
    body = body.strip()
    if not subject or not body:
        return RedirectResponse("/support-requests?error=" + quote("件名と内容を入力してください"), status_code=303)
    image_path = ""
    if image and image.filename:
        if not image.content_type or not image.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="添付は画像ファイルを選択してください")
        data = await image.read()
        if data:
            if len(data) > 10 * 1024 * 1024:
                raise HTTPException(status_code=400, detail="画像サイズは10MB以内にしてください")
            suffix = safe_upload_name(image.filename)
            filename = f"support_company{company_id_for(user)}_user{user['id']}_{app_now().strftime('%Y%m%d%H%M%S%f')}{suffix}"
            image_path, _ = save_inspection_image(data, filename, image.content_type or "", "support_requests", SUPPORT_UPLOAD_DIR)
    now = app_now().isoformat(timespec="seconds")
    execute(
        """INSERT INTO support_requests(company_id, sender_id, request_type, subject, body, urgency, image_path, status, created_at, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, '未対応', ?, ?)""",
        (company_id_for(user), user["id"], request_type, subject, body, urgency, image_path, now, now),
    )
    return RedirectResponse("/support-requests?message=" + quote("不具合・改善要望を送信しました"), status_code=303)


@app.get("/admin/support-requests", response_class=HTMLResponse)
def platform_support_requests(request: Request, message: str = "", error: str = "", user=Depends(require_platform_admin)):
    rows = attach_image_info(
        query_all(
            """SELECT r.*, c.name AS company_name, u.name AS sender_name
               FROM support_requests r
               JOIN companies c ON c.id=r.company_id
               JOIN users u ON u.id=r.sender_id
               ORDER BY r.created_at DESC
               LIMIT 200"""
        ),
        "image_path",
    )
    return render(
        request,
        "support_requests.html",
        {
            "requests": rows,
            "support_types": SUPPORT_TYPES,
            "support_urgencies": SUPPORT_URGENCIES,
            "support_statuses": SUPPORT_STATUSES,
            "platform_view": True,
            "message": message,
            "error": error,
        },
    )


@app.post("/admin/support-requests/{request_id}/update")
def update_support_request(request_id: int, status: str = Form(...), reply_note: str = Form(""), user=Depends(require_platform_admin)):
    if status not in SUPPORT_STATUSES:
        status = "未対応"
    execute(
        "UPDATE support_requests SET status=?, reply_note=?, updated_at=? WHERE id=?",
        (status, reply_note.strip(), app_now().isoformat(timespec="seconds"), request_id),
    )
    return RedirectResponse("/admin/support-requests?message=" + quote("状態と返信メモを更新しました"), status_code=303)


@app.get("/settings", response_class=HTMLResponse)
def member_settings_page(request: Request, message: str = "", error: str = "", user=Depends(require_user)):
    return render(request, "member_settings.html", {"message": message, "error": error})


@app.post("/settings/password")
def update_own_password(
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    user=Depends(require_user),
):
    if not verify_password(current_password, user["password_hash"]):
        return RedirectResponse("/settings?error=" + quote("現在のパスワードが違います"), status_code=303)
    if not new_password:
        return RedirectResponse("/settings?error=" + quote("新しいパスワードを入力してください"), status_code=303)
    if new_password != confirm_password:
        return RedirectResponse("/settings?error=" + quote("新しいパスワードが一致しません"), status_code=303)
    execute("UPDATE users SET password_hash=? WHERE id=?", (hash_password(new_password), user["id"]))
    return RedirectResponse("/settings?message=" + quote("パスワードを変更しました"), status_code=303)


@app.get("/my-room", response_class=HTMLResponse)
def my_room_page(request: Request, message: str = "", error: str = "", user=Depends(require_user)):
    if user["role"] == "admin":
        return RedirectResponse("/admin", status_code=303)
    company_id = company_id_for(user)
    rates = query_one("SELECT * FROM rates WHERE user_id=? AND company_id=?", (user["id"], company_id))
    month_summary = monthly_reward_summary(user["id"], app_today(), company_id)
    attendance = query_one(
        """SELECT COUNT(*) AS count FROM (
             SELECT work_date FROM work_logs WHERE user_id=? AND company_id=? AND COALESCE(is_deleted,0)=0
             UNION
             SELECT work_date FROM work_log_sessions WHERE user_id=? AND company_id=? AND COALESCE(is_deleted,0)=0
           ) days""",
        (user["id"], company_id, user["id"], company_id),
    )
    sessions = query_all(
        """SELECT s.*, v.plate_number FROM work_log_sessions s
           LEFT JOIN vehicles v ON v.id=s.vehicle_id AND v.company_id=s.company_id
           WHERE s.user_id=? AND s.company_id=? AND COALESCE(s.is_deleted,0)=0
           ORDER BY s.work_date DESC, s.session_no DESC LIMIT 12""",
        (user["id"], company_id),
    )
    corrections = query_all(
        """SELECT c.*, a.name AS actor_name FROM work_log_corrections c
           LEFT JOIN users a ON a.id=c.actor_id
           WHERE c.user_id=? AND c.company_id=?
           ORDER BY c.changed_at DESC LIMIT 20""",
        (user["id"], company_id),
    )
    vehicles, selected_vehicle = selected_vehicle_context(user)
    return render(
        request,
        "my_room.html",
        {
            "rates": rates,
            "month_summary": month_summary,
            "attendance_days": row_value(attendance, "count", 0),
            "sessions": sessions,
            "corrections": corrections,
            "vehicles": vehicles,
            "selected_vehicle": selected_vehicle,
            "work_state": today_work_state(user["id"], company_id, app_today().isoformat()),
            "message": message,
            "error": error,
        },
    )


@app.post("/my-room")
def update_my_room(
    phone: str = Form(""),
    email: str = Form(""),
    notification_email: str = Form(""),
    vehicle_id: int = Form(0),
    member_notes: str = Form(""),
    user=Depends(require_user),
):
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとしてログインしてください")
    company_id = company_id_for(user)
    before = query_one("SELECT * FROM users WHERE id=? AND company_id=?", (user["id"], company_id))
    vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=? AND active=1", (vehicle_id, company_id)) if vehicle_id else None
    with db() as conn:
        conn.execute(
            """UPDATE users SET phone=?, email=?, notification_email=?, last_vehicle_id=?,
               vehicle=COALESCE(NULLIF(?, ''), vehicle), member_notes=?
               WHERE id=? AND company_id=?""",
            (phone.strip(), email.strip(), notification_email.strip(), vehicle["id"] if vehicle else row_value(before, "last_vehicle_id"), row_value(vehicle, "name", ""), member_notes.strip(), user["id"], company_id),
        )
        after = conn.execute("SELECT * FROM users WHERE id=? AND company_id=?", (user["id"], company_id)).fetchone()
        audit_log(company_id, user, "member.my_room_update", "users", user["id"], "Member profile updated", before=row_to_dict(before), after=row_to_dict(after), conn=conn)
        conn.commit()
    return RedirectResponse("/my-room?message=" + quote("マイルームを更新しました"), status_code=303)


@app.post("/my-room/work-sessions/{session_id}/update")
def member_update_work_session(
    session_id: int,
    start_time: str = Form(""),
    end_time: str = Form(""),
    vehicle_id: int = Form(0),
    start_odometer: int = Form(0),
    end_odometer: int = Form(0),
    alcohol_check_start: str = Form(""),
    alcohol_check_end: str = Form(""),
    call_check_start: str = Form(""),
    call_check_end: str = Form(""),
    delivery_count: int = Form(0),
    transfer_count: int = Form(0),
    night_count: int = Form(0),
    pickup_count: int = Form(0),
    large_count: int = Form(0),
    inspection_image_path: str = Form(""),
    notes: str = Form(""),
    user=Depends(require_user),
):
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとしてログインしてください")
    update_work_session_from_form(
        session_id,
        company_id_for(user),
        user,
        user_id_guard=user["id"],
        start_time=start_time,
        end_time=end_time,
        vehicle_id=vehicle_id,
        start_odometer=start_odometer,
        end_odometer=end_odometer,
        alcohol_check_start=alcohol_check_start,
        alcohol_check_end=alcohol_check_end,
        call_check_start=call_check_start,
        call_check_end=call_check_end,
        delivery_count=delivery_count,
        transfer_count=transfer_count,
        night_count=night_count,
        pickup_count=pickup_count,
        large_count=large_count,
        inspection_image_path=inspection_image_path,
        notes=notes,
    )
    return RedirectResponse("/my-room?message=" + quote("勤務記録を修正しました"), status_code=303)


@app.post("/admin/work-sessions/{session_id}/update")
def admin_update_work_session(
    session_id: int,
    start_time: str = Form(""),
    end_time: str = Form(""),
    vehicle_id: int = Form(0),
    start_odometer: int = Form(0),
    end_odometer: int = Form(0),
    alcohol_check_start: str = Form(""),
    alcohol_check_end: str = Form(""),
    call_check_start: str = Form(""),
    call_check_end: str = Form(""),
    delivery_count: int = Form(0),
    transfer_count: int = Form(0),
    night_count: int = Form(0),
    pickup_count: int = Form(0),
    large_count: int = Form(0),
    inspection_image_path: str = Form(""),
    notes: str = Form(""),
    user=Depends(require_admin),
):
    require_feature(user, "safety")
    update_work_session_from_form(
        session_id,
        company_id_for(user),
        user,
        start_time=start_time,
        end_time=end_time,
        vehicle_id=vehicle_id,
        start_odometer=start_odometer,
        end_odometer=end_odometer,
        alcohol_check_start=alcohol_check_start,
        alcohol_check_end=alcohol_check_end,
        call_check_start=call_check_start,
        call_check_end=call_check_end,
        delivery_count=delivery_count,
        transfer_count=transfer_count,
        night_count=night_count,
        pickup_count=pickup_count,
        large_count=large_count,
        inspection_image_path=inspection_image_path,
        notes=notes,
    )
    return RedirectResponse("/admin/safety?message=" + quote("勤務セッションを修正しました"), status_code=303)


@app.get("/admin/rates", response_class=HTMLResponse)
def rates_page(request: Request, user=Depends(require_admin)):
    require_feature(user, "rewards")
    company_id = company_id_for(user)
    rows = query_all(
        """SELECT u.id, u.name, r.* FROM users u LEFT JOIN rates r ON r.user_id = u.id AND r.company_id=?
           WHERE u.role='member' AND u.active=1 AND u.company_id=? ORDER BY u.id LIMIT 200""",
        (company_id, company_id),
    )
    return render(request, "rates.html", {"rows": rows})


@app.post("/admin/rates")
def save_rates(
    user_id: int = Form(...),
    delivery_unit: int = Form(...),
    transfer_unit: int = Form(...),
    night_unit: int = Form(...),
    pickup_unit: int = Form(...),
    large_unit: int = Form(...),
    vehicle_rental_type: str = Form("none"),
    vehicle_daily_fee: int = Form(0),
    vehicle_monthly_fee: int = Form(0),
    user=Depends(require_admin),
):
    require_feature(user, "rewards")
    company_id = company_id_for(user)
    member = query_one("SELECT id FROM users WHERE id=? AND role='member' AND company_id=?", (user_id, company_id))
    if not member:
        raise HTTPException(status_code=404, detail="メンバーが見つかりません")
    before = query_one("SELECT * FROM rates WHERE user_id=? AND company_id=?", (user_id, company_id))
    execute(
        """INSERT INTO rates(user_id, company_id, delivery_unit, transfer_unit, night_unit, pickup_unit, large_unit,
           vehicle_rental_type, vehicle_daily_fee, vehicle_monthly_fee)
           VALUES (?,?,?,?,?,?,?,?,?,?)
           ON CONFLICT(user_id) DO UPDATE SET delivery_unit=excluded.delivery_unit,
           company_id=excluded.company_id,
           transfer_unit=excluded.transfer_unit, night_unit=excluded.night_unit,
           pickup_unit=excluded.pickup_unit, large_unit=excluded.large_unit,
           vehicle_rental_type=excluded.vehicle_rental_type,
           vehicle_daily_fee=excluded.vehicle_daily_fee,
           vehicle_monthly_fee=excluded.vehicle_monthly_fee""",
        (user_id, company_id, delivery_unit, transfer_unit, night_unit, pickup_unit, large_unit, vehicle_rental_type, vehicle_daily_fee, vehicle_monthly_fee),
    )
    after = query_one("SELECT * FROM rates WHERE user_id=? AND company_id=?", (user_id, company_id))
    audit_log(
        company_id,
        user,
        "payroll.rates_update",
        "rates",
        user_id,
        "Member reward rates updated",
        before=row_to_dict(before),
        after=row_to_dict(after),
    )
    return RedirectResponse("/admin/rates", status_code=303)


@app.get("/admin/vehicle-rates", response_class=HTMLResponse)
def vehicle_rates_page(request: Request, user=Depends(require_admin)):
    require_feature(user, "vehicle")
    rates = vehicle_rates_for_company(company_id_for(user))
    return render(request, "vehicle_rates.html", {"rates": rates})


@app.post("/admin/vehicle-rates")
def save_vehicle_rates(daily_fee: int = Form(...), monthly_fee: int = Form(...), user=Depends(require_admin)):
    require_feature(user, "vehicle")
    company_id = company_id_for(user)
    before = vehicle_rates_for_company(company_id)
    now = app_now().isoformat(timespec="seconds")
    execute(
        """INSERT INTO company_vehicle_rates(company_id, daily_fee, monthly_fee, updated_at)
           VALUES (?, ?, ?, ?)
           ON CONFLICT(company_id) DO UPDATE SET daily_fee=excluded.daily_fee,
           monthly_fee=excluded.monthly_fee, updated_at=excluded.updated_at""",
        (company_id, daily_fee, monthly_fee, now),
    )
    audit_log(
        company_id,
        user,
        "payroll.vehicle_rates_update",
        "company_vehicle_rates",
        company_id,
        "Company vehicle rates updated",
        before=row_to_dict(before),
        after={"daily_fee": daily_fee, "monthly_fee": monthly_fee},
    )
    return RedirectResponse("/admin/vehicle-rates", status_code=303)


@app.get("/admin/vehicles", response_class=HTMLResponse)
def admin_vehicles_page(request: Request, user=Depends(require_admin)):
    require_feature(user, "vehicle")
    company_id = company_id_for(user)
    target_month = request.query_params.get("month") or app_today().strftime("%Y-%m")
    month_start, month_end = vehicle_expense_month_bounds(target_month)
    recent_expenses = query_all(
        """SELECT e.*, v.name AS vehicle_name, v.plate_number, COALESCE(d.name, '') AS depot_name
           FROM vehicle_expenses e
           LEFT JOIN vehicles v ON v.id=e.vehicle_id AND v.company_id=e.company_id
           LEFT JOIN depots d ON d.id=e.depot_id AND d.company_id=e.company_id
           WHERE e.company_id=? AND e.expense_date BETWEEN ? AND ?
           ORDER BY e.expense_date DESC, e.id DESC LIMIT 50""",
        (company_id, month_start, month_end),
    )
    return render(
        request,
        "vehicles.html",
        {
            "vehicles": vehicle_dashboard_rows(company_id, target_month),
            "members": member_rows_for_company(company_id),
            "depots": depot_rows_for_company(company_id),
            "recent_expenses": recent_expenses,
            "vehicle_statuses": VEHICLE_STATUSES,
            "vehicle_expense_categories": VEHICLE_EXPENSE_CATEGORIES,
            "target_month": target_month,
            "is_admin_view": True,
        },
    )


@app.post("/admin/vehicles")
def save_vehicle(
    vehicle_id: str = Form(""),
    name: str = Form(...),
    plate_number: str = Form(""),
    last_odometer: int = Form(0),
    oil_change_odometer: int = Form(0),
    oil_change_date: str = Form(""),
    inspection_due: str = Form(""),
    tire_note: str = Form(""),
    active: Optional[str] = Form(None),
    user=Depends(require_admin),
):
    require_feature(user, "vehicle")
    company_id = company_id_for(user)
    now = app_now().isoformat(timespec="seconds")
    name = name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="車両名を入力してください")
    active_value = 1 if active == "1" else 0
    before = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=?", (vehicle_id, company_id)) if vehicle_id else query_one("SELECT * FROM vehicles WHERE company_id=? AND name=?", (company_id, name))
    if vehicle_id:
        execute(
            """UPDATE vehicles
               SET name=?, plate_number=?, last_odometer=?, oil_change_odometer=?, oil_change_date=?,
                   inspection_due=?, tire_note=?, active=?, updated_at=?
               WHERE id=? AND company_id=?""",
            (
                name,
                plate_number,
                max(last_odometer, 0),
                max(oil_change_odometer, 0),
                oil_change_date,
                inspection_due,
                tire_note,
                active_value,
                now,
                vehicle_id,
                company_id,
            ),
        )
    else:
        execute(
            """INSERT INTO vehicles(company_id, name, plate_number, last_odometer, oil_change_odometer,
               oil_change_date, inspection_due, tire_note, active, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)
               ON CONFLICT(company_id, name) DO UPDATE SET plate_number=excluded.plate_number,
               last_odometer=excluded.last_odometer, oil_change_odometer=excluded.oil_change_odometer,
               oil_change_date=excluded.oil_change_date, inspection_due=excluded.inspection_due,
               tire_note=excluded.tire_note, active=excluded.active, updated_at=excluded.updated_at""",
            (
                company_id,
                name,
                plate_number,
                max(last_odometer, 0),
                max(oil_change_odometer, 0),
                oil_change_date,
                inspection_due,
                tire_note,
                active_value,
                now,
                now,
            ),
        )
    after = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=?", (vehicle_id, company_id)) if vehicle_id else query_one("SELECT * FROM vehicles WHERE company_id=? AND name=?", (company_id, name))
    audit_log(
        company_id,
        user,
        "vehicle.update" if before else "vehicle.create",
        "vehicles",
        row_value(after, "id", vehicle_id),
        "Vehicle saved",
        before=row_to_dict(before),
        after=row_to_dict(after),
    )
    return RedirectResponse("/admin/vehicles", status_code=303)


@app.post("/admin/vehicles/detail")
def save_vehicle_detail(
    vehicle_id: str = Form(""),
    name: str = Form(...),
    plate_number: str = Form(""),
    model: str = Form(""),
    color: str = Form(""),
    depot_id: int = Form(0),
    primary_user_id: int = Form(0),
    vehicle_status: str = Form("使用中"),
    last_odometer: int = Form(0),
    oil_change_odometer: int = Form(0),
    oil_change_date: str = Form(""),
    oil_change_interval: int = Form(5000),
    oil_note: str = Form(""),
    inspection_due: str = Form(""),
    inspection_cost: int = Form(0),
    inspection_note: str = Form(""),
    tire_change_date: str = Form(""),
    tire_change_odometer: int = Form(0),
    tire_cost: int = Form(0),
    tire_status_note: str = Form(""),
    tire_note: str = Form(""),
    repair_cost: int = Form(0),
    parts_cost: int = Form(0),
    car_wash_cost: int = Form(0),
    insurance_fee: int = Form(0),
    parking_fee: int = Form(0),
    other_vehicle_cost: int = Form(0),
    admin_memo: str = Form(""),
    issue_memo: str = Form(""),
    active: Optional[str] = Form(None),
    user=Depends(require_admin),
):
    require_feature(user, "vehicle")
    company_id = company_id_for(user)
    name = name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="車両名を入力してください")
    if depot_id and not query_one("SELECT id FROM depots WHERE id=? AND company_id=?", (depot_id, company_id)):
        raise HTTPException(status_code=400, detail="拠点が正しくありません")
    if primary_user_id and not query_one("SELECT id FROM users WHERE id=? AND company_id=? AND role='member'", (primary_user_id, company_id)):
        raise HTTPException(status_code=400, detail="主な使用者が正しくありません")
    if vehicle_status not in VEHICLE_STATUSES:
        vehicle_status = "使用中"
    now = app_now().isoformat(timespec="seconds")
    active_value = 1 if active == "1" else 0
    before = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=?", (vehicle_id, company_id)) if vehicle_id else query_one("SELECT * FROM vehicles WHERE company_id=? AND name=?", (company_id, name))
    values = (
        name,
        plate_number.strip(),
        model.strip(),
        color.strip(),
        depot_id or None,
        primary_user_id or None,
        vehicle_status,
        max(last_odometer, 0),
        max(oil_change_odometer, 0),
        oil_change_date,
        max(oil_change_interval, 1),
        oil_note.strip(),
        inspection_due,
        max(inspection_cost, 0),
        inspection_note.strip(),
        tire_change_date,
        max(tire_change_odometer, 0),
        max(tire_cost, 0),
        tire_status_note.strip(),
        tire_note.strip(),
        max(repair_cost, 0),
        max(parts_cost, 0),
        max(car_wash_cost, 0),
        max(insurance_fee, 0),
        max(parking_fee, 0),
        max(other_vehicle_cost, 0),
        admin_memo.strip(),
        issue_memo.strip(),
        active_value,
        now,
    )
    with db() as conn:
        if vehicle_id:
            conn.execute(
                """UPDATE vehicles
                   SET name=?, plate_number=?, model=?, color=?, depot_id=?, primary_user_id=?, vehicle_status=?,
                       last_odometer=?, oil_change_odometer=?, oil_change_date=?, oil_change_interval=?, oil_note=?,
                       inspection_due=?, inspection_cost=?, inspection_note=?,
                       tire_change_date=?, tire_change_odometer=?, tire_cost=?, tire_status_note=?, tire_note=?,
                       repair_cost=?, parts_cost=?, car_wash_cost=?, insurance_fee=?, parking_fee=?, other_vehicle_cost=?,
                       admin_memo=?, issue_memo=?, active=?, updated_at=?
                   WHERE id=? AND company_id=?""",
                (*values, vehicle_id, company_id),
            )
            record_id = vehicle_id
        else:
            conn.execute(
                """INSERT INTO vehicles(company_id, name, plate_number, model, color, depot_id, primary_user_id, vehicle_status,
                   last_odometer, oil_change_odometer, oil_change_date, oil_change_interval, oil_note,
                   inspection_due, inspection_cost, inspection_note,
                   tire_change_date, tire_change_odometer, tire_cost, tire_status_note, tire_note,
                   repair_cost, parts_cost, car_wash_cost, insurance_fee, parking_fee, other_vehicle_cost,
                   admin_memo, issue_memo, active, created_at, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                   ON CONFLICT(company_id, name) DO UPDATE SET plate_number=excluded.plate_number,
                   model=excluded.model, color=excluded.color, depot_id=excluded.depot_id, primary_user_id=excluded.primary_user_id,
                   vehicle_status=excluded.vehicle_status, last_odometer=excluded.last_odometer,
                   oil_change_odometer=excluded.oil_change_odometer, oil_change_date=excluded.oil_change_date,
                   oil_change_interval=excluded.oil_change_interval, oil_note=excluded.oil_note,
                   inspection_due=excluded.inspection_due, inspection_cost=excluded.inspection_cost, inspection_note=excluded.inspection_note,
                   tire_change_date=excluded.tire_change_date, tire_change_odometer=excluded.tire_change_odometer,
                   tire_cost=excluded.tire_cost, tire_status_note=excluded.tire_status_note, tire_note=excluded.tire_note,
                   repair_cost=excluded.repair_cost, parts_cost=excluded.parts_cost, car_wash_cost=excluded.car_wash_cost,
                   insurance_fee=excluded.insurance_fee, parking_fee=excluded.parking_fee, other_vehicle_cost=excluded.other_vehicle_cost,
                   admin_memo=excluded.admin_memo, issue_memo=excluded.issue_memo, active=excluded.active, updated_at=excluded.updated_at""",
                (company_id, *values[:-1], now, now),
            )
            saved = conn.execute("SELECT id FROM vehicles WHERE company_id=? AND name=?", (company_id, name)).fetchone()
            record_id = row_value(saved, "id", "")
        after = conn.execute("SELECT * FROM vehicles WHERE id=? AND company_id=?", (record_id, company_id)).fetchone() if record_id else None
        audit_log(company_id, user, "vehicle.detail_update" if before else "vehicle.detail_create", "vehicles", record_id, "Vehicle detail saved", before=row_to_dict(before), after=row_to_dict(after), conn=conn)
        conn.commit()
    return RedirectResponse("/admin/vehicles", status_code=303)


@app.post("/admin/vehicle-expenses")
async def save_vehicle_expense(
    expense_date: str = Form(...),
    vehicle_id: int = Form(...),
    depot_id: int = Form(0),
    category: str = Form(...),
    amount: float = Form(...),
    payee: str = Form(""),
    memo: str = Form(""),
    receipt: Optional[UploadFile] = File(None),
    user=Depends(require_admin),
):
    require_feature(user, "vehicle")
    company_id = company_id_for(user)
    if not parse_iso_date(expense_date):
        raise HTTPException(status_code=400, detail="日付が正しくありません")
    vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=?", (vehicle_id, company_id))
    if not vehicle:
        raise HTTPException(status_code=404, detail="車両が見つかりません")
    if depot_id and not query_one("SELECT id FROM depots WHERE id=? AND company_id=?", (depot_id, company_id)):
        raise HTTPException(status_code=400, detail="拠点が正しくありません")
    if category not in VEHICLE_EXPENSE_CATEGORIES:
        category = "その他"
    receipt_path = await save_receipt_upload(receipt, company_id, "vehicle_expense")
    now = app_now().isoformat(timespec="seconds")
    with db() as conn:
        cur = conn.execute(
            """INSERT INTO vehicle_expenses(company_id, vehicle_id, depot_id, expense_date, category, amount, payee, memo, receipt_path, created_by, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (company_id, vehicle_id, depot_id or None, expense_date, category, max(float(amount), 0.0), payee.strip(), memo.strip(), receipt_path, user["id"], now),
        )
        record_id = cur.lastrowid
        after = conn.execute("SELECT * FROM vehicle_expenses WHERE id=? AND company_id=?", (record_id, company_id)).fetchone()
        audit_log(company_id, user, "vehicle.expense_create", "vehicle_expenses", record_id, "Vehicle expense saved", after=row_to_dict(after), conn=conn)
        conn.commit()
    return RedirectResponse("/admin/vehicles", status_code=303)


@app.get("/admin/depot-rates", response_class=HTMLResponse)
def depot_rates_page(request: Request, user=Depends(require_admin)):
    require_feature(user, "rewards")
    company_id = company_id_for(user)
    rows = query_all(
        """SELECT rs.*, dep.name AS depot_name
           FROM depot_rate_settings rs
           JOIN depots dep ON dep.id=rs.depot_id AND dep.company_id=rs.company_id
           WHERE rs.company_id=?
           ORDER BY dep.name, rs.effective_start DESC, rs.id DESC LIMIT 300""",
        (company_id,),
    )
    return render(
        request,
        "depot_rates.html",
        {
            "depots": depot_rows_for_company(company_id),
            "rows": rows,
        },
    )


@app.post("/admin/depot-rates")
def save_depot_rate(
    depot_id: int = Form(...),
    base_unit: float = Form(0),
    transfer_unit: float = Form(0),
    night_unit: float = Form(0),
    pickup_unit: float = Form(0),
    large_unit: float = Form(0),
    effective_start: str = Form(...),
    effective_end: str = Form(""),
    memo: str = Form(""),
    user=Depends(require_admin),
):
    require_feature(user, "rewards")
    company_id = company_id_for(user)
    if not query_one("SELECT id FROM depots WHERE id=? AND company_id=?", (depot_id, company_id)):
        raise HTTPException(status_code=404, detail="拠点が見つかりません")
    if not parse_iso_date(effective_start):
        raise HTTPException(status_code=400, detail="適用開始日が正しくありません")
    if effective_end and not parse_iso_date(effective_end):
        raise HTTPException(status_code=400, detail="適用終了日が正しくありません")
    now = app_now().isoformat(timespec="seconds")
    with db() as conn:
        cur = conn.execute(
            """INSERT INTO depot_rate_settings(company_id, depot_id, base_unit, transfer_unit, night_unit, pickup_unit, large_unit, effective_start, effective_end, memo, created_by, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (company_id, depot_id, max(float(base_unit), 0.0), max(float(transfer_unit), 0.0), max(float(night_unit), 0.0), max(float(pickup_unit), 0.0), max(float(large_unit), 0.0), effective_start, effective_end, memo.strip(), user["id"], now),
        )
        record_id = cur.lastrowid
        after = conn.execute("SELECT * FROM depot_rate_settings WHERE id=? AND company_id=?", (record_id, company_id)).fetchone()
        audit_log(company_id, user, "finance.depot_rate_create", "depot_rate_settings", record_id, "Depot revenue rate saved", after=row_to_dict(after), conn=conn)
        conn.commit()
    return RedirectResponse("/admin/depot-rates", status_code=303)


@app.get("/admin/finance", response_class=HTMLResponse)
def finance_dashboard_page(request: Request, user=Depends(require_admin)):
    require_feature(user, "rewards")
    company_id = company_id_for(user)
    target_month = request.query_params.get("month") or app_today().strftime("%Y-%m")
    dashboard = build_finance_dashboard(company_id, target_month)
    return render(
        request,
        "finance_dashboard.html",
        {
            "dashboard": dashboard,
            "depots": depot_rows_for_company(company_id),
            "company_expense_categories": COMPANY_EXPENSE_CATEGORIES,
            "target_month": dashboard["target_month"],
        },
    )


@app.post("/admin/company-expenses")
async def save_company_expense(
    expense_date: str = Form(...),
    depot_id: int = Form(0),
    category: str = Form(...),
    amount: float = Form(...),
    payee: str = Form(""),
    memo: str = Form(""),
    receipt: Optional[UploadFile] = File(None),
    user=Depends(require_admin),
):
    require_feature(user, "rewards")
    company_id = company_id_for(user)
    if not parse_iso_date(expense_date):
        raise HTTPException(status_code=400, detail="日付が正しくありません")
    if depot_id and not query_one("SELECT id FROM depots WHERE id=? AND company_id=?", (depot_id, company_id)):
        raise HTTPException(status_code=400, detail="拠点が正しくありません")
    if category not in COMPANY_EXPENSE_CATEGORIES:
        category = "その他"
    receipt_path = await save_receipt_upload(receipt, company_id, "company_expense")
    now = app_now().isoformat(timespec="seconds")
    with db() as conn:
        cur = conn.execute(
            """INSERT INTO company_expenses(company_id, depot_id, expense_date, category, amount, payee, memo, receipt_path, created_by, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (company_id, depot_id or None, expense_date, category, max(float(amount), 0.0), payee.strip(), memo.strip(), receipt_path, user["id"], now),
        )
        record_id = cur.lastrowid
        after = conn.execute("SELECT * FROM company_expenses WHERE id=? AND company_id=?", (record_id, company_id)).fetchone()
        audit_log(company_id, user, "finance.company_expense_create", "company_expenses", record_id, "Company expense saved", after=row_to_dict(after), conn=conn)
        conn.commit()
    return RedirectResponse("/admin/finance", status_code=303)


@app.get("/vehicles", response_class=HTMLResponse)
def member_vehicles_page(request: Request, user=Depends(require_user)):
    require_feature(user, "vehicle")
    company_id = company_id_for(user)
    vehicles = []
    for vehicle in vehicle_rows_for_company(company_id):
        item = dict(vehicle)
        item["alert"] = vehicle_alert(vehicle)
        vehicles.append(item)
    return render(request, "vehicles.html", {"vehicles": vehicles, "is_admin_view": user["role"] == "admin"})


@app.get("/mobile/vehicle", response_class=HTMLResponse)
def mobile_vehicle_change_page(request: Request, user=Depends(require_user)):
    require_feature(user, "vehicle")
    if user["role"] == "admin":
        return RedirectResponse("/admin/vehicles", status_code=303)
    vehicles, selected_vehicle = selected_vehicle_context(user)
    return render(request, "mobile_vehicle_change.html", {"vehicles": vehicles, "selected_vehicle": selected_vehicle})


@app.post("/mobile/vehicle")
def mobile_vehicle_change(vehicle_id: int = Form(...), user=Depends(require_user)):
    require_feature(user, "vehicle")
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとしてログインしてください")
    company_id = company_id_for(user)
    vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=? AND active=1", (vehicle_id, company_id))
    if not vehicle:
        raise HTTPException(status_code=404, detail="車両が見つかりません")
    before_user = row_to_dict(user)
    today_s = app_today().isoformat()
    with db() as conn:
        conn.execute("UPDATE users SET last_vehicle_id=?, vehicle=? WHERE id=? AND company_id=?", (vehicle["id"], vehicle["name"], user["id"], company_id))
        conn.execute(
            """UPDATE work_logs SET vehicle_id=?, vehicle_name=?
               WHERE id IN (
                   SELECT id FROM work_logs
                   WHERE user_id=? AND company_id=? AND work_date=? AND log_type='start' AND COALESCE(is_deleted,0)=0
                   ORDER BY logged_at DESC LIMIT 1
               )""",
            (vehicle["id"], vehicle_label(vehicle), user["id"], company_id, today_s),
        )
        audit_log(
            company_id,
            user,
            "vehicle.member_change",
            "users",
            user["id"],
            "Member selected vehicle",
            before=before_user,
            after={"last_vehicle_id": vehicle["id"], "vehicle": vehicle["name"]},
            conn=conn,
        )
        conn.commit()
    return RedirectResponse("/member", status_code=303)


@app.get("/member", response_class=HTMLResponse)
def member_home(request: Request, user=Depends(require_user)):
    if user["role"] == "admin":
        return RedirectResponse("/admin", status_code=303)
    now_dt = app_now()
    today_s = now_dt.date().isoformat()
    company_id = company_id_for(user)
    delivery = active_delivery_for_day(company_id, user["id"], today_s)
    rates = query_one("SELECT * FROM rates WHERE user_id=? AND company_id=?", (user["id"], company_id))
    vehicle_rates = vehicle_rates_for_company(company_id)
    reward = calc_reward(delivery, rates, vehicle_rates) if delivery else 0
    month_summary = monthly_reward_summary(user["id"], app_today(), company_id)
    shifts = upcoming_member_shifts(user["id"], today_s, company_id)
    today_shift = shifts[0] if shifts and shifts[0]["shift_date"] == today_s else None
    next_shifts = [item for item in shifts if item["shift_date"] != today_s][:2]
    comment = auto_comment(delivery, reward, bool(today_shift))
    work_state = today_work_state(user["id"], company_id, today_s)
    selected_vehicle = vehicle_for_user(user, company_id)
    selected_vehicle_alert = vehicle_alert(selected_vehicle)
    holiday_alert = holiday_deadline_alert(now_dt.date())
    last_odometer = latest_vehicle_odometer(company_id, row_value(selected_vehicle, "id")) or latest_odometer_for_user(user["id"], company_id)
    company = query_one("SELECT memo FROM companies WHERE id=?", (company_id,))
    company_notice = (row_value(company, "memo", "") or "").strip()
    return render(
        request,
        "member_home.html",
        {
            "delivery": delivery,
            "reward": reward,
            "comment": comment,
            "shift": today_shift,
            "month_summary": month_summary,
            "today_shift": today_shift,
            "next_shifts": next_shifts,
            "work_state": work_state,
            "last_odometer": last_odometer,
            "today": today_s,
            "today_label": f"{now_dt.year}年 {now_dt.month}月{now_dt.day}日",
            "current_time": now_dt.strftime("%H:%M"),
            "flow_message": home_flow_message(selected_vehicle_alert, holiday_alert),
            "vehicle_name": vehicle_label(selected_vehicle),
            "vehicle_alert": selected_vehicle_alert,
            "holiday_alert": holiday_alert,
            "oil_change_date": row_value(selected_vehicle, "oil_change_date", "") or row_value(user, "oil_change_date", "") or "未設定",
            "vehicle_inspection_due": row_value(selected_vehicle, "inspection_due", "") or row_value(user, "vehicle_inspection_due", "") or "未設定",
            "vehicles": [],
            "selected_vehicle": selected_vehicle,
            "company_notice": company_notice,
        },
    )


@app.get("/member/results", response_class=HTMLResponse)
def member_results_calendar_page(request: Request, month: Optional[str] = None, user=Depends(require_user)):
    require_feature(user, "deliveries")
    if user["role"] == "admin":
        return RedirectResponse("/admin", status_code=303)
    company_id = company_id_for(user)
    calendar_data = member_month_calendar(company_id, user["id"], month)
    return render(request, "member_results_calendar.html", {"calendar_data": calendar_data})


@app.get("/member/result-corrections", response_class=HTMLResponse)
def member_result_corrections_page(request: Request, user=Depends(require_user)):
    require_feature(user, "deliveries")
    if user["role"] == "admin":
        return RedirectResponse("/admin", status_code=303)
    company_id = company_id_for(user)
    corrections = query_all(
        """SELECT c.*, a.name AS actor_name FROM delivery_corrections c
           LEFT JOIN users a ON a.id=c.actor_id
           WHERE c.user_id=? AND c.company_id=?
           ORDER BY c.changed_at DESC LIMIT 100""",
        (user["id"], company_id),
    )
    deletions = query_all(
        """SELECT d.*, a.name AS actor_name FROM work_day_deletions d
           LEFT JOIN users a ON a.id=d.actor_id
           WHERE d.user_id=? AND d.company_id=?
           ORDER BY d.deleted_at DESC LIMIT 100""",
        (user["id"], company_id),
    )
    return render(request, "member_result_corrections.html", {"corrections": corrections, "deletions": deletions})


@app.get("/member/results/{work_date}", response_class=HTMLResponse)
def member_result_detail_page(request: Request, work_date: str, message: str = "", error: str = "", user=Depends(require_user)):
    require_feature(user, "deliveries")
    if user["role"] == "admin":
        return RedirectResponse("/admin", status_code=303)
    if not parse_iso_date(work_date):
        raise HTTPException(status_code=400, detail="日付が正しくありません")
    company_id = company_id_for(user)
    context = member_result_day_context(company_id, user["id"], work_date)
    context.update({"target": work_date, "message": message, "error": error})
    return render(request, "member_result_detail.html", context)


@app.post("/member/results/{work_date}/delivery")
def member_result_update_delivery(
    work_date: str,
    completed: int = Form(0),
    transfer: int = Form(0),
    night: int = Form(0),
    pickup: int = Form(0),
    large: int = Form(0),
    memo: str = Form(""),
    user=Depends(require_user),
):
    require_feature(user, "deliveries")
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとしてログインしてください")
    if not parse_iso_date(work_date):
        raise HTTPException(status_code=400, detail="日付が正しくありません")
    company_id = company_id_for(user)
    selected_vehicle = vehicle_for_user(user, company_id)
    upsert_delivery_counts(
        company_id,
        user["id"],
        work_date,
        completed,
        transfer,
        night,
        pickup,
        large,
        "none",
        memo,
        "",
        user["id"],
        "実績カレンダー修正",
        row_value(selected_vehicle, "id"),
        vehicle_label(selected_vehicle) if selected_vehicle else "",
    )
    return RedirectResponse(f"/member/results/{work_date}?message=" + quote("配達個数を更新しました"), status_code=303)


@app.post("/member/results/{work_date}/delete")
def member_result_delete_work_day(work_date: str, reason: str = Form(...), user=Depends(require_user)):
    require_feature(user, "deliveries")
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとしてログインしてください")
    if not parse_iso_date(work_date):
        raise HTTPException(status_code=400, detail="日付が正しくありません")
    logically_delete_work_day(company_id_for(user), user["id"], work_date, user, reason)
    month = work_date[:7]
    return RedirectResponse(f"/member/results?month={month}", status_code=303)


@app.get("/work-log", response_class=HTMLResponse)
def work_log_page(request: Request, type: str = "start", user=Depends(require_user)):
    require_feature(user, "safety")
    log_type = type if type in {"start", "end"} else "start"
    now = app_now()
    logs = query_all(
        "SELECT * FROM work_logs WHERE user_id=? AND company_id=? AND COALESCE(is_deleted,0)=0 ORDER BY logged_at DESC LIMIT 20",
        (user["id"], company_id_for(user)),
    )
    vehicles, selected_vehicle = selected_vehicle_context(user)
    return render(
        request,
        "work_log.html",
        {
            "logs": logs,
            "log_type": log_type,
            "work_date": now.date().isoformat(),
            "logged_at": now.strftime("%Y-%m-%dT%H:%M"),
            "vehicles": vehicles,
            "selected_vehicle": selected_vehicle,
        },
    )


@app.post("/work-log")
def save_work_log(
    work_date: str = Form(...),
    log_type: str = Form(...),
    logged_at: str = Form(...),
    vehicle_id: int = Form(0),
    odometer: int = Form(0),
    alcohol_result: str = Form(...),
    detector_used: str = Form(...),
    intoxicated: str = Form(...),
    health_status: str = Form(...),
    face_check: str = Form(...),
    breath_check: str = Form(...),
    voice_check: str = Form(...),
    admin_confirm: str = Form(""),
    notes: str = Form(""),
    user=Depends(require_user),
):
    require_feature(user, "safety")
    company_id = company_id_for(user)
    vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=? AND active=1", (vehicle_id, company_id)) if vehicle_id else None
    vehicle_name = vehicle_label(vehicle) if vehicle else ""
    now = app_now().isoformat(timespec="seconds")
    with db() as conn:
        start_odo, end_odo, distance_km = odometer_log_values(conn, company_id, user["id"], work_date, log_type, odometer)
        cur = conn.execute(
            """INSERT INTO work_logs(company_id, user_id, work_date, log_type, logged_at, vehicle_id, vehicle_name, odometer,
               start_odometer, end_odometer, distance_km, alcohol_result, detector_used, intoxicated, health_status,
               face_check, breath_check, voice_check, admin_confirm, notes, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (company_id, user["id"], work_date, log_type, logged_at, vehicle["id"] if vehicle else None, vehicle_name, max(odometer, 0), start_odo, end_odo, distance_km, alcohol_result, detector_used, intoxicated, health_status, face_check, breath_check, voice_check, admin_confirm, notes, now),
        )
        if vehicle:
            conn.execute("UPDATE users SET last_vehicle_id=?, vehicle=? WHERE id=? AND company_id=?", (vehicle["id"], vehicle["name"], user["id"], company_id))
        update_vehicle_odometer(conn, company_id, vehicle["id"] if vehicle else None, vehicle_name, odometer, now)
        audit_log(
            company_id,
            user,
            f"work_log.{log_type}",
            "work_logs",
            cur.lastrowid,
            "Work log saved",
            after={
                "work_date": work_date,
                "log_type": log_type,
                "logged_at": logged_at,
                "vehicle_id": vehicle["id"] if vehicle else None,
                "odometer": max(odometer, 0),
                "start_odometer": start_odo,
                "end_odometer": end_odo,
                "distance_km": distance_km,
            },
            conn=conn,
        )
        conn.commit()
    return RedirectResponse("/work-log", status_code=303)


@app.get("/mobile/work/start", response_class=HTMLResponse)
def mobile_work_start_page(request: Request, user=Depends(require_user)):
    require_feature(user, "safety")
    if user["role"] == "admin":
        return RedirectResponse("/admin", status_code=303)
    company_id = company_id_for(user)
    today_s = app_today().isoformat()
    state = today_work_state(user["id"], company_id, today_s)
    if state["status"] == "working":
        return RedirectResponse("/member", status_code=303)
    now_dt = app_now()
    vehicles, selected_vehicle = selected_vehicle_context(user)
    next_no = query_one(
        "SELECT COALESCE(MAX(session_no), 0) + 1 AS next_no FROM work_log_sessions WHERE company_id=? AND user_id=? AND work_date=?",
        (company_id, user["id"], today_s),
    )
    return render(
        request,
        "mobile_work_start.html",
        {
            "work_date": today_s,
            "logged_at": now_dt.strftime("%Y-%m-%dT%H:%M"),
            "last_odometer": latest_vehicle_odometer(company_id, row_value(selected_vehicle, "id")) or latest_odometer_for_user(user["id"], company_id),
            "vehicle_name": vehicle_label(selected_vehicle),
            "vehicles": vehicles,
            "selected_vehicle": selected_vehicle,
            "session_no": row_value(next_no, "next_no", 1),
            "is_restart": state["status"] == "finished",
        },
    )


@app.post("/mobile/work/start")
def mobile_work_start(
    work_date: str = Form(...),
    logged_at: str = Form(""),
    vehicle_id: int = Form(0),
    odometer: int = Form(0),
    alcohol_result: str = Form("0.00"),
    detector_used: str = Form("有"),
    health_status: str = Form("良好"),
    notes: str = Form(""),
    user=Depends(require_user),
):
    require_feature(user, "safety")
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとしてログインしてください")
    company_id = company_id_for(user)
    state = today_work_state(user["id"], company_id, work_date)
    if state["status"] == "working":
        return RedirectResponse("/member", status_code=303)
    vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=? AND active=1", (vehicle_id, company_id)) if vehicle_id else None
    vehicle_name = vehicle_label(vehicle) if vehicle else (row_value(user, "vehicle", "") or "")
    logged_at = logged_at or app_now().strftime("%Y-%m-%dT%H:%M")
    vehicle_note = f"車両: {vehicle_name or '未登録'}"
    notes = " / ".join(part for part in [vehicle_note, notes.strip()] if part)
    now = app_now().isoformat(timespec="seconds")
    with db() as conn:
        open_session = conn.execute(
            """SELECT id FROM work_log_sessions
               WHERE company_id=? AND user_id=? AND work_date=? AND status='working' AND COALESCE(is_deleted,0)=0
               ORDER BY id DESC LIMIT 1""",
            (company_id, user["id"], work_date),
        ).fetchone()
        if open_session:
            return RedirectResponse("/member", status_code=303)
        session_no = next_work_session_no(conn, company_id, user["id"], work_date)
        start_odo, end_odo, distance_km = odometer_log_values(conn, company_id, user["id"], work_date, "start", odometer)
        cur = conn.execute(
            """INSERT INTO work_logs(company_id, user_id, work_date, log_type, logged_at, vehicle_id, vehicle_name, odometer,
               start_odometer, end_odometer, distance_km, alcohol_result, detector_used,
               intoxicated, health_status, face_check, breath_check, voice_check, admin_confirm, notes, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                company_id,
                user["id"],
                work_date,
                "start",
                logged_at,
                vehicle["id"] if vehicle else None,
                vehicle_name,
                max(odometer, 0),
                start_odo,
                end_odo,
                distance_km,
                alcohol_result or "OK",
                detector_used,
                "無",
                health_status or "良好",
                "問題なし",
                "問題なし",
                "問題なし",
                "",
                notes,
                now,
            ),
        )
        session_cur = conn.execute(
            """INSERT INTO work_log_sessions(company_id, user_id, work_date, session_no, status, start_time,
               start_log_id, vehicle_id, vehicle_name, start_odometer, alcohol_check_start, call_check_start, notes, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                company_id,
                user["id"],
                work_date,
                session_no,
                "working",
                logged_at,
                cur.lastrowid,
                vehicle["id"] if vehicle else None,
                vehicle_name,
                start_odo,
                alcohol_result or "OK",
                health_status or "良好",
                notes,
                now,
                now,
            ),
        )
        if vehicle:
            conn.execute("UPDATE users SET last_vehicle_id=?, vehicle=? WHERE id=? AND company_id=?", (vehicle["id"], vehicle["name"], user["id"], company_id))
        update_vehicle_odometer(conn, company_id, vehicle["id"] if vehicle else None, vehicle_name, odometer, now)
        audit_log(
            company_id,
            user,
            "work_log.start",
            "work_logs",
            cur.lastrowid,
            "Mobile start work log saved",
            after={
                "work_date": work_date,
                "logged_at": logged_at,
                "vehicle_id": vehicle["id"] if vehicle else None,
                "odometer": max(odometer, 0),
                "start_odometer": start_odo,
                "end_odometer": end_odo,
                "distance_km": distance_km,
                "session_id": session_cur.lastrowid,
                "session_no": session_no,
            },
            conn=conn,
        )
        conn.commit()
    return RedirectResponse("/mobile/work/start/complete", status_code=303)


@app.get("/mobile/work/start/complete", response_class=HTMLResponse)
def mobile_work_start_complete(request: Request, user=Depends(require_user)):
    require_feature(user, "safety")
    if user["role"] == "admin":
        return RedirectResponse("/admin", status_code=303)
    return render(request, "mobile_work_start_complete.html", {})


@app.get("/mobile/work/end", response_class=HTMLResponse)
def mobile_work_end_page(request: Request, user=Depends(require_user)):
    require_feature(user, "safety")
    require_feature(user, "deliveries")
    if user["role"] == "admin":
        return RedirectResponse("/admin", status_code=303)
    company_id = company_id_for(user)
    today_s = app_today().isoformat()
    state = today_work_state(user["id"], company_id, today_s)
    if state["status"] == "not_started":
        return RedirectResponse("/mobile/work/start", status_code=303)
    if state["status"] == "finished":
        return RedirectResponse(f"/mobile/work/end/complete?day={today_s}", status_code=303)
    open_session = state.get("open_session")
    delivery = active_delivery_for_day(company_id, user["id"], today_s)
    sheet = query_one(
        """SELECT * FROM inspection_sheets
           WHERE user_id=? AND company_id=? AND (sheet_date=? OR COALESCE(delivery_date, '')=?)
           ORDER BY created_at DESC LIMIT 1""",
        (user["id"], company_id, today_s, today_s),
    )
    ocr_result = {"work_date": today_s, "completed": 0, "pickup": 0, "acceptance": 0, "received": 0, "raw_text": ""}
    if sheet:
        extracted = extract_inspection_fields(row_value(sheet, "ocr_text", "") or "", fallback_date=row_value(sheet, "delivery_date", "") or sheet["sheet_date"])
        ocr_result = {
            "work_date": row_value(sheet, "delivery_date", "") or extracted["work_date"] or today_s,
            "completed": row_value(sheet, "ocr_completed", 0) or extracted["completed"],
            "pickup": row_value(sheet, "ocr_pickup", 0) or extracted["pickup"],
            "acceptance": row_value(sheet, "ocr_acceptance", 0) or extracted["acceptance"],
            "received": row_value(sheet, "ocr_received", 0) or extracted["received"],
            "raw_text": (row_value(sheet, "ocr_text", "") or "").strip(),
        }
    if open_session:
        counts = {
            "completed": int_value(row_value(open_session, "delivery_count")),
            "transfer": int_value(row_value(open_session, "transfer_count")),
            "night": int_value(row_value(open_session, "night_count")),
            "pickup": int_value(row_value(open_session, "pickup_count")),
            "large": int_value(row_value(open_session, "large_count")),
        }
    else:
        counts = {
            "completed": delivery["completed"] if delivery else ocr_result["completed"],
            "transfer": delivery["transfer"] if delivery else 0,
            "night": delivery["night"] if delivery else 0,
            "pickup": delivery["pickup"] if delivery else ocr_result["pickup"],
            "large": delivery["large"] if delivery else 0,
        }
    now_dt = app_now()
    vehicles, selected_vehicle = selected_vehicle_context(user)
    if open_session and row_value(open_session, "vehicle_id"):
        selected_vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=?", (row_value(open_session, "vehicle_id"), company_id)) or selected_vehicle
    last_odometer = int_value(row_value(open_session, "start_odometer")) if open_session else 0
    if not last_odometer:
        last_odometer = latest_vehicle_odometer(company_id, row_value(selected_vehicle, "id")) or latest_odometer_for_user(user["id"], company_id)
    return render(
        request,
        "mobile_work_end.html",
        {
            "work_date": today_s,
            "logged_at": now_dt.strftime("%Y-%m-%dT%H:%M"),
            "last_odometer": last_odometer,
            "counts": counts,
            "ocr_result": ocr_result,
            "vehicles": vehicles,
            "selected_vehicle": selected_vehicle,
        },
    )


@app.post("/mobile/work/end/preview", response_class=HTMLResponse)
async def mobile_work_end_preview(
    request: Request,
    work_date: str = Form(...),
    logged_at: str = Form(""),
    vehicle_id: int = Form(0),
    odometer: int = Form(0),
    alcohol_result: str = Form("0.00"),
    detector_used: str = Form("有"),
    completed: int = Form(0),
    transfer: int = Form(0),
    night: int = Form(0),
    pickup: int = Form(0),
    large: int = Form(0),
    notes: str = Form(""),
    inspection_image: Optional[UploadFile] = File(None),
    user=Depends(require_user),
):
    require_feature(user, "safety")
    require_feature(user, "deliveries")
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとしてログインしてください")
    company_id = company_id_for(user)
    state = today_work_state(user["id"], company_id, work_date)
    if state["status"] == "finished":
        return RedirectResponse(f"/mobile/work/end/complete?day={work_date}", status_code=303)
    if state["status"] != "working":
        return RedirectResponse("/mobile/work/start", status_code=303)
    log_date = work_date
    image_path = ""
    original_filename = ""
    content_type = ""
    ocr_text = ""
    ocr_result = {"work_date": work_date, "completed": 0, "pickup": 0, "acceptance": 0, "received": 0, "raw_text": ""}
    if inspection_image and inspection_image.filename:
        if not inspection_image.content_type or not inspection_image.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="点検表は画像ファイルを選択してください")
        suffix = safe_upload_name(inspection_image.filename)
        if suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
            raise HTTPException(status_code=400, detail="点検表画像は jpg, jpeg, png, webp に対応しています")
        data = await inspection_image.read()
        if not data:
            raise HTTPException(status_code=400, detail="画像ファイルが空です")
        if len(data) > 15 * 1024 * 1024:
            data = data[: 15 * 1024 * 1024]
        now_dt = app_now()
        filename = f"user{user['id']}_{work_date.replace('-', '')}_{now_dt.strftime('%H%M%S%f')}{suffix}"
        image_path, disk_path = save_image_data_for_ocr(data, filename, inspection_image.content_type or "", "inspection_slips", SLIP_UPLOAD_DIR)
        original_filename = inspection_image.filename or ""
        content_type = inspection_image.content_type or ""
        ocr_text = extract_ocr_text(disk_path)
        extracted = extract_inspection_fields(ocr_text, fallback_date=work_date)
        ocr_result = {
            "work_date": extracted["work_date"] or work_date,
            "completed": extracted["completed"],
            "pickup": extracted["pickup"],
            "acceptance": extracted["acceptance"],
            "received": extracted["received"],
            "raw_text": ocr_text.strip(),
        }
        if completed == 0 and extracted["completed"]:
            completed = extracted["completed"]
        if pickup == 0 and extracted["pickup"]:
            pickup = extracted["pickup"]
        if ocr_result["work_date"]:
            work_date = ocr_result["work_date"]
    counts = {"completed": completed, "transfer": transfer, "night": night, "pickup": pickup, "large": large}
    return render(
        request,
        "mobile_work_end_confirm.html",
        {
            "work_date": work_date,
            "log_date": log_date,
            "logged_at": logged_at or app_now().strftime("%Y-%m-%dT%H:%M"),
            "vehicle_id": vehicle_id,
            "odometer": max(odometer, 0),
            "alcohol_result": alcohol_result or "OK",
            "detector_used": detector_used,
            "notes": notes,
            "counts": counts,
            "ocr_result": ocr_result,
            "image_path": image_path,
            "image_info": inspection_image_info(image_path),
            "original_filename": original_filename,
            "content_type": content_type,
        },
    )


@app.post("/mobile/work/end")
def mobile_work_end_save(
    work_date: str = Form(...),
    log_date: str = Form(""),
    logged_at: str = Form(""),
    vehicle_id: int = Form(0),
    odometer: int = Form(0),
    alcohol_result: str = Form("0.00"),
    detector_used: str = Form("有"),
    completed: int = Form(0),
    transfer: int = Form(0),
    night: int = Form(0),
    pickup: int = Form(0),
    large: int = Form(0),
    notes: str = Form(""),
    image_path: str = Form(""),
    original_filename: str = Form(""),
    content_type: str = Form(""),
    user=Depends(require_user),
):
    require_feature(user, "safety")
    require_feature(user, "deliveries")
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとしてログインしてください")
    company_id = company_id_for(user)
    log_work_date = log_date or app_today().isoformat()
    state = today_work_state(user["id"], company_id, log_work_date)
    if state["status"] != "working":
        return RedirectResponse(f"/mobile/work/end/complete?day={work_date}", status_code=303)
    open_session = state.get("open_session")
    if not vehicle_id and open_session and row_value(open_session, "vehicle_id"):
        vehicle_id = int_value(row_value(open_session, "vehicle_id"))
    vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=? AND active=1", (vehicle_id, company_id)) if vehicle_id else None
    vehicle_name = vehicle_label(vehicle) if vehicle else (row_value(open_session, "vehicle_name", "") if open_session else "")
    logged_at_value = logged_at or app_now().strftime("%Y-%m-%dT%H:%M")
    session_id = row_value(open_session, "id") if open_session else None
    now = app_now().isoformat(timespec="seconds")
    with db() as conn:
        start_odo, end_odo, distance_km = odometer_log_values(conn, company_id, user["id"], log_work_date, "end", odometer)
        if open_session:
            start_odo = int_value(row_value(open_session, "start_odometer")) or start_odo
            end_odo = max(odometer, 0)
            distance_km = max(end_odo - start_odo, 0) if start_odo else distance_km
        cur = conn.execute(
            """INSERT INTO work_logs(company_id, user_id, work_date, log_type, logged_at, vehicle_id, vehicle_name, odometer,
               start_odometer, end_odometer, distance_km, alcohol_result, detector_used,
               intoxicated, health_status, face_check, breath_check, voice_check, admin_confirm, notes, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                company_id,
                user["id"],
                log_work_date,
                "end",
                logged_at_value,
                vehicle["id"] if vehicle else None,
                vehicle_name,
                max(odometer, 0),
                start_odo,
                max(odometer, 0),
                distance_km,
                alcohol_result or "OK",
                detector_used,
                "無",
                "良好",
                "問題なし",
                "問題なし",
                "問題なし",
                "",
                notes,
                now,
            ),
        )
        if open_session:
            conn.execute(
                """UPDATE work_log_sessions
                   SET status='finished', end_time=?, end_log_id=?, vehicle_id=COALESCE(?, vehicle_id),
                       vehicle_name=COALESCE(NULLIF(?, ''), vehicle_name), end_odometer=?, distance_km=?,
                       alcohol_check_end=?, call_check_end=?, delivery_count=?, transfer_count=?,
                       night_count=?, pickup_count=?, large_count=?, inspection_image_path=COALESCE(NULLIF(?, ''), inspection_image_path),
                       notes=COALESCE(NULLIF(?, ''), notes), updated_at=?
                   WHERE id=? AND company_id=? AND user_id=? AND COALESCE(is_deleted,0)=0""",
                (
                    logged_at_value,
                    cur.lastrowid,
                    vehicle["id"] if vehicle else None,
                    vehicle_name,
                    max(odometer, 0),
                    distance_km,
                    alcohol_result or "OK",
                    detector_used or "実施",
                    max(completed, 0),
                    max(transfer, 0),
                    max(night, 0),
                    max(pickup, 0),
                    max(large, 0),
                    image_path,
                    notes,
                    now,
                    session_id,
                    company_id,
                    user["id"],
                ),
            )
        if vehicle:
            conn.execute("UPDATE users SET last_vehicle_id=?, vehicle=? WHERE id=? AND company_id=?", (vehicle["id"], vehicle["name"], user["id"], company_id))
        update_vehicle_odometer(conn, company_id, vehicle["id"] if vehicle else None, vehicle_name, odometer, now)
        audit_log(
            company_id,
            user,
            "work_log.end",
            "work_logs",
            cur.lastrowid,
            "Mobile end work log saved",
            after={
                "work_date": log_work_date,
                "logged_at": logged_at_value,
                "vehicle_id": vehicle["id"] if vehicle else None,
                "odometer": max(odometer, 0),
                "start_odometer": start_odo,
                "end_odometer": max(odometer, 0),
                "distance_km": distance_km,
                "session_id": session_id,
            },
            conn=conn,
        )
        conn.commit()
    if work_date == log_work_date and session_id:
        delivery = sync_delivery_from_sessions(company_id, user["id"], work_date, user["id"], "スマホ退勤フロー", vehicle["id"] if vehicle else None, vehicle_name, image_path)
    else:
        delivery = upsert_delivery_counts(company_id, user["id"], work_date, completed, transfer, night, pickup, large, "none", notes or "スマホ退勤フローから登録", image_path, user["id"], "スマホ退勤フロー", vehicle["id"] if vehicle else None, vehicle_name)
    slip_id = None
    if image_path and delivery:
        slip_id = save_inspection_slip_record(company_id, user["id"], delivery["id"], work_date, image_path, original_filename, content_type)
    if slip_id and session_id:
        execute("UPDATE work_log_sessions SET inspection_sheet_id=? WHERE id=? AND company_id=? AND user_id=? AND COALESCE(is_deleted,0)=0", (slip_id, session_id, company_id, user["id"]))
    return RedirectResponse(f"/mobile/work/end/complete?day={work_date}", status_code=303)


@app.get("/mobile/work/end/complete", response_class=HTMLResponse)
def mobile_work_complete(request: Request, day: Optional[str] = None, user=Depends(require_user)):
    require_feature(user, "deliveries")
    if user["role"] == "admin":
        return RedirectResponse("/admin", status_code=303)
    target = day or app_today().isoformat()
    company_id = company_id_for(user)
    delivery = active_delivery_for_day(company_id, user["id"], target)
    rates = query_one("SELECT * FROM rates WHERE user_id=? AND company_id=?", (user["id"], company_id))
    vehicle_rates = vehicle_rates_for_company(company_id)
    reward = calc_reward(delivery, rates, vehicle_rates) if delivery else 0
    return render(request, "mobile_work_complete.html", {"delivery": delivery, "reward": reward, "target": target})


@app.get("/deliveries", response_class=HTMLResponse)
def delivery_page(request: Request, day: Optional[str] = None, user=Depends(require_user)):
    require_feature(user, "deliveries")
    target = day or app_today().isoformat()
    if user["role"] == "admin":
        company_id = company_id_for(user)
        rows = query_all(
            """SELECT d.*, u.name, COALESCE(s.file_path, d.inspection_sheet_path) AS slip_path,
                      (SELECT COUNT(*) FROM delivery_corrections c WHERE c.delivery_id=d.id AND c.company_id=d.company_id) AS correction_count
               FROM deliveries d
               JOIN users u ON u.id=d.user_id
               LEFT JOIN inspection_slips s ON s.delivery_id=d.id AND s.company_id=d.company_id
               WHERE d.work_date=? AND d.company_id=? AND COALESCE(d.is_deleted,0)=0 ORDER BY u.name LIMIT 300""",
            (target, company_id),
        )
        rows = attach_image_info(rows, "slip_path")
        corrections = query_all(
            """SELECT c.*, u.name AS member_name, a.name AS actor_name
               FROM delivery_corrections c
               JOIN users u ON u.id=c.user_id
               JOIN users a ON a.id=c.actor_id
               WHERE c.company_id=? AND c.work_date=?
               ORDER BY c.changed_at DESC LIMIT 50""",
            (company_id, target),
        )
        return render(request, "admin_deliveries.html", {"rows": rows, "target": target, "corrections": corrections})
    company_id = company_id_for(user)
    delivery = active_delivery_for_day(company_id, user["id"], target)
    slip = query_one("SELECT * FROM inspection_slips WHERE user_id=? AND company_id=? AND slip_date=?", (user["id"], company_id, target))
    sheet = query_one("SELECT * FROM inspection_sheets WHERE user_id=? AND company_id=? AND sheet_date=?", (user["id"], company_id, target))
    rates = query_one("SELECT * FROM rates WHERE user_id=? AND company_id=?", (user["id"], company_id))
    vehicle_rates = vehicle_rates_for_company(company_id)
    reward = calc_reward(delivery, rates, vehicle_rates) if delivery else 0
    inspection_image_path = ""
    if slip:
        inspection_image_path = slip["file_path"]
    elif delivery and "inspection_sheet_path" in delivery.keys() and delivery["inspection_sheet_path"]:
        inspection_image_path = delivery["inspection_sheet_path"]
    elif sheet:
        inspection_image_path = sheet["file_path"]
    inspection_image = inspection_image_info(inspection_image_path)
    corrections = query_all(
        """SELECT c.*, a.name AS actor_name FROM delivery_corrections c
           JOIN users a ON a.id=c.actor_id
           WHERE c.user_id=? AND c.company_id=? AND c.work_date=?
           ORDER BY c.changed_at DESC LIMIT 20""",
        (user["id"], company_id, target),
    )
    return render(
        request,
        "deliveries.html",
        {
            "delivery": delivery,
            "slip": slip,
            "sheet": sheet,
            "inspection_image_path": inspection_image_path,
            "inspection_image": inspection_image,
            "target": target,
            "reward": reward,
            "comment": auto_comment(delivery, reward),
            "corrections": corrections,
        },
    )


@app.post("/deliveries")
async def save_delivery(
    work_date: str = Form(...),
    completed: int = Form(0),
    transfer: int = Form(0),
    night: int = Form(0),
    pickup: int = Form(0),
    large: int = Form(0),
    vehicle_rental: str = Form("none"),
    memo: str = Form(""),
    inspection_image: Optional[UploadFile] = File(None),
    user=Depends(require_user),
):
    require_feature(user, "deliveries")
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとしてログインしてください")
    now_dt = app_now()
    now = now_dt.isoformat(timespec="seconds")
    company_id = company_id_for(user)
    selected_vehicle = vehicle_for_user(user, company_id)
    delivery = upsert_delivery_counts(
        company_id,
        user["id"],
        work_date,
        completed,
        transfer,
        night,
        pickup,
        large,
        vehicle_rental,
        memo,
        "",
        user["id"],
        "配達入力",
        row_value(selected_vehicle, "id"),
        vehicle_label(selected_vehicle) if selected_vehicle else "",
    )
    if inspection_image and inspection_image.filename:
        if not inspection_image.content_type or not inspection_image.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="点検表は画像ファイルを選択してください")
        suffix = safe_upload_name(inspection_image.filename)
        if suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
            raise HTTPException(status_code=400, detail="点検表画像は jpg, jpeg, png, webp に対応しています")
        SLIP_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        data = await inspection_image.read()
        if data:
            max_size = 15 * 1024 * 1024
            if len(data) > max_size:
                data = data[:max_size]
            filename = f"user{user['id']}_{work_date.replace('-', '')}_{now_dt.strftime('%H%M%S%f')}{suffix}"
            public_path, _ = save_inspection_image(data, filename, inspection_image.content_type or "", "inspection_slips", SLIP_UPLOAD_DIR)
            save_inspection_slip_record(company_id, user["id"], delivery["id"], work_date, public_path, inspection_image.filename or filename, inspection_image.content_type or "")
    return RedirectResponse(f"/deliveries?day={work_date}", status_code=303)


@app.get("/admin/delivery-corrections", response_class=HTMLResponse)
def delivery_corrections_page(request: Request, day: Optional[str] = None, user=Depends(require_admin)):
    require_feature(user, "deliveries")
    company_id = company_id_for(user)
    target = day or app_today().isoformat()
    rows = query_all(
        """SELECT c.*, u.name AS member_name, a.name AS actor_name
           FROM delivery_corrections c
           JOIN users u ON u.id=c.user_id
           JOIN users a ON a.id=c.actor_id
           WHERE c.company_id=? AND c.work_date=?
           ORDER BY c.changed_at DESC LIMIT 200""",
        (company_id, target),
    )
    deletions = query_all(
        """SELECT d.*, u.name AS member_name, a.name AS actor_name
           FROM work_day_deletions d
           JOIN users u ON u.id=d.user_id AND u.company_id=d.company_id
           LEFT JOIN users a ON a.id=d.actor_id
           WHERE d.company_id=? AND d.work_date=?
           ORDER BY d.deleted_at DESC LIMIT 100""",
        (company_id, target),
    )
    return render(request, "delivery_corrections.html", {"rows": rows, "target": target, "deletions": deletions})


@app.get("/inspection-sheets", response_class=HTMLResponse)
def inspection_sheets_page(request: Request, day: Optional[str] = None, member_id: Optional[int] = None, user=Depends(require_user)):
    require_feature(user, "inspection")
    target = day or app_today().isoformat()
    if user["role"] == "admin":
        company_id = company_id_for(user)
        members = query_all("SELECT id, name FROM users WHERE role='member' AND active=1 AND COALESCE(purged,0)=0 AND company_id=? ORDER BY name LIMIT 200", (company_id,))
        params = [company_id, target, target]
        member_filter = ""
        if member_id:
            member_filter = " AND u.id=?"
            params.append(member_id)
        sheets = query_all(
            f"""SELECT s.*, u.name FROM inspection_sheets s
                JOIN users u ON u.id=s.user_id
                WHERE s.company_id=? AND (s.sheet_date=? OR COALESCE(s.delivery_date, '')=?) {member_filter}
                ORDER BY u.name LIMIT 200""",
            params,
        )
        sheets = attach_image_info(sheets)
        submitted_rows = query_all(
            f"""SELECT s.user_id FROM inspection_sheets s
                JOIN users u ON u.id=s.user_id
                WHERE s.company_id=? AND (s.sheet_date=? OR COALESCE(s.delivery_date, '')=?) {member_filter}""",
            params,
        )
        submitted_ids = {sheet["user_id"] for sheet in submitted_rows}
        missing_members = [member for member in members if member["id"] not in submitted_ids and (not member_id or member["id"] == member_id)]
        return render(request, "inspection_admin.html", {"target": target, "members": members, "member_id": member_id, "sheets": sheets, "missing_members": missing_members})
    sheets = attach_image_info(
        query_all(
            "SELECT * FROM inspection_sheets WHERE user_id=? AND company_id=? ORDER BY sheet_date DESC LIMIT 30",
            (user["id"], company_id_for(user)),
        )
    )
    return render(request, "inspection_member.html", {"target": target, "sheets": sheets})


@app.get("/admin/inspection-retention", response_class=HTMLResponse)
def inspection_retention_page(request: Request, user=Depends(require_admin)):
    require_feature(user, "inspection")
    company_id = company_id_for(user)
    today_s = app_today().isoformat()
    rows = query_all(
        """SELECT 'sheet' AS image_type, s.id, s.user_id, u.name, COALESCE(s.delivery_date, s.sheet_date) AS image_date,
                  s.file_path, s.storage_path, s.public_url, s.uploaded_at, s.retention_until
           FROM inspection_sheets s
           JOIN users u ON u.id=s.user_id AND u.company_id=s.company_id
           WHERE s.company_id=? AND COALESCE(s.retention_until, '')<>'' AND s.retention_until<?
           UNION ALL
           SELECT 'slip' AS image_type, s.id, s.user_id, u.name, s.slip_date AS image_date,
                  s.file_path, s.storage_path, s.public_url, s.uploaded_at, s.retention_until
           FROM inspection_slips s
           JOIN users u ON u.id=s.user_id AND u.company_id=s.company_id
           WHERE s.company_id=? AND COALESCE(s.retention_until, '')<>'' AND s.retention_until<?
           ORDER BY retention_until, uploaded_at
           LIMIT 200""",
        (company_id, today_s, company_id, today_s),
    )
    return render(request, "inspection_retention.html", {"rows": attach_image_info(rows), "today": today_s})


@app.post("/inspection-sheets")
async def upload_inspection_sheet(sheet_date: str = Form(...), image: UploadFile = File(...), user=Depends(require_user)):
    require_feature(user, "inspection")
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとしてアップロードしてください")
    if not image.content_type or not image.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="画像ファイルを選択してください")
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    now_dt = app_now()
    now = now_dt.isoformat(timespec="seconds")
    suffix = safe_upload_name(image.filename)
    filename = f"user{user['id']}_{sheet_date.replace('-', '')}_{now_dt.strftime('%H%M%S%f')}{suffix}"
    data = await image.read()
    if not data:
        raise HTTPException(status_code=400, detail="画像ファイルが空です")
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="画像サイズは10MB以内にしてください")
    public_path, disk_path = save_image_data_for_ocr(data, filename, image.content_type or "", "inspection_sheets", UPLOAD_DIR)
    storage_meta = inspection_storage_meta("inspection_sheets", filename, public_path, now_dt)
    ocr_text = extract_ocr_text(disk_path)
    extracted = extract_inspection_fields(ocr_text, fallback_date=sheet_date)
    delivery_date = extracted["work_date"] or sheet_date
    ocr_status = "OCR確認待ち" if ocr_text else "手入力待ち"
    company_id = company_id_for(user)
    with db() as conn:
        conn.execute(
            """INSERT INTO inspection_sheets(company_id, user_id, sheet_date, delivery_date, file_path, storage_path, public_url, signed_url,
               original_filename, content_type, ocr_status, ocr_text, ocr_completed, ocr_pickup, ocr_acceptance, ocr_received,
               uploaded_at, retention_until, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(user_id, sheet_date) DO UPDATE SET file_path=excluded.file_path,
               company_id=excluded.company_id, delivery_date=excluded.delivery_date, storage_path=excluded.storage_path,
               public_url=excluded.public_url, signed_url=excluded.signed_url,
               original_filename=excluded.original_filename, content_type=excluded.content_type,
               ocr_status=excluded.ocr_status, ocr_text=excluded.ocr_text,
               ocr_completed=excluded.ocr_completed, ocr_pickup=excluded.ocr_pickup,
               ocr_acceptance=excluded.ocr_acceptance, ocr_received=excluded.ocr_received,
               uploaded_at=excluded.uploaded_at, retention_until=excluded.retention_until,
               created_at=excluded.created_at""",
            (
                company_id,
                user["id"],
                sheet_date,
                delivery_date,
                public_path,
                storage_meta["storage_path"],
                storage_meta["public_url"],
                storage_meta["signed_url"],
                image.filename or "",
                image.content_type or "",
                ocr_status,
                ocr_text,
                extracted["completed"],
                extracted["pickup"],
                extracted["acceptance"],
                extracted["received"],
                storage_meta["uploaded_at"],
                storage_meta["retention_until"],
                now,
            ),
        )
        sheet = conn.execute("SELECT id FROM inspection_sheets WHERE user_id=? AND company_id=? AND sheet_date=?", (user["id"], company_id, sheet_date)).fetchone()
        audit_log(
            company_id,
            user,
            "inspection_sheet.upload",
            "inspection_sheets",
            sheet["id"],
            "Inspection sheet uploaded",
            after={"sheet_date": sheet_date, "delivery_date": delivery_date, "file_path": public_path},
            conn=conn,
        )
        conn.commit()
    return RedirectResponse(f"/inspection-sheets/{sheet['id']}/correct", status_code=303)


def inspection_sheet_for_user(sheet_id: int, user):
    if user["role"] == "admin":
        sheet = query_one(
            """SELECT s.*, u.name AS member_name FROM inspection_sheets s
               JOIN users u ON u.id=s.user_id
               WHERE s.id=? AND s.company_id=?""",
            (sheet_id, company_id_for(user)),
        )
    else:
        sheet = query_one(
            """SELECT s.*, u.name AS member_name FROM inspection_sheets s
               JOIN users u ON u.id=s.user_id
               WHERE s.id=? AND s.user_id=? AND s.company_id=?""",
            (sheet_id, user["id"], company_id_for(user)),
        )
    if not sheet:
        raise HTTPException(status_code=404, detail="点検表が見つかりません")
    return sheet


@app.get("/inspection-sheets/{sheet_id}/correct", response_class=HTMLResponse)
def inspection_sheet_correct_page(request: Request, sheet_id: int, user=Depends(require_user)):
    require_feature(user, "inspection")
    sheet = with_image_info(inspection_sheet_for_user(sheet_id, user))
    extracted_fields = extract_inspection_fields(row_value(sheet, "ocr_text", "") or "", fallback_date=row_value(sheet, "delivery_date", "") or sheet["sheet_date"])
    target_date = row_value(sheet, "delivery_date", "") or extracted_fields["work_date"] or sheet["sheet_date"]
    delivery = query_one("SELECT * FROM deliveries WHERE user_id=? AND company_id=? AND work_date=?", (sheet["user_id"], row_value(sheet, "company_id", company_id_for(user)), target_date))
    ocr_completed = row_value(sheet, "ocr_completed", 0) or extracted_fields["completed"]
    ocr_pickup = row_value(sheet, "ocr_pickup", 0) or extracted_fields["pickup"]
    counts = {
        "completed": ocr_completed or (delivery["completed"] if delivery else 0),
        "transfer": delivery["transfer"] if delivery else 0,
        "night": delivery["night"] if delivery else 0,
        "pickup": ocr_pickup or (delivery["pickup"] if delivery else 0),
        "large": delivery["large"] if delivery else 0,
    }
    ocr_result = {
        "work_date": target_date,
        "completed": ocr_completed,
        "pickup": ocr_pickup,
        "acceptance": row_value(sheet, "ocr_acceptance", 0) or extracted_fields["acceptance"],
        "received": row_value(sheet, "ocr_received", 0) or extracted_fields["received"],
        "raw_text": (row_value(sheet, "ocr_text", "") or "").strip(),
    }
    return render(
        request,
        "inspection_correct.html",
        {"sheet": sheet, "delivery": delivery, "counts": counts, "target_date": target_date, "ocr_result": ocr_result},
    )


@app.post("/inspection-sheets/{sheet_id}/correct")
def apply_inspection_sheet_counts(
    sheet_id: int,
    work_date: str = Form(...),
    completed: int = Form(0),
    transfer: int = Form(0),
    night: int = Form(0),
    pickup: int = Form(0),
    large: int = Form(0),
    user=Depends(require_user),
):
    require_feature(user, "inspection")
    sheet = inspection_sheet_for_user(sheet_id, user)
    company_id = int(row_value(sheet, "company_id", company_id_for(user)) or company_id_for(user))
    target_date = normalize_import_date(work_date)
    try:
        datetime.strptime(target_date, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="日付はYYYY-MM-DD形式で入力してください") from exc
    now = app_now().isoformat(timespec="seconds")
    existing = query_one("SELECT * FROM deliveries WHERE user_id=? AND company_id=? AND work_date=?", (sheet["user_id"], company_id, target_date))
    memo = existing["memo"] if existing and existing["memo"] else "点検表から反映"
    vehicle_rental = existing["vehicle_rental"] if existing else "none"
    values = {
        "completed": max(completed, 0),
        "transfer": max(transfer, 0),
        "night": max(night, 0),
        "pickup": max(pickup, 0),
        "large": max(large, 0),
    }
    upsert_delivery_counts(
        company_id,
        sheet["user_id"],
        target_date,
        values["completed"],
        values["transfer"],
        values["night"],
        values["pickup"],
        values["large"],
        vehicle_rental,
        memo,
        sheet["file_path"],
        user["id"],
        "点検表OCR補正",
    )
    with db() as conn:
        conn.execute(
            """UPDATE inspection_sheets
               SET delivery_date=?, ocr_status=?, ocr_completed=?, ocr_pickup=?, ocr_acceptance=?
               WHERE id=? AND company_id=?""",
            (
                target_date,
                "反映済み",
                values["completed"],
                values["pickup"],
                values["pickup"],
                sheet_id,
                company_id,
            ),
        )
        audit_log(
            company_id,
            user,
            "inspection_sheet.apply_counts",
            "inspection_sheets",
            sheet_id,
            "Inspection OCR counts applied",
            before=row_to_dict(sheet),
            after={"delivery_date": target_date, **values},
            conn=conn,
        )
        conn.commit()
    return RedirectResponse(f"/deliveries?day={target_date}", status_code=303)


def reward_target_for_user(user, member_id: Optional[int] = None):
    company_id = company_id_for(user)
    members = []
    viewing_user_id = user["id"]
    if user["role"] == "admin":
        members = query_all("SELECT id, name FROM users WHERE role='member' AND active=1 AND company_id=? ORDER BY name LIMIT 200", (company_id,))
        if member_id:
            selected = query_one("SELECT id FROM users WHERE id=? AND role='member' AND company_id=?", (member_id, company_id))
            viewing_user_id = selected["id"] if selected else (members[0]["id"] if members else user["id"])
        elif members:
            viewing_user_id = members[0]["id"]
    return company_id, viewing_user_id, members


def reward_daily_rows(company_id: int, viewing_user_id: int, start: date, end: date):
    rows = query_all(
        "SELECT * FROM deliveries WHERE user_id=? AND company_id=? AND work_date BETWEEN ? AND ? AND COALESCE(is_deleted,0)=0 ORDER BY work_date",
        (viewing_user_id, company_id, start.isoformat(), end.isoformat()),
    )
    rates = query_one("SELECT * FROM rates WHERE user_id=? AND company_id=?", (viewing_user_id, company_id))
    vehicle_rates = vehicle_rates_for_company(company_id)
    daily = []
    total = 0
    monthly_deducted = False
    rental_type = rates["vehicle_rental_type"] if rates and "vehicle_rental_type" in rates.keys() else "none"
    for row in rows:
        monthly = rental_type == "monthly" and not monthly_deducted
        reward = calc_reward(row, rates, vehicle_rates, monthly_mode=monthly)
        gross = calc_reward_gross(row, rates)
        monthly_deducted = monthly_deducted or monthly
        daily.append({"row": row, "gross": gross, "vehicle_deduction": gross - reward, "reward": reward})
        total += reward
    return daily, total


@app.get("/rewards", response_class=HTMLResponse)
def rewards_page(request: Request, ym: Optional[str] = None, member_id: Optional[int] = None, user=Depends(require_user)):
    require_feature(user, "rewards")
    start, end = month_bounds(ym)
    company_id, viewing_user_id, members = reward_target_for_user(user, member_id)
    daily, total = reward_daily_rows(company_id, viewing_user_id, start, end)
    return render(request, "rewards.html", {"daily": daily, "total": total, "ym": start.strftime("%Y-%m"), "members": members, "member_id": viewing_user_id})


@app.get("/rewards/statement.csv")
def rewards_statement_csv(ym: Optional[str] = None, member_id: Optional[int] = None, user=Depends(require_user)):
    require_feature(user, "rewards")
    start, end = month_bounds(ym)
    company_id, viewing_user_id, _ = reward_target_for_user(user, member_id)
    member = query_one("SELECT name FROM users WHERE id=? AND company_id=?", (viewing_user_id, company_id))
    daily, total = reward_daily_rows(company_id, viewing_user_id, start, end)
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(["対象月", start.strftime("%Y-%m")])
    writer.writerow(["メンバー", row_value(member, "name", "")])
    writer.writerow([])
    writer.writerow(["日付", "完了", "転送", "夜間", "集荷/引受", "大型", "総額", "車両代控除", "日次報酬"])
    for item in daily:
        row = item["row"]
        writer.writerow([row["work_date"], row["completed"], row["transfer"], row["night"], row["pickup"], row["large"], item["gross"], item["vehicle_deduction"], item["reward"]])
    writer.writerow([])
    writer.writerow(["月次報酬", "", "", "", "", "", "", "", total])
    content = "\ufeff" + buffer.getvalue()
    filename = f"reward_statement_{start.strftime('%Y%m')}_{viewing_user_id}.csv"
    return Response(content=content, media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/vehicle-issues", response_class=HTMLResponse)
def vehicle_issues_page(request: Request, user=Depends(require_user)):
    require_feature(user, "issues")
    ensure_vehicle_issue_schema()
    company_id = company_id_for(user)
    if user["role"] == "admin":
        issues = query_all("""SELECT v.*, u.name FROM vehicle_issues v JOIN users u ON u.id=v.user_id WHERE v.company_id=? ORDER BY v.created_at DESC LIMIT 100""", (company_id,))
        history_filter = "v.company_id=?"
        history_params = (company_id,)
    else:
        issues = query_all("SELECT * FROM vehicle_issues WHERE user_id=? AND company_id=? ORDER BY created_at DESC", (user["id"], company_id))
        history_filter = "v.company_id=? AND v.user_id=?"
        history_params = (company_id, user["id"])
    normalized_issues = []
    for issue in issues:
        item = dict(issue)
        item["status"] = normalize_issue_status(item.get("status", "未対応"))
        item["status_class"] = ISSUE_STATUS_CLASSES[item["status"]]
        normalized_issues.append(item)
    histories = query_all(
        """SELECT l.*, u.name AS changed_by_name FROM vehicle_issue_status_logs l
           JOIN users u ON u.id=l.changed_by
           JOIN vehicle_issues v ON v.id=l.issue_id
           WHERE """ + history_filter + " ORDER BY l.changed_at DESC LIMIT 200",
        history_params,
    )
    history_by_issue = {}
    for history in histories:
        item = dict(history)
        item["status"] = normalize_issue_status(item.get("status", "未対応"))
        history_by_issue.setdefault(item["issue_id"], []).append(item)
    return render(request, "vehicle_issues.html", {"issues": normalized_issues, "history_by_issue": history_by_issue, "statuses": ISSUE_STATUSES})


@app.post("/vehicle-issues")
def save_vehicle_issue(issue_date: str = Form(...), vehicle_name: str = Form(...), severity: str = Form(...), detail: str = Form(...), user=Depends(require_user)):
    require_feature(user, "issues")
    ensure_vehicle_issue_schema()
    company_id = company_id_for(user)
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとして入力してください")
    now = datetime.now().isoformat(timespec="seconds")
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO vehicle_issues(company_id, user_id, issue_date, vehicle_name, severity, detail, status, created_at) VALUES (?,?,?,?,?,?,?,?)",
            (company_id, user["id"], issue_date, vehicle_name, severity, detail, "未対応", now),
        )
        conn.execute(
            "INSERT INTO vehicle_issue_status_logs(company_id, issue_id, status, changed_by, changed_at) VALUES (?, ?, ?, ?, ?)",
            (company_id, cur.lastrowid, "未対応", user["id"], now),
        )
        audit_log(
            company_id,
            user,
            "vehicle_issue.create",
            "vehicle_issues",
            cur.lastrowid,
            "Vehicle issue submitted",
            after={"issue_date": issue_date, "vehicle_name": vehicle_name, "severity": severity},
            conn=conn,
        )
        conn.commit()
    return RedirectResponse("/vehicle-issues", status_code=303)


@app.post("/vehicle-issues/{issue_id}/status")
def update_vehicle_issue_status(issue_id: int, status: str = Form(...), user=Depends(require_admin)):
    require_feature(user, "issues")
    ensure_vehicle_issue_schema()
    if status not in ISSUE_STATUSES:
        raise HTTPException(status_code=400, detail="状態が正しくありません")
    company_id = company_id_for(user)
    issue = query_one("SELECT * FROM vehicle_issues WHERE id=? AND company_id=?", (issue_id, company_id))
    if not issue:
        raise HTTPException(status_code=404, detail="車両不具合が見つかりません")
    now = datetime.now().isoformat(timespec="seconds")
    with db() as conn:
        conn.execute("UPDATE vehicle_issues SET status=? WHERE id=? AND company_id=?", (status, issue_id, company_id))
        conn.execute(
            "INSERT INTO vehicle_issue_status_logs(company_id, issue_id, status, changed_by, changed_at) VALUES (?, ?, ?, ?, ?)",
            (company_id, issue_id, status, user["id"], now),
        )
        audit_log(
            company_id,
            user,
            "vehicle_issue.status_update",
            "vehicle_issues",
            issue_id,
            "Vehicle issue status changed",
            before=row_to_dict(issue),
            after={"status": status},
            conn=conn,
        )
        conn.commit()
    return RedirectResponse("/vehicle-issues", status_code=303)


@app.get("/holidays", response_class=HTMLResponse)
def holidays_page(request: Request, user=Depends(require_user)):
    require_feature(user, "holidays")
    ym = request.query_params.get("ym")
    start, end = month_bounds(ym)
    if user["role"] == "admin":
        company_id = company_id_for(user)
        rows = query_all(
            """SELECT h.*, u.name FROM holiday_requests h JOIN users u ON u.id=h.user_id
               WHERE h.company_id=? AND h.request_date BETWEEN ? AND ? ORDER BY h.request_date, u.name LIMIT 300""",
            (company_id, start.isoformat(), end.isoformat()),
        )
    else:
        rows = query_all(
            "SELECT * FROM holiday_requests WHERE user_id=? AND company_id=? AND request_date BETWEEN ? AND ? ORDER BY request_date",
            (user["id"], company_id_for(user), start.isoformat(), end.isoformat()),
        )
    holiday_days = simple_japanese_holidays(start.year)
    rows_by_date = {}
    for row in rows:
        rows_by_date.setdefault(row["request_date"], []).append(row)
    return render(
        request,
        "holidays.html",
        {
            "rows": rows,
            "ym": start.strftime("%Y-%m"),
            "days": calendar_days_for_month(start),
            "rows_by_date": rows_by_date,
            "holiday_dates": {d.isoformat() for d in holiday_days},
            "holiday_alert": holiday_deadline_alert(),
            "holiday_status": holiday_submission_status(company_id_for(user), start) if user["role"] == "admin" else None,
        },
    )


@app.post("/holidays")
def save_holiday(request_date: str = Form(...), reason: str = Form(""), user=Depends(require_user)):
    require_feature(user, "holidays")
    if user["role"] == "admin":
        raise HTTPException(status_code=403, detail="メンバーとして入力してください")
    execute(
        """INSERT INTO holiday_requests(company_id, user_id, request_date, reason, created_at) VALUES (?,?,?,?,?)
           ON CONFLICT(user_id, request_date) DO UPDATE SET reason=excluded.reason, company_id=excluded.company_id""",
        (company_id_for(user), user["id"], request_date, reason, datetime.now().isoformat(timespec="seconds")),
    )
    return RedirectResponse("/holidays", status_code=303)


@app.get("/admin/shifts")
def admin_shifts_link(user=Depends(require_admin)):
    require_feature(user, "shifts")
    return RedirectResponse("/shifts", status_code=303)


@app.get("/member/shifts")
def member_shifts_link(user=Depends(require_user)):
    require_feature(user, "shifts")
    if user["role"] == "admin":
        return RedirectResponse("/shifts", status_code=303)
    return RedirectResponse("/shifts", status_code=303)


@app.get("/shifts", response_class=HTMLResponse)
def shifts_page(request: Request, ym: Optional[str] = None, message: str = "", user=Depends(require_user)):
    require_feature(user, "shifts")
    start, end = month_bounds(ym)
    if user["role"] == "admin":
        company_id = company_id_for(user)
        members = query_all("SELECT id, name FROM users WHERE role='member' AND active=1 AND company_id=? ORDER BY name LIMIT 200", (company_id,))
        shifts = query_all(
            """SELECT s.*, u.name, d.name AS district_name, t.name AS town_name
               FROM shifts s
               JOIN users u ON u.id=s.user_id
               LEFT JOIN districts d ON d.id=s.district_id AND d.company_id=s.company_id
               LEFT JOIN towns t ON t.id=s.town_id AND t.company_id=s.company_id
               WHERE s.company_id=? AND s.shift_date BETWEEN ? AND ? ORDER BY s.shift_date, u.name""",
            (company_id, start.isoformat(), end.isoformat()),
        )
        shifts = attach_shift_areas(shifts)
        holidays = query_all("""SELECT h.*, u.name FROM holiday_requests h JOIN users u ON u.id=h.user_id WHERE h.company_id=? AND h.request_date BETWEEN ? AND ? ORDER BY h.request_date LIMIT 300""", (company_id, start.isoformat(), end.isoformat()))
        return render(
            request,
            "admin_shifts.html",
            {
                "members": members,
                "shifts": shifts,
                "holidays": holidays,
                "ym": start.strftime("%Y-%m"),
                "days": calendar_days_for_month(start),
                "towns": active_towns(company_id),
                "shifts_by_date": shifts_by_date(shifts),
                "message": message,
                "holiday_status": holiday_submission_status(company_id, start),
            },
        )
    shifts = query_all(
        """SELECT s.*, d.name AS district_name, t.name AS town_name
           FROM shifts s
           LEFT JOIN districts d ON d.id=s.district_id AND d.company_id=s.company_id
           LEFT JOIN towns t ON t.id=s.town_id AND t.company_id=s.company_id
           WHERE s.user_id=? AND s.company_id=? AND s.shift_date BETWEEN ? AND ? AND s.decided=1
           ORDER BY s.shift_date""",
        (user["id"], company_id_for(user), start.isoformat(), end.isoformat()),
    )
    shifts = attach_shift_areas(shifts)
    return render(request, "member_shifts.html", {"shifts": shifts, "ym": start.strftime("%Y-%m"), "days": calendar_days_for_month(start), "shifts_by_date": shifts_by_date(shifts)})


@app.get("/admin/shifts/import-template")
def download_shift_import_template(user=Depends(require_admin)):
    require_feature(user, "shifts")
    rows = [
        SHIFT_IMPORT_HEADERS,
        ["2026-05-05", "笹本", "出勤", "高松1・2・3", "1便", ""],
        ["2026-05-05", "関口", "出勤", "高松4・5", "2便", ""],
    ]
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerows(rows)
    content = "\ufeff" + buffer.getvalue()
    headers = {"Content-Disposition": 'attachment; filename="shift_import_template.csv"'}
    return Response(content=content, media_type="text/csv; charset=utf-8", headers=headers)


@app.post("/admin/shifts/import/preview", response_class=HTMLResponse)
async def preview_shift_import(request: Request, import_file: UploadFile = File(...), user=Depends(require_admin)):
    require_feature(user, "shifts")
    try:
        raw_rows = await load_shift_import_rows(import_file)
        preview = build_shift_import_preview(raw_rows, company_id_for(user))
    except HTTPException as exc:
        return render(
            request,
            "shift_import_confirm.html",
            {
                "rows": [],
                "payload": "[]",
                "has_errors": True,
                "has_conflicts": False,
                "error": exc.detail,
                "filename": import_file.filename or "",
            },
        )
    return render(
        request,
        "shift_import_confirm.html",
        {
            **preview,
            "error": "",
            "filename": import_file.filename or "",
        },
    )


@app.post("/admin/shifts/import/commit", response_class=HTMLResponse)
def commit_shift_import(request: Request, payload: str = Form(...), overwrite_conflicts: Optional[str] = Form(None), user=Depends(require_admin)):
    require_feature(user, "shifts")
    try:
        rows = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="取り込み内容を読み取れませんでした") from exc
    if not isinstance(rows, list) or not rows:
        return render(
            request,
            "shift_import_confirm.html",
            {
                "rows": [],
                "payload": "[]",
                "has_errors": True,
                "has_conflicts": False,
                "error": "登録できる行がありません",
                "filename": "",
            },
        )

    company_id = company_id_for(user)
    conflicts = current_shift_conflicts(rows, company_id)
    for row in rows:
        row["errors"] = []
        row["conflict"] = bool(row.get("member_id") and row.get("shift_date") and (int(row["member_id"]), row["shift_date"]) in conflicts)
    if conflicts and not overwrite_conflicts:
        return render(
            request,
            "shift_import_confirm.html",
            {
                "rows": rows,
                "payload": json.dumps([{key: value for key, value in row.items() if key != "errors"} for row in rows], ensure_ascii=False),
                "has_errors": False,
                "has_conflicts": True,
                "error": "既存シフトがあります。上書きする場合はチェックを入れて確定してください。",
                "filename": "",
            },
        )

    towns_by_id = {int(town["id"]): town for town in active_towns(company_id)}
    with db() as conn:
        for row in rows:
            status = row.get("status", "")
            if status not in SHIFT_ALLOWED_STATUSES:
                raise HTTPException(status_code=400, detail="出勤区分が正しくありません")
            member = conn.execute(
                "SELECT id FROM users WHERE id=? AND role='member' AND company_id=? AND active=1 AND COALESCE(purged,0)=0",
                (int(row["member_id"]), company_id),
            ).fetchone()
            if not member:
                raise HTTPException(status_code=403, detail="他社または無効なメンバーのシフトは登録できません")
            town_ids = [int(town_id) for town_id in row.get("town_ids", [])]
            town_rows = [towns_by_id[town_id] for town_id in town_ids if town_id in towns_by_id]
            if status != "休み" and len(town_rows) != len(set(town_ids)):
                raise HTTPException(status_code=400, detail="配達エリアが正しくありません")
            upsert_shift_with_towns(conn, int(row["member_id"]), row["shift_date"], status, town_rows, row.get("note", ""), company_id, user)
        conn.commit()
    ym = rows[0]["shift_date"][:7] if rows and rows[0].get("shift_date") else app_today().strftime("%Y-%m")
    message = quote(f"{len(rows)}件のシフトを取り込みました")
    return RedirectResponse(f"/shifts?ym={ym}&message={message}", status_code=303)


@app.post("/shifts")
def save_shift(
    member_id: int = Form(...),
    shift_date: str = Form(...),
    status: str = Form(...),
    town_ids: List[int] = Form(default=[]),
    user=Depends(require_admin),
):
    require_feature(user, "shifts")
    upsert_shift_record(member_id, shift_date, status, town_ids, company_id=company_id_for(user), actor=user)
    return RedirectResponse("/shifts", status_code=303)


@app.get("/admin/areas", response_class=HTMLResponse)
def areas_page(request: Request, user=Depends(require_admin)):
    require_feature(user, "multi_depot")
    company_id = company_id_for(user)
    depots = query_all("SELECT * FROM depots WHERE active=1 AND company_id=? ORDER BY name", (company_id,))
    districts = query_all(
        """SELECT d.*, p.name AS depot_name FROM districts d
           LEFT JOIN depots p ON p.id=d.depot_id AND p.company_id=d.company_id
           WHERE d.active=1 AND d.company_id=? ORDER BY p.name, d.name""",
        (company_id,),
    )
    towns = query_all(
        """SELECT t.*, d.name AS district_name, p.name AS depot_name FROM towns t
           JOIN districts d ON d.id=t.district_id AND d.company_id=t.company_id
           LEFT JOIN depots p ON p.id=d.depot_id AND p.company_id=d.company_id
           WHERE t.active=1 AND d.active=1 AND t.company_id=? ORDER BY p.name, d.name, t.name""",
        (company_id,),
    )
    return render(request, "areas.html", {"depots": depots, "districts": districts, "towns": towns})


@app.post("/admin/areas/depots")
def save_depot(depot_id: str = Form(""), name: str = Form(...), user=Depends(require_admin)):
    require_feature(user, "multi_depot")
    company_id = company_id_for(user)
    name = name.strip()
    if not name:
        return RedirectResponse("/admin/areas", status_code=303)
    now = datetime.now().isoformat(timespec="seconds")
    try:
        if depot_id:
            execute("UPDATE depots SET name=? WHERE id=? AND company_id=?", (name, depot_id, company_id))
        else:
            execute("INSERT INTO depots(company_id, name, created_at) VALUES (?, ?, ?)", (company_id, name, now))
    except DB_INTEGRITY_ERROR:
        raise HTTPException(status_code=400, detail="同じ拠点名が既に使われています")
    return RedirectResponse("/admin/areas", status_code=303)


@app.post("/admin/areas/depots/{depot_id}/delete")
def delete_depot(depot_id: int, user=Depends(require_admin)):
    require_feature(user, "multi_depot")
    company_id = company_id_for(user)
    execute("UPDATE depots SET active=0 WHERE id=? AND company_id=?", (depot_id, company_id))
    execute("UPDATE districts SET active=0 WHERE depot_id=? AND company_id=?", (depot_id, company_id))
    execute(
        """UPDATE towns SET active=0
           WHERE company_id=? AND district_id IN (SELECT id FROM districts WHERE depot_id=? AND company_id=?)""",
        (company_id, depot_id, company_id),
    )
    return RedirectResponse("/admin/areas", status_code=303)


@app.post("/admin/areas/districts")
def save_district(district_id: str = Form(""), depot_id: int = Form(...), name: str = Form(...), user=Depends(require_admin)):
    require_feature(user, "multi_depot")
    company_id = company_id_for(user)
    depot = query_one("SELECT id FROM depots WHERE id=? AND company_id=? AND active=1", (depot_id, company_id))
    if not depot:
        raise HTTPException(status_code=403, detail="他社の拠点は変更できません")
    now = datetime.now().isoformat(timespec="seconds")
    if district_id:
        execute("UPDATE districts SET depot_id=?, name=? WHERE id=? AND company_id=?", (depot_id, name, district_id, company_id))
    else:
        execute("INSERT OR IGNORE INTO districts(company_id, depot_id, name, created_at) VALUES (?, ?, ?, ?)", (company_id, depot_id, name, now))
    return RedirectResponse("/admin/areas", status_code=303)


@app.post("/admin/areas/districts/{district_id}/delete")
def delete_district(district_id: int, user=Depends(require_admin)):
    require_feature(user, "multi_depot")
    company_id = company_id_for(user)
    execute("UPDATE districts SET active=0 WHERE id=? AND company_id=?", (district_id, company_id))
    execute("UPDATE towns SET active=0 WHERE district_id=? AND company_id=?", (district_id, company_id))
    return RedirectResponse("/admin/areas", status_code=303)


@app.post("/admin/areas/towns")
def save_town(town_id: str = Form(""), district_id: int = Form(...), name: str = Form(...), user=Depends(require_admin)):
    require_feature(user, "multi_depot")
    company_id = company_id_for(user)
    district = query_one("SELECT id FROM districts WHERE id=? AND company_id=? AND active=1", (district_id, company_id))
    if not district:
        raise HTTPException(status_code=403, detail="他社の配達エリアは変更できません")
    now = datetime.now().isoformat(timespec="seconds")
    if town_id:
        execute("UPDATE towns SET district_id=?, name=? WHERE id=? AND company_id=?", (district_id, name, town_id, company_id))
    else:
        execute("INSERT OR IGNORE INTO towns(company_id, district_id, name, created_at) VALUES (?, ?, ?, ?)", (company_id, district_id, name, now))
    return RedirectResponse("/admin/areas", status_code=303)


@app.post("/admin/areas/towns/{town_id}/delete")
def delete_town(town_id: int, user=Depends(require_admin)):
    require_feature(user, "multi_depot")
    execute("UPDATE towns SET active=0 WHERE id=? AND company_id=?", (town_id, company_id_for(user)))
    return RedirectResponse("/admin/areas", status_code=303)


def build_daily_report_rows(company_id: int, start_date: str, end_date: str, member_id: int = 0, vehicle_id: int = 0):
    member_clause = ""
    vehicle_clause_logs = ""
    vehicle_clause_deliveries = ""
    vehicle_clause_sessions = ""
    log_params = [company_id, start_date, end_date]
    delivery_params = [company_id, start_date, end_date]
    session_params = [company_id, start_date, end_date]
    if member_id:
        member_clause = " AND u.id=?"
        log_params.append(member_id)
        delivery_params.append(member_id)
        session_params.append(member_id)
    if vehicle_id:
        vehicle_clause_logs = " AND COALESCE(w.vehicle_id, 0)=?"
        vehicle_clause_deliveries = " AND COALESCE(d.vehicle_id, 0)=?"
        vehicle_clause_sessions = " AND COALESCE(s.vehicle_id, 0)=?"
        log_params.append(vehicle_id)
        delivery_params.append(vehicle_id)
        session_params.append(vehicle_id)
    sessions = query_all(
        f"""SELECT s.*, u.name AS driver_name, u.vehicle AS user_vehicle, v.name AS vehicle_master_name, v.plate_number
            FROM work_log_sessions s
            JOIN users u ON u.id=s.user_id AND u.company_id=s.company_id
            LEFT JOIN vehicles v ON v.id=s.vehicle_id AND v.company_id=s.company_id
            WHERE s.company_id=? AND s.work_date BETWEEN ? AND ? AND COALESCE(s.is_deleted,0)=0 {member_clause} {vehicle_clause_sessions}
            ORDER BY s.work_date, u.name, s.session_no""",
        session_params,
    )
    logs = query_all(
        f"""SELECT w.*, u.name AS driver_name, u.vehicle AS user_vehicle, v.name AS vehicle_master_name, v.plate_number
            FROM work_logs w
            JOIN users u ON u.id=w.user_id AND u.company_id=w.company_id
            LEFT JOIN vehicles v ON v.id=w.vehicle_id AND v.company_id=w.company_id
            WHERE w.company_id=? AND w.work_date BETWEEN ? AND ? AND COALESCE(w.is_deleted,0)=0 {member_clause} {vehicle_clause_logs}
            ORDER BY w.work_date, u.name, w.logged_at""",
        log_params,
    )
    deliveries = query_all(
        f"""SELECT d.*, u.name AS driver_name, u.vehicle AS user_vehicle, v.name AS vehicle_master_name, v.plate_number
            FROM deliveries d
            JOIN users u ON u.id=d.user_id AND u.company_id=d.company_id
            LEFT JOIN vehicles v ON v.id=d.vehicle_id AND v.company_id=d.company_id
            WHERE d.company_id=? AND d.work_date BETWEEN ? AND ? AND COALESCE(d.is_deleted,0)=0 {member_clause} {vehicle_clause_deliveries}
            ORDER BY d.work_date, u.name""",
        delivery_params,
    )
    records = {}
    session_user_dates = {(session["user_id"], session["work_date"]) for session in sessions}
    for session in sessions:
        key = ("session", session["id"])
        vehicle_text = row_value(session, "vehicle_name") or row_value(session, "vehicle_master_name") or row_value(session, "user_vehicle", "")
        plate = row_value(session, "plate_number", "")
        records[key] = {
            "date": session["work_date"],
            "driver": session["driver_name"],
            "logs": {},
            "delivery": None,
            "session": session,
            "vehicle": f"{vehicle_text} {plate}".strip() if plate else vehicle_text,
        }
    for log in logs:
        if (log["user_id"], log["work_date"]) in session_user_dates:
            continue
        key = (log["user_id"], log["work_date"])
        record = records.setdefault(key, {"date": log["work_date"], "driver": log["driver_name"], "logs": {}, "delivery": None, "vehicle": ""})
        vehicle_text = row_value(log, "vehicle_name") or row_value(log, "vehicle_master_name") or row_value(log, "user_vehicle", "")
        plate = row_value(log, "plate_number", "")
        record["vehicle"] = record["vehicle"] or (f"{vehicle_text} {plate}".strip() if plate else vehicle_text)
        log_type = row_value(log, "log_type", "")
        if log_type == "start":
            current = record["logs"].get("start")
            if not current or str(log["logged_at"]) < str(current["logged_at"]):
                record["logs"]["start"] = log
        elif log_type == "end":
            current = record["logs"].get("end")
            if not current or str(log["logged_at"]) > str(current["logged_at"]):
                record["logs"]["end"] = log
    for delivery in deliveries:
        if (delivery["user_id"], delivery["work_date"]) in session_user_dates:
            continue
        key = (delivery["user_id"], delivery["work_date"])
        record = records.setdefault(key, {"date": delivery["work_date"], "driver": delivery["driver_name"], "logs": {}, "delivery": None, "vehicle": ""})
        vehicle_text = row_value(delivery, "vehicle_name") or row_value(delivery, "vehicle_master_name") or row_value(delivery, "user_vehicle", "")
        plate = row_value(delivery, "plate_number", "")
        record["vehicle"] = record["vehicle"] or (f"{vehicle_text} {plate}".strip() if plate else vehicle_text)
        record["delivery"] = delivery
    rows = []
    for _, record in sorted(records.items(), key=lambda item: (item[1]["date"], item[1]["driver"])):
        start_log = record["logs"].get("start")
        end_log = record["logs"].get("end")
        delivery = record.get("delivery")
        session = record.get("session")
        start_odo = int_value(row_value(session, "start_odometer") or row_value(end_log, "start_odometer") or row_value(start_log, "start_odometer") or row_value(start_log, "odometer"))
        end_odo = int_value(row_value(session, "end_odometer") or row_value(end_log, "end_odometer") or row_value(end_log, "odometer"))
        distance_km = int_value(row_value(session, "distance_km") or row_value(end_log, "distance_km"))
        if not distance_km and start_odo and end_odo:
            distance_km = max(end_odo - start_odo, 0)
        notes = []
        for log in (start_log, end_log):
            if row_value(log, "notes"):
                notes.append(row_value(log, "notes"))
        if row_value(session, "notes"):
            notes.append(row_value(session, "notes"))
        if row_value(delivery, "memo"):
            notes.append(row_value(delivery, "memo"))
        rows.append(
            {
                "date": record["date"],
                "driver": record["driver"],
                "vehicle": record["vehicle"],
                "start_time": (row_value(session, "start_time", "") or row_value(start_log, "logged_at", ""))[11:16] if (row_value(session, "start_time", "") or row_value(start_log, "logged_at", "")) else "",
                "end_time": (row_value(session, "end_time", "") or row_value(end_log, "logged_at", ""))[11:16] if (row_value(session, "end_time", "") or row_value(end_log, "logged_at", "")) else "",
                "roll_call": "実施" if session or start_log or end_log else "",
                "alcohol": " / ".join([value for value in [row_value(session, "alcohol_check_start", ""), row_value(session, "alcohol_check_end", "")] if value]) if session else " / ".join([row_value(log, "alcohol_result", "") for log in (start_log, end_log) if row_value(log, "alcohol_result", "")]),
                "start_odometer": start_odo or "",
                "end_odometer": end_odo or "",
                "distance_km": distance_km or "",
                "completed": int_value(row_value(session, "delivery_count")) if session else (int_value(row_value(delivery, "completed")) if delivery else 0),
                "transfer": int_value(row_value(session, "transfer_count")) if session else (int_value(row_value(delivery, "transfer")) if delivery else 0),
                "night": int_value(row_value(session, "night_count")) if session else (int_value(row_value(delivery, "night")) if delivery else 0),
                "pickup": int_value(row_value(session, "pickup_count")) if session else (int_value(row_value(delivery, "pickup")) if delivery else 0),
                "large": int_value(row_value(session, "large_count")) if session else (int_value(row_value(delivery, "large")) if delivery else 0),
                "rest_minutes": int_value(row_value(end_log, "rest_minutes") or row_value(start_log, "rest_minutes")),
                "waiting_minutes": int_value(row_value(end_log, "waiting_minutes") or row_value(start_log, "waiting_minutes")),
                "loading_minutes": int_value(row_value(end_log, "loading_minutes") or row_value(start_log, "loading_minutes")),
                "notes": " / ".join(notes),
            }
        )
    return rows


@app.post("/admin/work-logs/{log_id}/update")
def admin_update_work_log(
    log_id: int,
    logged_at: str = Form(""),
    vehicle_id: int = Form(0),
    odometer: int = Form(0),
    alcohol_result: str = Form(""),
    detector_used: str = Form(""),
    health_status: str = Form(""),
    notes: str = Form(""),
    user=Depends(require_admin),
):
    require_feature(user, "safety")
    company_id = company_id_for(user)
    before = query_one("SELECT * FROM work_logs WHERE id=? AND company_id=? AND COALESCE(is_deleted,0)=0", (log_id, company_id))
    if not before:
        raise HTTPException(status_code=404, detail="Work log not found")
    vehicle = query_one("SELECT * FROM vehicles WHERE id=? AND company_id=?", (vehicle_id, company_id)) if vehicle_id else None
    vehicle_name = vehicle_label(vehicle) if vehicle else row_value(before, "vehicle_name", "")
    work_date = row_value(before, "work_date", "")
    log_type = row_value(before, "log_type", "")
    odometer = max(int_value(odometer), 0)
    logged_at = logged_at or row_value(before, "logged_at", "")
    alcohol_result = alcohol_result or row_value(before, "alcohol_result", "")
    detector_used = detector_used or row_value(before, "detector_used", "")
    health_status = health_status or row_value(before, "health_status", "")
    with db() as conn:
        start_odo, end_odo, distance_km = odometer_log_values(conn, company_id, before["user_id"], work_date, log_type, odometer)
        conn.execute(
            """UPDATE work_logs
               SET logged_at=?, vehicle_id=?, vehicle_name=?, odometer=?, start_odometer=?, end_odometer=?,
                   distance_km=?, alcohol_result=?, detector_used=?, health_status=?, notes=?
               WHERE id=? AND company_id=?""",
            (logged_at, vehicle["id"] if vehicle else row_value(before, "vehicle_id"), vehicle_name, odometer, start_odo, end_odo, distance_km, alcohol_result, detector_used, health_status, notes, log_id, company_id),
        )
        if log_type == "start":
            end_log = conn.execute(
                """SELECT * FROM work_logs
                   WHERE company_id=? AND user_id=? AND work_date=? AND log_type='end'
                     AND COALESCE(is_deleted,0)=0
                   ORDER BY logged_at DESC LIMIT 1""",
                (company_id, before["user_id"], work_date),
            ).fetchone()
            if end_log:
                end_start, end_end, end_distance = odometer_log_values(conn, company_id, before["user_id"], work_date, "end", row_value(end_log, "odometer"))
                conn.execute(
                    "UPDATE work_logs SET start_odometer=?, end_odometer=?, distance_km=? WHERE id=? AND company_id=?",
                    (end_start, end_end, end_distance, end_log["id"], company_id),
                )
        if vehicle:
            conn.execute(
                "UPDATE users SET last_vehicle_id=?, vehicle=? WHERE id=? AND company_id=?",
                (vehicle["id"], vehicle["name"], before["user_id"], company_id),
            )
        update_vehicle_odometer(conn, company_id, vehicle["id"] if vehicle else row_value(before, "vehicle_id"), vehicle_name, odometer, app_now().isoformat(timespec="seconds"))
        after = conn.execute("SELECT * FROM work_logs WHERE id=? AND company_id=? AND COALESCE(is_deleted,0)=0", (log_id, company_id)).fetchone()
        audit_log(
            company_id,
            user,
            "work_log.update",
            "work_logs",
            log_id,
            "Work log corrected",
            before=row_to_dict(before),
            after=row_to_dict(after),
            conn=conn,
        )
        conn.commit()
    return RedirectResponse("/admin/safety", status_code=303)


@app.get("/admin/safety", response_class=HTMLResponse)
def admin_safety(request: Request, message: str = "", user=Depends(require_admin)):
    require_feature(user, "safety")
    company_id = company_id_for(user)
    logs = query_all("""SELECT w.*, u.name FROM work_logs w JOIN users u ON u.id=w.user_id WHERE w.company_id=? AND COALESCE(w.is_deleted,0)=0 ORDER BY w.logged_at DESC LIMIT 100""", (company_id,))
    members = query_all("SELECT id, name FROM users WHERE role='member' AND company_id=? ORDER BY active DESC, name LIMIT 200", (company_id,))
    vehicles = vehicle_rows_for_company(company_id, include_inactive=True)
    sessions = query_all(
        """SELECT s.*, u.name FROM work_log_sessions s
           JOIN users u ON u.id=s.user_id AND u.company_id=s.company_id
           WHERE s.company_id=? AND COALESCE(s.is_deleted,0)=0
           ORDER BY s.work_date DESC, s.session_no DESC, s.id DESC LIMIT 100""",
        (company_id,),
    )
    audit_rows = query_all(
        """SELECT a.*, u.name AS actor_name FROM audit_logs a
           LEFT JOIN users u ON u.id=a.actor_id
           WHERE a.company_id=? AND a.action IN (
             'work_log.update', 'delivery.update', 'shift.update', 'vehicle.update',
             'inspection_sheet.apply_counts', 'inspection_slip.upsert'
           )
           ORDER BY a.created_at DESC LIMIT 50""",
        (company_id,),
    )
    return render(request, "admin_safety.html", {"logs": logs, "members": members, "vehicles": vehicles, "sessions": sessions, "audit_rows": audit_rows, "message": message})


@app.get("/admin/safety/daily-report.pdf")
def admin_safety_daily_report_pdf(
    start_date: str = "",
    end_date: str = "",
    member_id: int = 0,
    vehicle_id: int = 0,
    user=Depends(require_admin),
):
    require_feature(user, "safety")
    company_id = company_id_for(user)
    today = app_today()
    start = parse_iso_date(start_date) or today.replace(day=1)
    end = parse_iso_date(end_date) or today
    if end < start:
        start, end = end, start
    if member_id and not query_one("SELECT id FROM users WHERE id=? AND company_id=? AND role='member'", (member_id, company_id)):
        raise HTTPException(status_code=404, detail="メンバーが見つかりません")
    if vehicle_id and not query_one("SELECT id FROM vehicles WHERE id=? AND company_id=?", (vehicle_id, company_id)):
        raise HTTPException(status_code=404, detail="車両が見つかりません")
    rows = build_daily_report_rows(company_id, start.isoformat(), end.isoformat(), member_id, vehicle_id)
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    except ImportError:
        raise HTTPException(status_code=500, detail="PDF出力には reportlab が必要です")
    buffer = io.BytesIO()
    font_name = "HeiseiKakuGo-W5"
    try:
        pdfmetrics.registerFont(UnicodeCIDFont(font_name))
    except Exception:
        font_name = "Helvetica"
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = font_name
        style.fontSize = 7
        style.leading = 9
    company = query_one("SELECT name FROM companies WHERE id=?", (company_id,))
    title = Paragraph(f"{row_value(company, 'name', APP_NAME)} 業務記録 日報 ({start.isoformat()} - {end.isoformat()})", styles["Title"])
    headers = ["日付", "運転者", "車両番号", "出勤", "退勤", "点呼", "アルコール", "開始km", "終了km", "走行km", "完了", "転送", "夜間", "集荷", "大型", "休憩", "待機", "荷役", "備考"]
    table_rows = [headers]
    for row in rows:
        table_rows.append(
            [
                row["date"],
                row["driver"],
                row["vehicle"],
                row["start_time"],
                row["end_time"],
                row["roll_call"],
                row["alcohol"],
                row["start_odometer"],
                row["end_odometer"],
                row["distance_km"],
                row["completed"],
                row["transfer"],
                row["night"],
                row["pickup"],
                row["large"],
                row["rest_minutes"],
                row["waiting_minutes"],
                row["loading_minutes"],
                Paragraph(row["notes"] or "", styles["BodyText"]),
            ]
        )
    if len(table_rows) == 1:
        table_rows.append(["対象データなし"] + [""] * (len(headers) - 1))
    widths = [38, 50, 62, 32, 32, 32, 58, 36, 36, 36, 26, 26, 26, 26, 26, 28, 28, 28, 126]
    table = Table(table_rows, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 6),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]
        )
    )
    doc.build([title, Spacer(1, 8), table])
    filename = f"daily_report_{start.isoformat()}_{end.isoformat()}.pdf"
    headers_out = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(buffer.getvalue(), media_type="application/pdf", headers=headers_out)


@app.get("/admin/safety/print", response_class=HTMLResponse)
def admin_safety_print(request: Request, ym: Optional[str] = None, member_id: Optional[int] = None, user=Depends(require_admin)):
    require_feature(user, "safety")
    start, end = month_bounds(ym)
    company_id = company_id_for(user)
    member = None
    if member_id:
        member = query_one("SELECT * FROM users WHERE id=? AND role='member' AND company_id=?", (member_id, company_id))
    if not member:
        member = query_one("SELECT * FROM users WHERE role='member' AND company_id=? ORDER BY active DESC, name LIMIT 1", (company_id,))
    if not member:
        return render(request, "safety_print.html", {"records": [], "ym": start.strftime("%Y-%m"), "period": f"{start.isoformat()} ～ {end.isoformat()}", "member": None})
    logs = query_all(
        """SELECT w.*, u.name FROM work_logs w JOIN users u ON u.id=w.user_id
           WHERE w.user_id=? AND w.company_id=? AND w.work_date BETWEEN ? AND ? ORDER BY w.work_date, w.log_type""",
        (member["id"], company_id, start.isoformat(), end.isoformat()),
    )
    issues = query_all(
        """SELECT v.*, u.name FROM vehicle_issues v JOIN users u ON u.id=v.user_id
           WHERE v.user_id=? AND v.company_id=? AND v.issue_date BETWEEN ? AND ? ORDER BY v.issue_date""",
        (member["id"], company_id, start.isoformat(), end.isoformat()),
    )
    issue_map = {}
    for issue in issues:
        issue_map.setdefault(issue["issue_date"], []).append(issue["detail"])
    log_map = {}
    for log in logs:
        log_map.setdefault(log["work_date"], {})[log["log_type"]] = log
    records = []
    for day in calendar_days_for_month(start):
        day_s = day.isoformat()
        start_log = log_map.get(day_s, {}).get("start")
        end_log = log_map.get(day_s, {}).get("end")
        day_issues = issue_map.get(day_s, [])
        notes = []
        for log in (start_log, end_log):
            if log and log["notes"]:
                notes.append(log["notes"])
        if day_issues:
            notes.append("車両不具合: " + " / ".join(day_issues))
        records.append(
            {
                "date": date_label(day),
                "vehicle_number": member["vehicle"],
                "vehicle_status": "不具合あり" if day_issues else "良好",
                "roll_call": "実施" if start_log or end_log else "",
                "alcohol": "/".join([log["alcohol_result"] for log in (start_log, end_log) if log]) or "",
                "start_time": start_log["logged_at"][11:16] if start_log and len(start_log["logged_at"]) >= 16 else "",
                "end_time": end_log["logged_at"][11:16] if end_log and len(end_log["logged_at"]) >= 16 else "",
                "start_point": "",
                "end_point": "",
                "distance": "",
                "via": "",
                "notes": " / ".join(notes),
            }
        )
    return render(
        request,
        "safety_print.html",
        {"records": records, "ym": start.strftime("%Y-%m"), "period": f"{start.isoformat()} ～ {end.isoformat()}", "member": member},
    )
